#!/usr/bin/env python3
"""Tests for the one-off innovative attendance merge helper."""

import shutil
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from scripts.merge_innovation_attendance import merge_workbooks


SOURCE_GLOB = "dochazka_inovativni_vyuka_ZS_SD_ve_vyuce*.xlsx"
OUTPUT_NAME = "dochazka-inovativni_vzdelavani_MIMO-VYUKU-dochazka_OPJAK_II_TK.xlsx"


class MergeInnovationAttendanceTests(unittest.TestCase):
    def setUp(self):
        self.source_dir = Path("tmp")
        self.source_files = sorted(
            path for path in self.source_dir.glob(SOURCE_GLOB) if not path.name.startswith("~$")
        )
        self.template_file = self.source_dir / OUTPUT_NAME
        if len(self.source_files) != 3 or not self.template_file.exists():
            self.skipTest("Manual attendance merge fixtures are not available in tmp/")

    def merge_in_temp_dir(self):
        temp_context = tempfile.TemporaryDirectory()
        temp_dir = Path(temp_context.name)
        for source_file in self.source_files:
            shutil.copy2(source_file, temp_dir / source_file.name)
        output_file = temp_dir / OUTPUT_NAME
        shutil.copy2(self.template_file, output_file)
        merge_workbooks(temp_dir, output_file)
        return temp_context, output_file

    def test_merge_preserves_source_student_hour_totals(self):
        expected_totals = []
        for source_file in self.source_files:
            workbook = load_workbook(source_file, data_only=True)
            sheet = workbook["zdroj-dochazka"]
            try:
                for row in range(12, sheet.max_row + 1):
                    name = sheet.cell(row=row, column=2).value
                    if name in (None, ""):
                        if row > 12:
                            break
                        continue
                    expected_totals.append((str(name).strip(), sheet.cell(row=row, column=1).value))
            finally:
                workbook.close()

        temp_context, output_file = self.merge_in_temp_dir()
        with temp_context:
            output_workbook = load_workbook(output_file, data_only=False)
            output_sheet = output_workbook["zdroj-dochazka"]
            try:
                actual_totals = [
                    (
                        str(output_sheet.cell(row=12 + index, column=2).value).strip(),
                        output_sheet.cell(row=12 + index, column=1).value,
                    )
                    for index in range(len(expected_totals))
                ]
            finally:
                output_workbook.close()

        self.assertEqual(expected_totals, actual_totals)

    def test_merge_fills_class_book_dates_times_and_teachers(self):
        temp_context, output_file = self.merge_in_temp_dir()
        with temp_context:
            workbook = load_workbook(output_file, data_only=False)
            class_book = workbook["Třídní kniha"]
            try:
                rows = [
                    (class_book.cell(row=row, column=2).value, class_book.cell(row=row, column=4).value)
                    for row in range(7, 35)
                    if class_book.cell(row=row, column=2).value not in (None, "")
                ]
            finally:
                workbook.close()

        self.assertEqual(28, len(rows))
        self.assertEqual(("18.2.2026 8:30-11:30", "Eva Zlotá"), rows[0])
        self.assertEqual(("25.6.2026 8:30-12:15", "Petra Stiller"), rows[-1])


if __name__ == "__main__":
    unittest.main()
