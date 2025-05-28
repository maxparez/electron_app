#!/usr/bin/env python3
"""
Inspect real templates
"""

from openpyxl import load_workbook
import sys

def inspect_excel(file_path):
    print(f"=== Inspecting: {file_path} ===")
    
    try:
        wb = load_workbook(file_path, read_only=True)
        print(f"Worksheets: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\n--- Sheet: {sheet_name} ---")
            
            # Check cells that might contain version markers
            important_cells = ['B1', 'B6', 'B7', 'C4', 'C6']
            for cell_ref in important_cells:
                cell = sheet[cell_ref]
                if cell.value is not None:
                    print(f"{cell_ref}='{cell.value}'")
            
            # Look for headers around row 10-15
            for row in range(10, 16):
                row_data = []
                for col in range(1, 11):  # A-J
                    cell = sheet.cell(row=row, column=col)
                    if cell.value is not None:
                        row_data.append(f"{cell.coordinate}='{cell.value}'")
                
                if row_data:
                    print(f"Row {row}: {', '.join(row_data)}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error reading file: {e}")

if __name__ == "__main__":
    files = [
        "legacy_code/inv/16_hodin_inovativniho_vzdelavani_sablona.xlsx",
        "legacy_code/inv/32_hodin_inovativniho_vzdelavani_sablona.xlsx", 
        "legacy_code/inv/source.xlsx"
    ]
    
    for file in files:
        inspect_excel(file)
        print("\n" + "="*50 + "\n")