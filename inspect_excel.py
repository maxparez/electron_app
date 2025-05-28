#!/usr/bin/env python3
"""
Inspect Excel file structure to understand the format
"""

from openpyxl import load_workbook
import sys

def inspect_excel(file_path):
    print(f"=== Inspecting: {file_path} ===")
    
    try:
        wb = load_workbook(file_path, read_only=True)
        print(f"Worksheets: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\n--- Sheet: {sheet_name} ---")
            
            # Check first 15 rows and columns A-J for content
            for row in range(1, 16):
                row_data = []
                for col in range(1, 11):  # A-J
                    cell = sheet.cell(row=row, column=col)
                    if cell.value is not None:
                        row_data.append(f"{cell.coordinate}='{cell.value}'")
                
                if row_data:
                    print(f"Row {row}: {', '.join(row_data)}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error reading file: {e}")

if __name__ == "__main__":
    files = [
        "template_16_hodin.xlsx",
        "legacy_code/inv/16_hodin_inovativniho_vzdelavani_ZŠ_2x.xlsx"
    ]
    
    for file in files:
        inspect_excel(file)
        print("\n" + "="*50 + "\n")