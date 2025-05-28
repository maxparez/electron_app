#!/usr/bin/env python3
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src', 'python'))

import pandas as pd
from openpyxl import load_workbook

def debug_excel_file(filepath):
    """Debug Excel file structure"""
    print(f"=== Debugging {os.path.basename(filepath)} ===")
    
    try:
        # Check if file exists
        if not os.path.exists(filepath):
            print(f"File not found: {filepath}")
            return
            
        # Load with openpyxl to check sheets
        wb = load_workbook(filepath)
        print(f"Sheets: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\nSheet '{sheet_name}':")
            
            # Check key cells
            cells_to_check = [
                ("A1", sheet["A1"].value),
                ("B1", sheet["B1"].value), 
                ("B2", sheet["B2"].value),
                ("B6", sheet["B6"].value),
                ("B7", sheet["B7"].value)
            ]
            
            for cell_ref, value in cells_to_check:
                print(f"  {cell_ref}: {value}")
                
            # Check if this looks like a source data sheet
            if sheet_name == "Seznam aktivit":
                print(f"  Found 'Seznam aktivit' sheet")
                # Try to read with pandas
                try:
                    df = pd.read_excel(filepath, sheet_name=sheet_name)
                    print(f"  Pandas columns: {list(df.columns)}")
                    print(f"  Pandas shape: {df.shape}")
                    if len(df) > 0:
                        print(f"  First row data: {df.iloc[0].to_dict()}")
                except Exception as e:
                    print(f"  Pandas read error: {e}")
                    
        wb.close()
        
    except Exception as e:
        print(f"Error reading file: {e}")

if __name__ == "__main__":
    files_to_debug = [
        "/root/vyvoj_sw/electron_app/legacy_code/inv/source.xlsx",
        "/root/vyvoj_sw/electron_app/legacy_code/inv/16_hodin_inovativniho_vzdelavani_ZŠ_2x.xlsx",
        "/root/vyvoj_sw/electron_app/legacy_code/inv/16_hodin_inovativniho_vzdelavani_sablona.xlsx"
    ]
    
    for filepath in files_to_debug:
        debug_excel_file(filepath)
        print("\n" + "="*60 + "\n")