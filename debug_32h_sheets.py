#!/usr/bin/env python3
"""
Debug script to check sheets in 32h files
"""

import openpyxl

file_path = "tests/regression/inputs/32h/valid_basic.xlsx"

wb = openpyxl.load_workbook(file_path)
print(f"Available sheets in {file_path}:")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")
    
# Try first sheet
sheet = wb[wb.sheetnames[0]]
print(f"\nUsing sheet: {sheet.title}")

# Print relevant rows with 1-based indices
for row in range(1, 15):
    print(f"\nRow {row}:")
    for col in range(1, 5):  # Columns A-D
        cell_value = sheet.cell(row=row, column=col).value
        col_letter = openpyxl.utils.get_column_letter(col)
        if cell_value is not None:
            str_val = str(cell_value)[:40] + "..." if len(str(cell_value)) > 40 else str(cell_value)
            print(f"  {col_letter}: {repr(str_val)}")
wb.close()