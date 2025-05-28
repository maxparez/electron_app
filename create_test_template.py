#!/usr/bin/env python3
"""
Create test templates for 16 and 32 hour courses
"""

from openpyxl import Workbook
import os

def create_16_hour_template():
    """Create 16 hour template"""
    wb = Workbook()
    ws = wb.active
    ws.title = "zdroj-dochazka"
    
    # Template markers
    ws['B1'] = "16 hodin"
    ws['B6'] = "datum aktivity"
    ws['B7'] = "čas zahájení"
    
    # Headers for data (row 11, starting from column B)
    headers = ["datum", "cas", "forma", "tema", "ucitel", "hodin"]
    for i, header in enumerate(headers, start=2):  # Column B = 2
        ws.cell(row=11, column=i, value=header)
    
    # Save template
    template_path = "template_16_hodin.xlsx"
    wb.save(template_path)
    print(f"Created: {template_path}")
    return template_path

def create_32_hour_template():
    """Create 32 hour template"""
    wb = Workbook()
    ws = wb.active
    ws.title = "List1"
    
    # Template markers
    ws['B1'] = "32 hodin"
    ws['B6'] = "datum aktivity"
    
    # Headers for data (row 10, starting from column B)  
    headers = ["datum", "forma", "tema", "ucitel", "hodin"]
    for i, header in enumerate(headers, start=2):  # Column B = 2
        ws.cell(row=10, column=i, value=header)
    
    # Save template
    template_path = "template_32_hodin.xlsx"
    wb.save(template_path)
    print(f"Created: {template_path}")
    return template_path

if __name__ == "__main__":
    print("Creating test templates...")
    create_16_hour_template()
    create_32_hour_template()
    print("Templates created successfully!")