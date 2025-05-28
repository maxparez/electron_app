from genericpath import isfile
import pandas as pd
import glob
from openpyxl import load_workbook
from pretty_html_table import build_table
#import argparse
import os
import warnings
from pathlib import Path
import PySimpleGUI as sg

version = "0.16"

        
date_ranges = [
    ('2022-2023', '2022-09-01', '2023-08-31'),
    ('2023-2024', '2023-09-01', '2024-08-31'),
    ('2024-2025', '2024-09-01', '2025-08-31'),
    ('2025-2026', '2025-09-01', '2026-08-31')
]

tema_order = [
        "čtenářská pre/gramotnost",
        "matematická pre/gramotnost",
        "umělecká gramotnost",
        "mediální gramotnost",
        "cizí jazyky/komunikace v cizím jazyce",
        "inkluze včetně primární prevence",
        "přírodovědné a technické vzdělávání",
        "evvo a vzdělávání pro udržitelný rozvoj",
        "vzdělávání s využitím nových technologií",
        "kulturní povědomí a vyjádření",
        "historické povědomí, výuka moderních dějin",
        "rozvoj podnikavosti a kreativity",
        "well-being a psychohygiena",
        "genderová tematika v obsahu vzdělávání",
        "kariérové poradenství včetně identifikace a rozvoje nadání",
        "občanské vzdělávání a demokratické myšlení"
    ]

def is_valid_path(filepath):
    if filepath and Path(filepath).exists():
        return True
    sg.popup_error("Filepath not correct")
    return False

def main_window():
    # ------ Menu Definition ------ #
    menu_def =[["Help",["About"]]]


    # ------ GUI Definition ------ #
    layout = [[sg.MenubarCustom(menu_def, tearoff=False)],
              [sg.T("Source Folder:", s=15, justification="r"), sg.I(key="-OUT-"), sg.FolderBrowse(initial_folder=os.getcwd())],
              [sg.T("Name List:", s=15, justification="r"), sg.I(key="-LST-", disabled=True, disabled_readonly_background_color="Gray"), sg.FileBrowse(initial_folder=os.getcwd(),disabled=True, file_types=(("TXT Files", "*.txt"),))],
              [sg.Exit(s=16, button_color="tomato"), sg.B("Run", s=16)],]

    window_title = "ZoR SDP Calculator"
    window = sg.Window(window_title, layout, use_custom_titlebar=True)


    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, "Exit"):
            break
        if event == "About":
            window.disappear()
            sg.popup(window_title, "Version "+version, "Specifické datové položky pro ZoR", grab_anywhere=True)
            window.reappear()
        if event == "Run":
            if (is_valid_path(values["-OUT-"])):
                generate_report(values["-OUT-"], values["-LST-"])

    window.close()



class HtmlOutput():
    def __init__(self,**kwargs) -> None:
        self.html_body=""
        self.html = """
        <!DOCTYPE html>
        <html lang="cs">
        <head>
            <meta charset="windows-1250 ">
            <meta http-equiv="X-UA-Compatible" content="IE=edge">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Specifické datové položky - report</title>
            <style type="text/css">
                h1,h2{{color:navy;
                     font-family: Century Gothic, sans-serif; }}

            </style>
        </head>
        <body>
            {}
        </body>
        </html>
        """

    def add_html(self,html, before=False):
        self.html_body = html+self.html_body if before else self.html_body+html

    def add_h(self, level, text, before=False):
        h_template = "<h{0}>{1}</h{0}>"
        if before:
           self.html_body = h_template.format(level,text) + self.html_body
        else:
            self.html_body += h_template.format(level,text)

    @property
    def html_output(self):
        return self.html.format(self.html_body)

class SPC:
    def __init__(self,source_dir:str, name_list:str, **kwargs) -> None:
        self.source_dir=source_dir
        self.name_list_source = name_list
        self.sheet_name = kwargs.get('sheet_name', "Přehled")
        self.cols_names = ["ca","jmena","datum","pocet_hodin","forma","tema"]
        self.cols_agg = ["forma/téma","čislo","cena","typ"]
        self.result_cols_names = ["forma/téma","číslo celkem","cena celkem","typ"]
        self.html_report = HtmlOutput()
        #self.agg_function = [ {"jmena":"nunique"},{"pocet_hodin":"sum"}]
        self.agg_function = [ {"hash_jmena":"nunique"},{"pocet_hodin":"sum"}]
        self.source_file = kwargs.get('source_file')
        self.name_list =[]

    #check file is ok
    def is_expected_file_format(self, file_name):
        if os.path.isfile(file_name) and self.sheet_exists(file_name):
            return [file_name]
        return []

    # get only relavant files
    def get_files(self):
        # exclude ~ excel's temporary files
        file_names = [f for f in glob.glob(self.source_dir+'\\*.xlsx') if "~" not in f]
        return [f for f in file_names if self.sheet_exists(f)]

    #check if there is "Přehled" sheet
    def sheet_exists(self,excel_file):
        if not os.path.isfile(excel_file):
            sg.popup_error("File or sheet 'Přehled' not found.")
            print(f"soubor nenalezen: {excel_file}")
            return False
        wb = load_workbook(excel_file, read_only=True)   # open an Excel file and return a workbook
        result = self.sheet_name in wb.sheetnames
        wb.close()
        return result
    #get Vyber sablony on Uvod a postup vyplnovani sheet C4
    def get_template_name(fself,excel_file):
        if not os.path.isfile(excel_file):
            print(f"soubor nenalezen: {excel_file}")
            return False
        wb = load_workbook(excel_file, read_only=True)   # open an Excel file and return a workbook
        value =  wb.worksheets[0].cell(row=4,column=3).value
        wb.close()
        return value        

    #calculate sub report
    def calculate_subreport(self,excel_file):
        warnings.simplefilter(action='ignore', category=UserWarning)
        db = pd.read_excel(excel_file,sheet_name=self.sheet_name,usecols="C,D,E,F:H",names=self.cols_names,skiprows=1)
        db = db.applymap(lambda x: x.lower().strip() if isinstance(x, str) else x)
        
        # Nahradit "projektové vzdělávání (ve škole / mimo školu)" za "projektové vzdělávání / projektová výuka"
        db['forma'] = db['forma'].replace('projektové vzdělávání (ve škole / mimo školu)', 'projektové vzdělávání / projektová výuka')
        
        # Nahradit "propojování formálního a neformálního vzdělávání" za "propojování neformálního a formálního vzdělávání"
        db['forma'] = db['forma'].replace('propojování formálního a neformálního vzdělávání', 'propojování neformálního a formálního vzdělávání')
        
        db['ca'] = db['ca'].astype(str)+str(hash(excel_file))
        db=db.dropna()
        db['pocet_hodin'] = db['pocet_hodin'].astype(int)
        db['hash_jmena'] = db['jmena'].apply(self.custom_hash)
        #16.8.2024 pridan sloupec s nazvem sablony
        template_name = self.get_template_name(excel_file)
        db['sablona'] = template_name        
        forma = self.aggregate(db, "forma", "forma")
        tema = self.aggregate(db, "tema", "téma")
        return db, pd.concat([forma, tema], axis="rows")
    
    #vypocita hash retezce
    def custom_hash(self, value:str) -> int:
        hash_value = "".join([x.lower() for x in value if x.isalnum()])
        hash_value = hash("".join(sorted(hash_value)))
        return hash_value

    #consolidates report
    def generate_report(self):  # sourcery skip: raise-specific-error
        excel_files = self.get_files() if self.source_file is None else self.is_expected_file_format(self.source_file)
        if not len(excel_files):
            sg.popup_error("No file to process! Folder: "+self.source_dir)
            return False
            #raise Exception("Nenalezeny žádné soubory ke zpracování")
        print("Soubory ke zpracování:")
        for f in excel_files:
            print(f"{os.path.split(f)[1]}")

        self.load_unique_names()

        concatenated = pd.DataFrame()

        for f in excel_files:
            db, r = self.calculate_subreport(f)
            concatenated = pd.concat([concatenated, db], axis="rows")
            self.html_report.add_h(2,f)
            self.html_report.add_html(build_table(r, 'blue_light'))

        # jestli name_list neni prazdny, vyrad radky s jmeny v name_list
        if len(self.name_list):
                concatenated = concatenated[~concatenated['jmena'].isin(self.name_list)]


        #vypocitam hash jmena - vyloucim variantu prehozeni jmeno/prijmeni
        concatenated['hash_jmena'] = concatenated['jmena'].apply(self.custom_hash)
        
        #pro evaluaci spocitam pocty zaku v roce a tematu
        # ev = self.aggregate_evaluate(concatenated)
        
        xxx = self.aggregate_data_by_sablona(concatenated)
        print(xxx)
        #self.html_report.add_html(build_table(ev,'blue_light'),True)

        for sablona, data in xxx.items():
            self.html_report.add_html(build_table(data,'blue_light'),True)            
            self.html_report.add_h(3,f"Šablona: {sablona}", True)
        self.html_report.add_h(2,f"SDP ZZoR", True)    


        r1 = self.aggregate(concatenated, "forma", "forma")
        r2 = self.aggregate(concatenated, "tema", "téma")
        un, un_count = self.get_unique_names(concatenated)
        #unique_names = self.merge_unique_names(un,self.name_list)

        result =  pd.concat([r1, r2], axis="rows")
        result.columns = self.result_cols_names
        self.html_report.add_html(build_table(result, 'blue_light'),True)
        #self.html_report.add_h(3,f"Unikátní žáci do další ZoR: {len(unique_names)}", True)
        #self.html_report.add_h(3,f"Unikátní žáci z minulých ZoR: {len(self.name_list)}", True)
        #self.html_report.add_h(3,f"Unikátní žáci v aktuální ZoR: {len(un)}", True)
        self.html_report.add_h(3,f"Unikátní žáci v ZoR: {un_count}", True)
        
        self.html_report.add_h(2,"Údaje do ZoR", True)
        self.html_report.add_h(1,"Specifické datové položky pro ZoR", True)

        return self.html_report.html_output, un


    def aggregate(self, df, type, col_name):
        r1 = df.groupby([type], group_keys=False).agg(self.agg_function[0]).reset_index()
        r2 = df.groupby(["ca",type,"pocet_hodin"], group_keys=False).agg(self.agg_function[0]).reset_index()\
            .groupby([type],group_keys=False).agg(self.agg_function[1]).reset_index()
        result = r1.merge(r2, on=type, how="left")
        result['typ'] = col_name
        result.columns  = self.cols_agg
        return result
    
    def aggregate_evaluate(self, df):
        # Create a copy of the DataFrame
        df_copy = df.copy()
        
        # Convert 'datum' to datetime if it's not already
        df_copy['datum'] = pd.to_datetime(df_copy['datum'], format='%d.%m.%Y')
        
        # Define the date ranges
        date_ranges = [
            ('2022-2023', '2022-09-01', '2023-08-31'),
            ('2023-2024', '2023-09-01', '2024-08-31'),
            ('2024-2025', '2024-09-01', '2025-08-31'),
            ('2025-2026', '2025-09-01', '2026-08-31')
        ]
        
        # Create a new column 'period' based on the date ranges
        df_copy['period'] = pd.cut(
            df_copy['datum'],
            bins=[pd.to_datetime(start) for _, start, _ in date_ranges] + [pd.to_datetime('2026-09-01')],
            labels=[label for label, _, _ in date_ranges],
            include_lowest=True
        )
        
        # Group by tema and period, then count unique 'hash_jmena' values
        result = df_copy.groupby(['tema', 'period'])['hash_jmena'].nunique().unstack(fill_value=0)
        
        # Calculate the total sum for each row
        result['total'] = result.sum(axis=1)
        
        # Sort the result by the total column, in descending order
        result = result.sort_values('total', ascending=False)
        
        # Remove the total column
        result = result.drop('total', axis=1)
        
        # Reset the index to make 'tema' a regular column
        result = result.reset_index()
        
        # Reorder columns to have 'tema' first
        cols = ['tema'] + [col for col in result.columns if col != 'tema']
        result = result[cols]
        
        # Remove the name from the column index
        result.columns.name = None
        
        return result

    def aggregate_data_by_sablona(self, df):
        df_copy = df.copy()
        df_copy['datum'] = pd.to_datetime(df_copy['datum'], format='%d.%m.%Y')
        
        df_copy['period'] = pd.cut(
            df_copy['datum'],
            bins=[pd.to_datetime(start) for _, start, _ in date_ranges] + [pd.to_datetime('2026-09-01')],
            labels=[label for label, _, _ in date_ranges],
            include_lowest=True
        )
        
        def aggregate_single_sablona(sablona_df):
            sablona_df = sablona_df.sort_values('datum')
            
            # Dictionary to keep track of the first period for each participant-tema combination
            participant_tema_first_period = {}
            
            result = pd.DataFrame(columns=['tema'] + [label for label, _, _ in date_ranges])
            
            for period, _, _ in date_ranges:
                period_df = sablona_df[sablona_df['period'] == period]
                
                tema_counts = {}
                for _, row in period_df.iterrows():
                    participant = row['hash_jmena']
                    tema = row['tema']
                    key = (participant, tema)
                   
                    if key not in participant_tema_first_period:
                        participant_tema_first_period[key] = period
                        tema_counts[tema] = tema_counts.get(tema, 0) + 1
                #13.11.2024 - Sarka Bubelini viz mail
                # participant_tema_first_period = {}

                tema_counts_df = pd.DataFrame(list(tema_counts.items()), columns=['tema', period])
                if result.empty:
                    result = tema_counts_df
                else:
                    result = result.merge(tema_counts_df, on='tema', how='outer')
            
            result = result.fillna(0)
            
            # Ensure all period columns are present
            for period, _, _ in date_ranges:
                if period not in result.columns:
                    result[period] = 0
            
            # Reorder columns
            cols = ['tema'] + [label for label, _, _ in date_ranges]
            result = result[cols]
                       
            # Convert count columns to integers
            for col in result.columns[1:]:
                result[col] = result[col].astype(int)


            # Create a categorical column for sorting
            result['tema_cat'] = pd.Categorical(result['tema'], categories=tema_order, ordered=True)
            
            # Sort by the categorical column and drop it
            result = result.sort_values('tema_cat').drop('tema_cat', axis=1)
            

            # Ensure all temas are present, even if they have no data
            for tema in tema_order:
                if tema not in result['tema'].values:
                    new_row = pd.DataFrame({'tema': [tema], **{col: [0] for col in result.columns if col != 'tema'}})
                    result = pd.concat([result, new_row], ignore_index=True)
            
            # Final sort to ensure correct order
            result = result.set_index('tema').loc[tema_order].reset_index()                

            # Remove rows where all count values are 0
            result = result[result.iloc[:, 1:].sum(axis=1) > 0]

            return result
        
        grouped = df_copy.groupby('sablona')
        aggregated_data = {sablona: aggregate_single_sablona(group) for sablona, group in grouped}
        
        return aggregated_data



    # extract unique names
    def get_unique_names(self, df):
        if not "jmena" in df.columns:
            return []
        data = df.drop_duplicates(subset=["jmena"])
        #oprava duplicit 21.6.2024
        data = df.drop_duplicates(subset=["hash_jmena"]).sort_values(by=["jmena"])
        unique_hashes_count = df["hash_jmena"].nunique()
        return data[["jmena","hash_jmena"]].values.tolist(), unique_hashes_count
    
    # reads names from file
    def load_unique_names(self):
        self.name_list=[]
        rfile=self.name_list_source
        if not os.path.isfile(rfile):
            return        
        try:
            with open(rfile, 'r',encoding="utf-8") as f:
                for line in f:
                    self.name_list.append(line.strip())
                f.close
        except Exception as e:
            print("Error: ", e)
            self.name_list=[]

    def merge_unique_names(self,l1,l2):
        return list(set(l1+l2))


def generate_report(my_dir,name_list):
    report = SPC(my_dir,name_list)
    result, unique_names = report.generate_report()
    if not result:
        return
    rfile = os.path.join(my_dir,"result.html")
    with open(rfile, 'w',encoding="cp1250") as f:
        f.write(result)
        f.flush()
        f.close()        
    print(f'Result saved: {rfile}')

    rfile = os.path.join(my_dir,"seznam_zaku.txt")
    with open(rfile, 'w',encoding="utf-8") as f:
        for value in unique_names:
            f.write(f"{value[0]}\t\t\t{value[1]}\n")
        f.flush()
        f.close()
    

    sg.popup_no_titlebar("Done! "+rfile )


def main():
    """
    parser = argparse.ArgumentParser(description='Generate report from xlxs files')
    parser.add_argument('path', nargs="*", help="provide path to files")
    args = parser.parse_args()
    my_dir = os.getcwd()
    excel_file = None
    if len(args.path):
         if os.path.isdir(args.path[0]):
            my_dir = args.path[0]
         if os.path.isfile(args.path[0]):
            excel_file = args.path[0]
            my_dir = os.path.split(excel_file)[0]

    report = SPC(my_dir,source_file=excel_file)
    result = report.generate_report()
    rfile = os.path.join(my_dir,"result.html")
    with open(rfile, 'w') as f:
        f.write(result)
    print(f'Result saved: {rfile}')
    """
    warnings.simplefilter('ignore')
    theme = "DarkTeal10"
    font_family = "Arial"
    font_size = 14
    sg.theme(theme)

    main_window()


if __name__ == "__main__":
    main()
