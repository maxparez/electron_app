from genericpath import isfile
import pandas as pd
import numpy as np
import glob
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
import argparse
import os
import warnings
from pathlib import Path
import PySimpleGUI as sg
import xlwings as xw
import warnings
import time
import datetime
from output_manager import OutputManager
warnings.filterwarnings('ignore')

version = "0.16"


# Version-specific constants

VERSIONS = {
    "32": {
        "hours": 32,
        "template":{
                    "B1" : "32 hodin"
                    },
        "output_prefix": "32_hodin_inovativniho_vzdelavani",
        "short_prefix": "32_inv",
        "sheet" : "List1",
        "cells" : {
                    "B6" : "datum aktivity",
                },
         "skiprows": 9,
        "hours_cell": "C10",
        "hours_total_cell" : "A10",
        "version_identifier": "I",
        "column_head_names": ["to_drop","datum","forma","tema","ucitel","hodin"],
        "export_columns": ["datum","hodin","forma","tema","ucitel"]
    },
    "16": {
        "hours": 16,
        "template":{
                    "B1" : "16 hodin"
                    },
        "output_prefix": "16_hodin_inovativniho_vzdelavani",
        "short_prefix": "16_inv",
        "sheet" : "zdroj-dochazka",
        "cells" : {
                    "B6" : "datum aktivity",
                    "B7" : "čas zahájení"
                },
        "skiprows": 10,
        "hours_cell": "C11",
        "hours_total_cell" : "A11",
        "version_identifier": "II",
        "column_head_names": ["to_drop","datum","cas","forma","tema","ucitel","hodin"],
        "export_columns": ["datum","cas","hodin","forma","tema","ucitel"]
    }
}


#čísla sloupců v excelu C=3, AF=32
column_numbers = range(2,50)
column_names = [get_column_letter(num) for num in column_numbers]
column_head_names=["to_drop","datum","forma","tema","ucitel","hodin"]
skiprows = [x for x in range(11) if x !=5]
sdp_sum_forma_range = "C4:C10"
sdp_sum_tema_range = "C12:C27"
sdp_sheet_name = "SDP"
hours_total_cell = "A10"



def is_valid_path(filepath):
    if filepath and Path(filepath).exists():
        return True
    sg.popup_error("Filepath not correct")
    return False

def main_window(output_manager=None):
    # ------ Menu Definition ------ #
    menu_def =[["Help",["About"]]]


    # ------ GUI Definition ------ #
    layout = [[sg.MenubarCustom(menu_def, tearoff=False)],
              [sg.T("Složka s docházkou:", s=15, justification="r"), sg.I(key="-FOLDER-"), sg.FolderBrowse(initial_folder=os.getcwd())],
              [sg.T("Vzor/Template:", s=15, justification="r"), sg.I(key="-TEMPLATEFILE-"),
               sg.FileBrowse(initial_folder=os.getcwd(),file_types=(("XLSX Files", "*.xlsx"),))],
              [sg.Checkbox('Optimalizovat', default=False ,key="-OPTMZ-")],
              [sg.Checkbox('Zachovat název souboru', default=True ,key="-KEEPN-")],
              [sg.Exit(s=16, button_color="tomato"), sg.B("Run", s=16)],
              [sg.Multiline(size=(70, 10), key='-OUTPUT-', autoscroll=True, auto_refresh=True)],
              [sg.Button('Clear')]
              ]

    window_title = "Docházka -> MŠMT report"
    window = sg.Window(window_title, layout, use_custom_titlebar=True,finalize=True)
    if output_manager is not None:
        output_manager.start(window)

    while True:
        win, event, values = sg.read_all_windows()
        if win==window:
            if event in (sg.WINDOW_CLOSED, "Exit"):
                break
            if event == "About":
                window.disappear()
                sg.popup(window_title, "Verze "+version, "Z docházky pro šablonu Inovativní vzdělávání vytvoří MŠMT report ", grab_anywhere=True)
                window.reappear()
            if event == "Run":
                if (is_valid_path(values["-FOLDER-"])):
                    generate_org_file(gui=True, my_dir=values["-FOLDER-"],
                    template=values["-TEMPLATEFILE-"],keep_filename=bool(values["-KEEPN-"]),
                    optimize=bool(values["-OPTMZ-"]),
                    output_manager=output_manager)
        if win == output_manager.window:
            output_manager.do_event(win,event)


    window.close()

class CopyToOrg:
    def __init__(self,source_dir:str,**kwargs) -> None:
        self.source_dir=source_dir
        self.source_file = kwargs.get('source_file', None)
        self.template = kwargs.get('template', None)
        self.output_file = kwargs.get('output_file')
        self.keep_filename = kwargs.get('keep_filename', None)
        self.optimize_source = kwargs.get('optimize', None)
        self.om = kwargs.get('output_manager', None)
        self.hours_total = 0
        self.version = None

    def _detect_template_version(self, template_path: str) -> str:
        """Detect version from template content (Excel file), checking only first sheet."""
        try:
            wb = load_workbook(template_path, read_only=True)
            sheet = wb.worksheets[0]  # Get the first sheet
            for version, config in VERSIONS.items():
                match = True
                for cell, expected_value in config["template"].items():
                    actual_value = sheet[cell].value
                    if expected_value.lower() not in str(actual_value).lower():
                        match = False
                        break
                if match:
                     wb.close()
                     return version
            wb.close()
            return None
        except Exception as e:
            self.om.print(f"Chyba při detekci verze šablony: {str(e)}", "red")
            return None

    def _detect_source_version(self, source_file: str) -> str:
        """Detect version from source content (Excel file)."""
        try:
            wb = load_workbook(source_file, read_only=True)
            sheet = wb.worksheets[0] # Get the first sheet
            for version, config in VERSIONS.items():
                match = True # Reset the match before each version check
                for cell, expected_value in config["cells"].items():
                    actual_value = sheet[cell].value
                    if expected_value.lower() not in str(actual_value).lower():
                        match = False
                        break # break the inner loop
                if match:
                    wb.close()
                    return version
            wb.close()
            return None
        except Exception as e:
            self.om.print(f"Chyba při detekci verze zdrojového souboru: {str(e)}", "red")
            return None

    def _validate_version_match(self, source_file: str, template_path: str) -> bool:
        """Validate that source data version matches template version."""
        try:
            template_version = self._detect_template_version(template_path)
            source_version = self._detect_source_version(source_file)

            if source_version is None:
                self.om.print(f"Nepodařilo se detekovat verzi zdrojového souboru: {os.path.basename(source_file)}", "red")
                self.om.print("Zkontrolujte formát zdrojového souboru", "red")
                return False

            if template_version is None:
                self.om.print(f"Nepodařilo se detekovat verzi šablony: {os.path.basename(template_path)}", "red")
                self.om.print("Zkontrolujte formát šablony", "red")
                return False

            # Check if hours match
            template_hours = VERSIONS[template_version]["hours"]
            source_hours = VERSIONS[source_version]["hours"]

            if template_hours != source_hours:
                self.om.print(f"Neshoda verzí: Zdrojový soubor '{os.path.basename(source_file)}' má {source_hours} hodin, ale šablona '{os.path.basename(template_path)}' je pro {template_hours} hodin", "red")
                self.om.print("Použijte prosím správnou verzi šablony", "red")
                return False

            # For 16-hour version, check if it's the correct variant (I or II)
            if template_hours == 16:
                template_identifier = VERSIONS[template_version]["version_identifier"]
                source_identifier = VERSIONS[source_version]["version_identifier"]

                if template_identifier != source_identifier:
                    self.om.print(f"Neshoda verzí: Zdrojový soubor '{os.path.basename(source_file)}' je verze {source_identifier}, ale šablona '{os.path.basename(template_path)}' je pro verzi {template_identifier}", "red")
                    self.om.print("Použijte prosím správnou verzi šablony", "red")
                    return False

            self.version = source_version
            return True

        except Exception as e:
            self.om.print(f"Chyba při ověřování verzí: {str(e)}", "red")
            return False


    #check file is ok
    def is_expected_file_format(self, file_name):
        if not os.path.isfile(file_name):
            self.om.print(f"Soubor nenalezen: {file_name}", "red")
            return []

        if not self.sheet_exists(file_name):
            self.om.print(f"List 'List1' nebo zdroj-dochazka nenalezen v: {file_name}", "red")
            return []

        # Check the version before returning the file name
        version = self._detect_source_version(file_name)
        if version is None:
            self.om.print(f"Nepodařilo se detekovat verzi zdrojového souboru: {file_name}", "red")
            return []

        return [file_name]

    # get only relavant files
    def get_files(self):
        # exclude ~ excel's temporary files
        file_names = [f for f in glob.glob(os.path.join(self.source_dir, '*.xls*')) if "~" not in f]
        return [f for f in file_names if self.sheet_exists(f)]

    #check if there is "Přehled" sheet
    def sheet_exists(self,excel_file):
        if not os.path.isfile(excel_file):
            sg.popup_error("File or sheet 'List1' not found.")
            self.om.print(f"soubor nenalezen: {excel_file}")
            return False
        wb = load_workbook(excel_file, read_only=True)   # open an Excel file and return a workbook
        value =  wb.worksheets[0].cell(row=6,column=2).value
        result = (value is not None) and ("datum" in str(value).lower())
        wb.close()
        return result

    def first_emtpy_column(self,excel_file,row=6):
        wb = load_workbook(excel_file, read_only=True)   # open an Excel file and return a workbook
        for colmn in range(3,50) :
         if wb.worksheets[0].cell(row=row,column=colmn).value is None:
             break
        return colmn-1

    def is_valid_date(self,date_str):
        try:
            # Attempt to parse the date string with the specified format
            datetime.datetime.strptime(date_str, "%d.%m.%Y")
            return True
        except ValueError:
            return False

    def verify_dates(self,source_file):
        last_column = self.first_emtpy_column(source_file,6)
        wb = load_workbook(source_file)
        sheet =  wb.worksheets[0]
        result = True
        invalid_dates=[]
        for row in sheet.iter_rows(min_row=6, max_row=6, min_col=3, max_col=last_column):
            for cell in row:
                if not cell.is_date  and not self.is_valid_date(str(cell.value)):
                    result = False
                    column_letter = get_column_letter(cell.column)
                    row_number = cell.row
                    invalid_dates.append(f"{column_letter}{row_number} - {str(cell.value)}")
        if not result:
            self.om.print(f"Chybný formát dat: {', '.join(invalid_dates)}")
            self.om.print("Soubor nebude zpracován, prosím opravte.")
            self.om.print("------------------------------------------------")

        return result


    def sum_cell_range(self,worksheet, cell_range):
        total = 0
        for row in worksheet[cell_range]:
            for cell in row:
                if cell.value is not None:
                    total += cell.value
        return total

    def get_hours_total(self, source_file):
        wb = load_workbook(source_file, keep_vba=True, data_only=True)
        # Determine total hours based on the detected source version from config
        if self.version and "hours_total_cell" in VERSIONS[self.version]:
            hours_total_cell = VERSIONS[self.version]["hours_total_cell"]
        else:
            hours_total_cell = "A10"  # Default to A10 if version is not detected or config has no "hours_total_cell" info
            self.om.print(f"Hours_total_cell config not found in version config using default {hours_total_cell}", "yellow")        
        value = wb.worksheets[0][hours_total_cell].value
        wb.close()
        return value


    def prepare_data(self, source_file):
        # replace strings containing 'ano' with 1, case insensitive
        def replace_ano_with_1(cell_value):
            result = 0
            if isinstance(cell_value, str) and 'ano' == cell_value.lower().strip():
                result =  1
            return result

        AH = column_index_from_string("AH")
        AW = column_index_from_string("AW")
        last_column = self.first_emtpy_column(source_file, 6)
        last_column_letter = "AW" if last_column > AW else get_column_letter(last_column)
        if not self.optimize_source:
            last_column_letter = get_column_letter(min(last_column, AH))

        usecols = f"B:{last_column_letter}"

        # Determine skiprows based on the detected source version from config
        if self.version and "skiprows" in VERSIONS[self.version]:
            skiprows = VERSIONS[self.version]["skiprows"]
        else:
            skiprows = 9  # Default to 9 if version is not detected or config has no "skiprows" info
            self.om.print(f"Skiprows config not found in version config using default {skiprows}", "yellow")

        # dochazka
        df = pd.read_excel(source_file, usecols=usecols, skiprows=skiprows, dtype=str)
        df.dropna(how='all', inplace=True)
        df.fillna(0, inplace=True)
        # df=df.replace("ano",1)
        # nahradim ano za 1
        df.iloc[0:, 1:] = df.iloc[0:, 1:].applymap(replace_ano_with_1)
        df.columns = column_names[:df.shape[1]]
        #odebere radky kdy jsou vsechny hodiny 0
        df = self.drop_zero_rows(df,df.columns[1:])
        # vyberu zaky
        students = df["B"]
        df = df.drop(["B"], axis=1)
        df[df.columns] = df[df.columns].astype(int)
        # hlavicka tabulky
        usecols = f"C:{last_column_letter}"
        df_head_t = pd.read_excel(source_file, usecols=usecols, skiprows=4, dtype=str)
        
        # Determine column_head_names based on the detected source version from config
        if self.version and "column_head_names" in VERSIONS[self.version]:
            head_names = VERSIONS[self.version]["column_head_names"]
            df_head_t = df_head_t.drop(df_head_t.index[(len(head_names)-1):]) # Drop rows after number of columns
        else:
            df_head_t = df_head_t.drop(df_head_t.index[5:]) # default behaviour
            self.om.print(f"column_head_names config not found in version config using default", "yellow")
            head_names = ["to_drop","datum","forma","tema","ucitel","hodin"] # default values

        df_head = df_head_t.transpose().reset_index()
        df_head.columns = head_names
        df_head["new_index"] = df.columns
        df_head = df_head.set_index('new_index')
        df_head["datum"] = pd.to_datetime(df_head['datum']).dt.date
        df_head = df_head.drop("to_drop", axis=1)
        df_head.dropna(how='all', inplace=True)
        # pocty hodin z hlavicky vynasobim sloupce
        # upravim pocet sloupcu
        df = df[df_head.index.values]
        df = df.multiply(df_head["hodin"].astype(int).values, axis=1)
        self.df = df
        self.df_head = df_head
        self.students = students
        self.hours_total = self.get_hours_total(source_file)

    def drop_zero_rows(self, df, columns):
        """Drops rows from a DataFrame where all values are zero.
        Args:
            df: The pandas DataFrame.
        Returns:
            A new DataFrame with the zero rows removed.
        """
        # Use any(axis=1) to detect if any value in a row is not zero, then ~ (Not) to select zero rows.
        non_zero_rows = df[columns].any(axis=1)
        df_filtered = df[non_zero_rows]
        return df_filtered

    def optimize(self,tdf:pd.DataFrame,optimize=True):
        data = tdf
        if optimize:
            # pokud je sloupcu/aktivit vice jak 32
            # pokusim se vybrat ty aktivity, aby co nejvice zaku splnilo 32h
            data['row_sum']=0
            # Calculate the sum of values for each row
            data['row_sum'] = data.sum(axis=1)
            # Filter the rows where the sum is greater than or equal to 32
            filtered_data = data[data['row_sum'] >= 32]
            # Select the top 32 columns with the highest row sums in the filtered data
            selected_columns = filtered_data.iloc[:, :-1].sum().nlargest(32).index.tolist()
        else:
            selected_columns = data.columns[:32].tolist()
        # return the sorted selected columns
        selected_columns = [c for c in tdf.columns if c in selected_columns]
        to_remove_columns = [c for c in tdf.columns if c not in selected_columns]
        if 'row_sum' in to_remove_columns:
            to_remove_columns.remove('row_sum')
        if to_remove_columns:
            self.om.print(f"Sloupce k smazání v excelu {', '.join(to_remove_columns)}")
        return selected_columns, to_remove_columns

    def delete_columns_and_save(self,filename, columns):
        wb = xw.Book(filename)
        sheet = wb.sheets[0]
        for column in columns[::-1]:
            sheet.range(f'{column}:{column}').api.Delete()

        nfn = os.path.join(self.source_dir,f"upraveno_{os.path.split(filename)[1]}")
        wb.save(nfn)
        wb.close()

    def create_orginal_file(self, output_file="",template_file=""):
        app = xw.App(visible=True)
        # kopirovani do vyslednoho souboru
        template_file = template_file if template_file !="" else self.template
        wb = xw.Book(template_file)

        #sheet = wb['Seznam účastníků']
        sheet = wb.sheets['Seznam účastníků']
        #vloz jmena
        sheet.range("B4").options(ndim="expand", transpose=True).value = self.students.values.tolist()

        #aktivity
        # Determine export_columns based on the detected source version from config
        if self.version and "export_columns" in VERSIONS[self.version]:
            export_columns = VERSIONS[self.version]["export_columns"][:]           
        else:
            self.om.print(f"export_columns config not found in version config using default", "yellow")
            export_columns = ["datum","forma","tema","ucitel","hodin"] # default values

        sheet = wb.sheets['Seznam aktivit']
        data = self.df_head.loc[self.selected_columns][export_columns]
        sheet.range("C3").options(ndim="expand").value = data.values

        #prehled
        sheet =  wb.sheets['Přehled']
        cislo_aktivity=1
        df_prehled=self.df[self.selected_columns]
        df_prehled.loc[:,("zaci")] = self.students
        #jedu po sloupcich a vytvarim seznam zaku
        z=c=[]
        for selected_col in self.selected_columns:
            z1 = df_prehled[df_prehled[selected_col]>0]["zaci"].values.tolist()
            c1 = np.full(len(z1),cislo_aktivity, dtype=int)
            z = np.append(z,z1)
            c = np.append(c,c1)
            cislo_aktivity+=1
        result = [[x, y] for x, y in zip(c, z)]
        sheet.range("C3").options(ndim="expand").value = result

        wb.save(output_file)
        self.om.print(f"Soubor {output_file} byl uložen.")
        #kontrola sdp tema, forma
        sdp_sheet = wb.sheets[sdp_sheet_name]
        sdp_forma_total = self.sum_cell_range(sdp_sheet, sdp_sum_forma_range)
        sdp_tema_total = self.sum_cell_range(sdp_sheet, sdp_sum_tema_range)
        time.sleep(0.1)
        wb.close()
        app.quit()
        return sdp_forma_total, sdp_tema_total

    def create_output(self):
        excel_files = self.get_files() if self.source_file is None else self.is_expected_file_format(self.source_file)
        if not excel_files:
            sg.popup_error("Soubory ke zpracování nebyly rozpoznány: "+self.source_dir + "\nOčekává se sešit Excelu, data začínají od B6")
            return False

        excel_files = [excel_files] if type(excel_files) is str else excel_files
        self.om.print("Rozpoznané soubory ke zpracování:")
        self.om.print("---------------------------------")
        remove_files=[]
        for f in excel_files:
            self.om.print(f"{os.path.split(f)[1]}")
            if not self.verify_dates(f):
                remove_files.append(f)
            time.sleep(0.1)
        if remove_files:
            excel_files =[item for item in excel_files if item not in remove_files]

        self.om.print("\nUkládám výsledné soubory:")
        self.om.print("-------------------------")
        for c, f in enumerate(excel_files, start=1):
            try:
                # Validate version match, and set self.version
                if not self._validate_version_match(f, self.template):
                    continue # skip to the next file

                self.prepare_data(f)
                self.selected_columns, to_remove_columns = self.optimize(self.df,optimize=self.optimize_source)
                if to_remove_columns:
                    self.delete_columns_and_save(f,to_remove_columns)

                # Get output prefix based on the detected version
                output_prefix = VERSIONS[self.version]["output_prefix"]
                short_prefix = VERSIONS[self.version]["short_prefix"]

                fn = os.path.join(self.source_dir, f"{output_prefix}_{c}.xlsx")
                org_file_name = (os.path.split(f)[1]).split(".")[0]
                kfn = os.path.join(self.source_dir, f"{short_prefix}_{org_file_name}.xlsx")
                output_filename = kfn if self.keep_filename else fn

                sdp_forma_total, sdp_tema_total = self.create_orginal_file(output_filename)

                if sdp_forma_total != self.hours_total or sdp_tema_total != self.hours_total:
                    self.om.print(f"Nesouhlasí SDP s počtem inovativních hodin!", "red")
                    self.om.print(f"Počet inv. hodin: {self.hours_total}")
                    self.om.print(f"SDP forma : {sdp_forma_total}")
                    self.om.print(f"SDP téma: {sdp_tema_total}")
                    self.om.print(f"Zkontrolujte aktivity na listu 'Seznam aktivit'\n")

            except Exception as e:
                print("-----------------")
                print(f"Chyba v souboru {f}")
                print("Error: ", e)
                print("-----------------")
                self.om.print("-----------------")
                self.om.print(f"Chyba v souboru {f}")
                self.om.print(f"Error: {e}")
                self.om.print("-----------------")
                time.sleep(0.1)

        return True



def generate_org_file(my_dir,template,source_file=None,gui=True,keep_filename=False, optimize=True,output_manager=None):
    gen = CopyToOrg(my_dir,template=template,source_file=source_file,keep_filename=keep_filename,optimize=optimize,
                    output_manager=output_manager)
    result = gen.create_output()
    if gui:
        sg.popup_no_titlebar("Done!")



def main():
    parser = argparse.ArgumentParser(description='Generates info file for project')
    parser.add_argument('-g', help="GUI", default=1, required=False, type=int)
    parser.add_argument('-f', help="working folder or source file")
    parser.add_argument('-t', help="template file")
    parser.add_argument('-k', help="keep file name", default=0, required=False, type=int)
    parser.add_argument('-o', help="optimalize", default=1, required=False, type=int)
    args = parser.parse_args()
    my_dir = os.getcwd()
    args = parser.parse_args()
    gui = bool(args.g)

    #print(args)

    if args.f:
        my_dir = args.f
    if args.t:
        template = args.t
    keep_filename = bool(args.k)
    optimize = bool(args.o)
    output_manager = OutputManager()
    if gui:
        theme = "DarkTeal10"
        font_family = "Arial"
        font_size = 14
        sg.theme(theme)
        main_window(output_manager)
        return
    source_file = None
    if os.path.isfile(my_dir):
        source_file = my_dir
        my_dir = os.path.dirname(source_file)

    generate_org_file(my_dir,template,source_file=source_file,gui=False,keep_filename=keep_filename,optimize=optimize,output_manager=output_manager)


if __name__ == "__main__":
    main()