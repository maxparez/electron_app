"""
Zpracování specifických datových položek pro ZoR
Version: 0.15
"""

from __future__ import annotations
import os
import glob
import warnings
import argparse
from pathlib import Path
from typing import List, Tuple, Dict, Optional, Any
from dataclasses import dataclass, field

import pandas as pd
from openpyxl import load_workbook
from pretty_html_table import build_table


# ==================== Configuration ====================

VERSION = "0.15"
DEFAULT_SHEET_NAME = "Přehled"
DEFAULT_ENCODING = "cp1250"
OUTPUT_ENCODING = "utf-8"

# Date ranges for periods
DATE_RANGES = [
    ('2022-2023', '2022-09-01', '2023-08-31'),
    ('2023-2024', '2023-09-01', '2024-08-31'),
    ('2024-2025', '2024-09-01', '2025-08-31'),
    ('2025-2026', '2025-09-01', '2026-08-31')
]

# Theme order for sorting
TEMA_ORDER = [
    "čtenářská pre/gramotnost",
    "matematická pre/gramotnost",
    "umělecká gramotnost",
    "mediální gramotnost",
    "cizí jazyky/komunikace v cizím jazyce",
    "inkluze včetně primární prevence",
    "přírodovědné a technické vzdělávání",
    "evvo a vzdělávání pro udržitelný rozvoj",
    "vzdělávání s využitím nových technologií",
    "kulturní povědomí a vyjádření",
    "historické povědomí, výuka moderních dějin",
    "rozvoj podnikavosti a kreativity",
    "well-being a psychohygiena",
    "genderová tematika v obsahu vzdělávání",
    "kariérové poradenství včetně identifikace a rozvoje nadání",
    "občanské vzdělávání a demokratické myšlení"
]

# Column mappings
COLUMN_NAMES = ["ca", "jmena", "datum", "pocet_hodin", "forma", "tema"]
AGGREGATE_COLUMNS = ["forma/téma", "čislo", "cena", "typ"]
RESULT_COLUMN_NAMES = ["forma/téma", "číslo celkem", "cena celkem", "typ"]

# Text replacements for normalization
TEXT_REPLACEMENTS = {
    'forma': {
        'projektové vzdělávání (ve škole / mimo školu)': 'projektové vzdělávání / projektová výuka',
        'propojování formálního a neformálního vzdělávání': 'propojování neformálního a formálního vzdělávání'
    }
}


# ==================== Data Classes ====================

@dataclass
class ProcessingResult:
    """Result of processing Excel files"""
    html_output: str
    unique_names: List[Tuple[str, int]]
    success: bool = True
    error_message: Optional[str] = None


@dataclass
class AggregationConfig:
    """Configuration for data aggregation"""
    agg_functions: List[Dict[str, str]] = field(default_factory=lambda: [
        {"hash_jmena": "nunique"},
        {"pocet_hodin": "sum"}
    ])


# ==================== Utility Functions ====================

def is_valid_path(filepath: str) -> bool:
    """Check if filepath exists and is valid"""
    if filepath and Path(filepath).exists():
        return True
    print(f"Error: Filepath not correct: {filepath}")
    return False


def custom_hash(value: str) -> int:
    """Calculate hash of a string for name comparison"""
    hash_value = "".join([x.lower() for x in value if x.isalnum()])
    hash_value = hash("".join(sorted(hash_value)))
    return hash_value


def normalize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize dataframe by applying text replacements and transformations"""
    df = df.copy()
    
    # Apply lowercase and strip to string columns
    df = df.applymap(lambda x: x.lower().strip() if isinstance(x, str) else x)
    
    # Apply text replacements
    for column, replacements in TEXT_REPLACEMENTS.items():
        if column in df.columns:
            for old_text, new_text in replacements.items():
                df[column] = df[column].replace(old_text, new_text)
    
    return df


# ==================== HTML Report Generator ====================

class HtmlReportGenerator:
    """Generates HTML reports with consistent styling"""
    
    def __init__(self):
        self.body_content = ""
        self.template = """
        <!DOCTYPE html>
        <html lang="cs">
        <head>
            <meta charset="windows-1250 ">
            <meta http-equiv="X-UA-Compatible" content="IE=edge">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Specifické datové položky - report</title>
            <style type="text/css">
                h1,h2{{color:navy;
                     font-family: Century Gothic, sans-serif; }}
            </style>
        </head>
        <body>
            {}
        </body>
        </html>
        """
    
    def add_html(self, html: str, prepend: bool = False) -> None:
        """Add HTML content to the report"""
        if prepend:
            self.body_content = html + self.body_content
        else:
            self.body_content += html
    
    def add_heading(self, level: int, text: str, prepend: bool = False) -> None:
        """Add a heading to the report"""
        heading = f"<h{level}>{text}</h{level}>"
        self.add_html(heading, prepend)
    
    @property
    def html_output(self) -> str:
        """Get the complete HTML output"""
        return self.template.format(self.body_content)


# ==================== Excel File Handler ====================

class ExcelFileHandler:
    """Handles Excel file operations"""
    
    def __init__(self, sheet_name: str = DEFAULT_SHEET_NAME):
        self.sheet_name = sheet_name
    
    def is_valid_excel_file(self, file_path: str) -> bool:
        """Check if file exists and contains the required sheet"""
        if not os.path.isfile(file_path):
            print(f"Error: File not found: {file_path}")
            return False
        
        if "~" in file_path:  # Skip temporary Excel files
            return False
        
        return self.sheet_exists(file_path)
    
    def sheet_exists(self, excel_file: str) -> bool:
        """Check if the required sheet exists in the Excel file"""
        try:
            wb = load_workbook(excel_file, read_only=True)
            result = self.sheet_name in wb.sheetnames
            wb.close()
            return result
        except Exception as e:
            print(f"Error reading Excel file {excel_file}: {e}")
            return False
    
    def get_template_name(self, excel_file: str) -> Optional[str]:
        """Get template name from Excel file (cell C4 on first sheet)"""
        try:
            wb = load_workbook(excel_file, read_only=True)
            value = wb.worksheets[0].cell(row=4, column=3).value
            wb.close()
            return value
        except Exception as e:
            print(f"Error getting template name from {excel_file}: {e}")
            return None
    
    def get_excel_files(self, directory: str) -> List[str]:
        """Get all valid Excel files from directory"""
        pattern = os.path.join(directory, '*.xlsx')
        all_files = glob.glob(pattern)
        return [f for f in all_files if self.is_valid_excel_file(f)]


# ==================== Data Processor ====================

class DataProcessor:
    """Processes and aggregates data from Excel files"""
    
    def __init__(self, aggregation_config: AggregationConfig):
        self.agg_config = aggregation_config
        self.excel_handler = ExcelFileHandler()
    
    def read_excel_data(self, excel_file: str) -> pd.DataFrame:
        """Read and preprocess Excel data"""
        warnings.simplefilter(action='ignore', category=UserWarning)
        
        df = pd.read_excel(
            excel_file,
            sheet_name=DEFAULT_SHEET_NAME,
            usecols="C,D,E,F:H",
            names=COLUMN_NAMES,
            skiprows=1
        )
        
        df = normalize_dataframe(df)
        df['ca'] = df['ca'].astype(str) + str(hash(excel_file))
        df = df.dropna()
        df['pocet_hodin'] = df['pocet_hodin'].astype(int)
        df['hash_jmena'] = df['jmena'].apply(custom_hash)
        
        # Add template name
        template_name = self.excel_handler.get_template_name(excel_file)
        df['sablona'] = template_name
        
        return df
    
    def aggregate_by_type(self, df: pd.DataFrame, group_by: str, type_name: str) -> pd.DataFrame:
        """Aggregate data by specified type (forma or tema)"""
        # First aggregation: unique names
        r1 = df.groupby([group_by], group_keys=False).agg(
            self.agg_config.agg_functions[0]
        ).reset_index()
        
        # Second aggregation: sum of hours
        r2 = df.groupby(["ca", group_by, "pocet_hodin"], group_keys=False).agg(
            self.agg_config.agg_functions[0]
        ).reset_index().groupby([group_by], group_keys=False).agg(
            self.agg_config.agg_functions[1]
        ).reset_index()
        
        result = r1.merge(r2, on=group_by, how="left")
        result['typ'] = type_name
        result.columns = AGGREGATE_COLUMNS
        
        return result
    
    def aggregate_by_sablona(self, df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
        """Aggregate data by template (šablona) and period"""
        df_copy = df.copy()
        df_copy['datum'] = pd.to_datetime(df_copy['datum'], format='%d.%m.%Y')
        
        # Create period column
        df_copy['period'] = pd.cut(
            df_copy['datum'],
            bins=[pd.to_datetime(start) for _, start, _ in DATE_RANGES] + [pd.to_datetime('2026-09-01')],
            labels=[label for label, _, _ in DATE_RANGES],
            include_lowest=True
        )
        
        grouped = df_copy.groupby('sablona')
        return {
            sablona: self._aggregate_single_sablona(group) 
            for sablona, group in grouped
        }
    
    def _aggregate_single_sablona(self, sablona_df: pd.DataFrame) -> pd.DataFrame:
        """Aggregate data for a single template"""
        sablona_df = sablona_df.sort_values('datum')
        
        # Track first occurrence of participant-tema combination
        participant_tema_first_period = {}
        result = pd.DataFrame(columns=['tema'] + [label for label, _, _ in DATE_RANGES])
        
        for period, _, _ in DATE_RANGES:
            period_df = sablona_df[sablona_df['period'] == period]
            
            tema_counts = {}
            for _, row in period_df.iterrows():
                participant = row['hash_jmena']
                tema = row['tema']
                key = (participant, tema)
                
                if key not in participant_tema_first_period:
                    participant_tema_first_period[key] = period
                    tema_counts[tema] = tema_counts.get(tema, 0) + 1
            
            tema_counts_df = pd.DataFrame(
                list(tema_counts.items()), 
                columns=['tema', period]
            )
            
            if result.empty:
                result = tema_counts_df
            else:
                result = result.merge(tema_counts_df, on='tema', how='outer')
        
        result = result.fillna(0)
        
        # Ensure all period columns are present
        for period, _, _ in DATE_RANGES:
            if period not in result.columns:
                result[period] = 0
        
        # Reorder columns
        cols = ['tema'] + [label for label, _, _ in DATE_RANGES]
        result = result[cols]
        
        # Convert count columns to integers
        for col in result.columns[1:]:
            result[col] = result[col].astype(int)
        
        # Sort by tema order
        result['tema_cat'] = pd.Categorical(
            result['tema'], 
            categories=TEMA_ORDER, 
            ordered=True
        )
        result = result.sort_values('tema_cat').drop('tema_cat', axis=1)
        
        # Ensure all temas are present
        for tema in TEMA_ORDER:
            if tema not in result['tema'].values:
                new_row = pd.DataFrame({
                    'tema': [tema], 
                    **{col: [0] for col in result.columns if col != 'tema'}
                })
                result = pd.concat([result, new_row], ignore_index=True)
        
        # Final sort and remove empty rows
        result = result.set_index('tema').loc[TEMA_ORDER].reset_index()
        result = result[result.iloc[:, 1:].sum(axis=1) > 0]
        
        return result
    
    def get_unique_participants(self, df: pd.DataFrame) -> Tuple[List[Tuple[str, int]], int]:
        """Extract unique participants from dataframe"""
        if 'hash_jmena' not in df.columns:
            return [], 0
        
        unique_df = df.drop_duplicates(subset=['hash_jmena']).sort_values(by=['jmena'])
        unique_count = df['hash_jmena'].nunique()
        
        return unique_df[['jmena', 'hash_jmena']].values.tolist(), unique_count


# ==================== Report Generator ====================

class ReportGenerator:
    """Main class for generating reports"""
    
    def __init__(self, source_dir: str, name_list_file: str = ""):
        self.source_dir = source_dir
        self.name_list_file = name_list_file
        self.excel_handler = ExcelFileHandler()
        self.data_processor = DataProcessor(AggregationConfig())
        self.html_generator = HtmlReportGenerator()
        self.excluded_names = []
    
    def load_excluded_names(self) -> None:
        """Load names to exclude from file"""
        if not self.name_list_file or not os.path.isfile(self.name_list_file):
            return
        
        try:
            with open(self.name_list_file, 'r', encoding='utf-8') as f:
                self.excluded_names = [line.strip() for line in f]
        except Exception as e:
            print(f"Error loading name list: {e}")
            self.excluded_names = []
    
    def process_single_file(self, excel_file: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """Process a single Excel file"""
        df = self.data_processor.read_excel_data(excel_file)
        
        forma_agg = self.data_processor.aggregate_by_type(df, "forma", "forma")
        tema_agg = self.data_processor.aggregate_by_type(df, "tema", "téma")
        
        combined_agg = pd.concat([forma_agg, tema_agg], axis="rows")
        
        return df, combined_agg
    
    def generate_report(self, single_file: Optional[str] = None) -> ProcessingResult:
        """Generate the complete report"""
        # Get files to process
        if single_file:
            if not self.excel_handler.is_valid_excel_file(single_file):
                return ProcessingResult(
                    html_output="",
                    unique_names=[],
                    success=False,
                    error_message=f"Invalid file: {single_file}"
                )
            excel_files = [single_file]
        else:
            excel_files = self.excel_handler.get_excel_files(self.source_dir)
        
        if not excel_files:
            return ProcessingResult(
                html_output="",
                unique_names=[],
                success=False,
                error_message=f"No files to process in: {self.source_dir}"
            )
        
        print("Files to process:")
        for f in excel_files:
            print(f"  {os.path.basename(f)}")
        
        # Load excluded names
        self.load_excluded_names()
        
        # Process all files
        all_data = pd.DataFrame()
        
        for excel_file in excel_files:
            df, aggregated = self.process_single_file(excel_file)
            all_data = pd.concat([all_data, df], axis="rows")
            
            # Add to report
            self.html_generator.add_heading(2, excel_file)
            self.html_generator.add_html(build_table(aggregated, 'blue_light'))
        
        # Exclude names if provided
        if self.excluded_names:
            all_data = all_data[~all_data['jmena'].isin(self.excluded_names)]
        
        # Process aggregations by template
        sablona_aggregations = self.data_processor.aggregate_by_sablona(all_data)
        
        for sablona, data in sablona_aggregations.items():
            self.html_generator.add_html(build_table(data, 'blue_light'), prepend=True)
            self.html_generator.add_heading(3, f"Šablona: {sablona}", prepend=True)
        
        self.html_generator.add_heading(2, "SDP ZZoR", prepend=True)
        
        # Final aggregations
        final_forma = self.data_processor.aggregate_by_type(all_data, "forma", "forma")
        final_tema = self.data_processor.aggregate_by_type(all_data, "tema", "téma")
        final_result = pd.concat([final_forma, final_tema], axis="rows")
        final_result.columns = RESULT_COLUMN_NAMES
        
        # Get unique participants
        unique_participants, unique_count = self.data_processor.get_unique_participants(all_data)
        
        # Add final sections to report
        self.html_generator.add_html(build_table(final_result, 'blue_light'), prepend=True)
        self.html_generator.add_heading(3, f"Unikátní žáci v ZoR: {unique_count}", prepend=True)
        self.html_generator.add_heading(2, "Údaje do ZoR", prepend=True)
        self.html_generator.add_heading(1, "Specifické datové položky pro ZoR", prepend=True)
        
        return ProcessingResult(
            html_output=self.html_generator.html_output,
            unique_names=unique_participants,
            success=True
        )


# ==================== File Output Handler ====================

class FileOutputHandler:
    """Handles writing output files"""
    
    @staticmethod
    def save_html_report(filepath: str, content: str) -> bool:
        """Save HTML report to file"""
        try:
            with open(filepath, 'w', encoding=DEFAULT_ENCODING) as f:
                f.write(content)
            print(f"Result saved: {filepath}")
            return True
        except Exception as e:
            print(f"Error saving HTML report: {e}")
            return False
    
    @staticmethod
    def save_name_list(filepath: str, names: List[Tuple[str, int]]) -> bool:
        """Save list of unique names to file"""
        try:
            with open(filepath, 'w', encoding=OUTPUT_ENCODING) as f:
                for name, hash_value in names:
                    f.write(f"{name}\t\t\t{hash_value}\n")
            print(f"Seznam žáků saved: {filepath}")
            return True
        except Exception as e:
            print(f"Error saving name list: {e}")
            return False


# ==================== Main Entry Point ====================

def process_directory(directory: str, name_list_file: str = "", single_file: Optional[str] = None) -> None:
    """Process directory or single file and generate reports"""
    report_generator = ReportGenerator(directory, name_list_file)
    result = report_generator.generate_report(single_file)
    
    if not result.success:
        print(f"Error: {result.error_message}")
        return
    
    # Save outputs
    output_handler = FileOutputHandler()
    
    html_path = os.path.join(directory, "result.html")
    output_handler.save_html_report(html_path, result.html_output)
    
    names_path = os.path.join(directory, "seznam_zaku.txt")
    output_handler.save_name_list(names_path, result.unique_names)
    
    print("Done!")


def main():
    """Main entry point for command line interface"""
    parser = argparse.ArgumentParser(
        description='Generate report from xlsx files',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s                      # Process current directory
  %(prog)s /path/to/directory   # Process specific directory
  %(prog)s file.xlsx            # Process single file
  %(prog)s -n names.txt         # Exclude names from file
        """
    )
    
    parser.add_argument(
        'path',
        nargs="?",
        default=os.getcwd(),
        help="Path to files or directory (default: current directory)"
    )
    parser.add_argument(
        '--names', '-n',
        help="Path to file with name list to exclude",
        default=""
    )
    parser.add_argument(
        '--version', '-v',
        action='version',
        version=f'%(prog)s {VERSION}'
    )
    
    args = parser.parse_args()
    
    # Determine processing mode
    if os.path.isdir(args.path):
        directory = args.path
        single_file = None
    elif os.path.isfile(args.path):
        single_file = args.path
        directory = os.path.dirname(single_file)
    else:
        print(f"Error: Path '{args.path}' does not exist")
        return
    
    # Validate name list file
    name_list = args.names
    if name_list and not os.path.isfile(name_list):
        print(f"Warning: Name list file '{name_list}' not found, continuing without it")
        name_list = ""
    
    # Print processing information
    print(f"Processing directory: {directory}")
    if single_file:
        print(f"Processing single file: {single_file}")
    if name_list:
        print(f"Using name list from: {name_list}")
    
    # Suppress warnings and process
    warnings.simplefilter('ignore')
    process_directory(directory, name_list, single_file)


if __name__ == "__main__":
    main()