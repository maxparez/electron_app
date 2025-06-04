#!/usr/bin/env python3
"""
Quick test to find where dates actually are
"""

import pandas as pd
import sys
import os

test_file = "tests/regression/inputs/16h/valid_basic.xlsx"
if not os.path.exists(test_file):
    print(f"File not found: {test_file}")
    sys.exit(1)

# Read Excel with openpyxl to see exact structure
from openpyxl import load_workbook

wb = load_workbook(test_file, data_only=True)
sheet = wb['zdroj-dochazka']

print("Checking cells for 16h format:")
print(f"B5: {sheet['B5'].value}")  # Should be 'datum aktivity'
print(f"B6: {sheet['B6'].value}")  # Should be 'čas zahájení'

# Check where actual dates are
print("\nChecking row 6 (dates) from column C onwards:")
for col in range(3, 15):  # C to N
    cell = sheet.cell(row=6, column=col)
    if cell.value:
        print(f"  Column {col} (Excel {chr(64+col)}): {cell.value}")

print("\nChecking row 11 (hours) from column C onwards:")
for col in range(3, 15):  # C to N
    cell = sheet.cell(row=11, column=col)
    if cell.value:
        print(f"  Column {col} (Excel {chr(64+col)}): {cell.value}")

wb.close()