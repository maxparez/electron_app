#!/usr/bin/env python3
"""
Test ZorSpecDatProcessor functionality
"""

import sys
import os
import tempfile
from pathlib import Path

# Add src path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src', 'python'))

from tools.zor_spec_dat_processor import ZorSpecDatProcessor

def test_processor_initialization():
    """Test basic processor initialization"""
    print("=== Testing ZorSpecDatProcessor Initialization ===")
    
    processor = ZorSpecDatProcessor()
    
    print(f"✅ Processor created successfully")
    print(f"   Sheet name: {processor.sheet_name}")
    print(f"   Column names: {processor.cols_names}")
    print(f"   Result columns: {processor.result_cols_names}")
    
    # Test message handling
    processor.add_info("Test info message")
    processor.add_warning("Test warning message")
    processor.add_error("Test error message")
    
    print(f"   Info messages: {len(processor.info_messages)}")
    print(f"   Warning messages: {len(processor.warnings)}")
    print(f"   Error messages: {len(processor.errors)}")

def test_validation():
    """Test input validation"""
    print("\n=== Testing Input Validation ===")
    
    processor = ZorSpecDatProcessor()
    
    # Test with no inputs
    valid = processor.validate_inputs([], {})
    print(f"❌ No inputs: {'Valid' if valid else 'Invalid'} (expected: Invalid)")
    
    # Test with valid directory (if exists)
    test_dir = "/root/vyvoj_sw/electron_app/legacy_code/inv"
    if os.path.exists(test_dir):
        processor.clear_messages()
        valid = processor.validate_inputs([], {"source_dir": test_dir})
        print(f"📁 Valid directory: {'Valid' if valid else 'Invalid'}")
        if processor.info_messages:
            for info in processor.info_messages:
                print(f"   ℹ️ {info}")
    
    # Test with invalid directory
    processor.clear_messages()
    valid = processor.validate_inputs([], {"source_dir": "/nonexistent/path"})
    print(f"❌ Invalid directory: {'Valid' if valid else 'Invalid'} (expected: Invalid)")
    for error in processor.errors:
        print(f"   ❌ {error}")

def test_file_discovery():
    """Test file discovery in directory"""
    print("\n=== Testing File Discovery ===")
    
    processor = ZorSpecDatProcessor()
    
    # Test with legacy directory
    test_dir = "/root/vyvoj_sw/electron_app/legacy_code"
    if os.path.exists(test_dir):
        print(f"📂 Testing directory: {test_dir}")
        
        # Get all xlsx files
        import glob
        all_xlsx = glob.glob(os.path.join(test_dir, "**", "*.xlsx"), recursive=True)
        print(f"   Found {len(all_xlsx)} Excel files total")
        
        # Test directory processing
        try:
            files = processor._get_files_from_directory(test_dir)
            print(f"   Valid files with 'Přehled' sheet: {len(files)}")
            
            for file_path in files[:3]:  # Show first 3
                print(f"   ✅ {os.path.basename(file_path)}")
                
            if processor.warnings:
                print(f"   Warnings:")
                for warning in processor.warnings[:3]:  # Show first 3
                    print(f"   ⚠️ {warning}")
                    
        except Exception as e:
            print(f"   ❌ Error: {e}")

def test_sheet_validation():
    """Test sheet existence validation"""
    print("\n=== Testing Sheet Validation ===")
    
    processor = ZorSpecDatProcessor()
    
    # Test with known files from inv directory
    test_files = [
        "/root/vyvoj_sw/electron_app/legacy_code/inv/16_hodin_inovativniho_vzdelavani_sablona.xlsx",
        "/root/vyvoj_sw/electron_app/legacy_code/inv/source.xlsx"
    ]
    
    for file_path in test_files:
        if os.path.exists(file_path):
            has_sheet = processor._sheet_exists(file_path)
            filename = os.path.basename(file_path)
            status = "✅" if has_sheet else "❌"
            print(f"   {status} {filename}: {'Has' if has_sheet else 'Missing'} 'Přehled' sheet")

def test_name_hashing():
    """Test custom name hashing function"""
    print("\n=== Testing Name Hashing ===")
    
    processor = ZorSpecDatProcessor()
    
    test_names = [
        "Jan Novák",
        "Novák Jan",  # Should have same hash
        "jan novák",  # Should have same hash
        "Marie Svobodová",
        "Petr Dvořák",
        "",  # Empty string
        None  # None value
    ]
    
    hashes = {}
    for name in test_names:
        hash_val = processor._custom_hash(name)
        if hash_val in hashes:
            print(f"   🔄 '{name}' has same hash as '{hashes[hash_val]}': {hash_val}")
        else:
            hashes[hash_val] = name
            print(f"   🔢 '{name}': {hash_val}")

def create_test_data():
    """Create test Excel file for validation"""
    print("\n=== Creating Test Data ===")
    
    try:
        import pandas as pd
        from openpyxl import Workbook
        
        # Create test data
        test_data = {
            'ca': [1, 2, 3, 4, 5],
            'jmena': ['Jan Novák', 'Marie Svobodová', 'Petr Dvořák', 'Jana Krásná', 'Pavel Horný'],
            'datum': ['15.09.2024', '20.09.2024', '25.09.2024', '30.09.2024', '05.10.2024'],
            'pocet_hodin': [2, 3, 2, 1, 4],
            'forma': ['prezenční', 'online', 'prezenční', 'blended', 'online'],
            'tema': ['čtenářská pre/gramotnost', 'matematická pre/gramotnost', 
                    'cizí jazyky/komunikace v cizím jazyce', 'inkluze včetně primární prevence',
                    'přírodovědné a technické vzdělávání']
        }
        
        df = pd.DataFrame(test_data)
        
        # Create temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            temp_path = tmp_file.name
            
        # Create workbook with multiple sheets
        wb = Workbook()
        
        # First sheet (for template name)
        ws1 = wb.active
        ws1.title = "Úvod a postup vyplňování"
        ws1['C4'] = "Test Template Name"
        
        # Create Přehled sheet
        ws2 = wb.create_sheet("Přehled")
        
        # Add headers
        headers = ['', '', 'CA', 'Jména', 'Datum', 'Počet hodin', 'Forma', 'Téma']
        for col, header in enumerate(headers, 1):
            ws2.cell(row=1, column=col, value=header)
            
        # Add data
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            ws2.cell(row=row_idx, column=3, value=row['ca'])
            ws2.cell(row=row_idx, column=4, value=row['jmena'])
            ws2.cell(row=row_idx, column=5, value=row['datum'])
            ws2.cell(row=row_idx, column=6, value=row['pocet_hodin'])
            ws2.cell(row=row_idx, column=7, value=row['forma'])
            ws2.cell(row=row_idx, column=8, value=row['tema'])
            
        wb.save(temp_path)
        print(f"✅ Test file created: {temp_path}")
        
        return temp_path
        
    except Exception as e:
        print(f"❌ Failed to create test data: {e}")
        return None

def test_processing_pipeline():
    """Test complete processing pipeline with test data"""
    print("\n=== Testing Processing Pipeline ===")
    
    # Create test file
    test_file = create_test_data()
    if not test_file:
        print("❌ Cannot test processing without test data")
        return
        
    try:
        processor = ZorSpecDatProcessor()
        
        with tempfile.TemporaryDirectory() as temp_dir:
            print(f"📂 Output directory: {temp_dir}")
            
            # Test processing
            options = {
                'output_dir': temp_dir
            }
            
            result = processor.process([test_file], options)
            
            print(f"🎯 Processing result:")
            print(f"   Success: {'✅' if result['success'] else '❌'}")
            
            if result.get('data'):
                data = result['data']
                print(f"   Files processed: {data.get('files_processed', 0)}")
                print(f"   Unique students: {data.get('unique_students', 0)}")
                
            # Show messages
            if processor.info_messages:
                print(f"📋 Info messages:")
                for info in processor.info_messages:
                    print(f"   ℹ️ {info}")
                    
            if processor.warnings:
                print(f"⚠️ Warnings:")
                for warning in processor.warnings:
                    print(f"   ⚠️ {warning}")
                    
            if processor.errors:
                print(f"❌ Errors:")
                for error in processor.errors:
                    print(f"   ❌ {error}")
                    
            # Check output files
            if result.get('data', {}).get('output_files'):
                print(f"📁 Output files:")
                for output_file in result['data']['output_files']:
                    if os.path.exists(output_file):
                        size = os.path.getsize(output_file)
                        print(f"   ✅ {os.path.basename(output_file)} ({size} bytes)")
                    else:
                        print(f"   ❌ {os.path.basename(output_file)} (missing)")
                        
    except Exception as e:
        print(f"❌ Processing failed: {e}")
        import traceback
        traceback.print_exc()
    finally:
        # Clean up test file
        if os.path.exists(test_file):
            os.unlink(test_file)
            print(f"🧹 Cleaned up test file")

if __name__ == "__main__":
    test_processor_initialization()
    test_validation()
    test_file_discovery() 
    test_sheet_validation()
    test_name_hashing()
    test_processing_pipeline()