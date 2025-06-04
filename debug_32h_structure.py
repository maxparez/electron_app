#!/usr/bin/env python3
"""
Debug script to verify row structure for 32h format
"""

import pandas as pd
import openpyxl

file_path = "tests/regression/inputs/32h/valid_basic.xlsx"

print("=== OPENPYXL (1-based indices) ===")
wb = openpyxl.load_workbook(file_path)
sheet = wb['zdroj-dochazka']

# Print relevant rows with 1-based indices
for row in range(1, 15):
    print(f"\nRow {row}:")
    for col in range(1, 5):  # Columns A-D
        cell_value = sheet.cell(row=row, column=col).value
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"  {col_letter}: {repr(cell_value)}")
wb.close()

print("\n\n=== PANDAS (0-based indices) ===")
df = pd.read_excel(file_path, sheet_name='zdroj-dochazka')

# Print relevant rows with 0-based indices
for idx in range(0, 13):
    print(f"\nPandas index {idx} (Excel row {idx + 2}):")
    for col in range(0, 4):  # First 4 columns
        if col < len(df.columns):
            value = df.iloc[idx, col]
            print(f"  Col{col}: {repr(value)}")