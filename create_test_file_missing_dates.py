#!/usr/bin/env python3
"""
Create a test 32-hour file with missing dates to test error display
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
import os

def create_test_file():
    """Create a test file with some missing dates"""
    
    wb = Workbook()
    
    # Create zdroj-dochazka sheet
    ws = wb.active
    ws.title = "zdroj-dochazka"
    
    # Headers
    ws['B6'] = 'datum aktivity'
    ws['B7'] = 'Forma výuky'
    ws['B8'] = 'Téma výuky'
    ws['B9'] = 'Jméno pedagoga'
    ws['B10'] = 'počet hodin'
    
    # Activity data with some missing dates
    # Activity 1 - has date
    ws['C6'] = '10.09.2024'
    ws['C7'] = 'Zážitková pedagogika'
    ws['C8'] = 'Přírodovědné a technické vzdělávání'
    ws['C9'] = 'Imrich'
    ws['C10'] = 1
    
    # Activity 2 - MISSING DATE
    ws['D6'] = None  # Missing date!
    ws['D7'] = 'Propojování formálního a neformálního vzdělávání'
    ws['D8'] = 'Přírodovědné a technické vzdělávání'
    ws['D9'] = 'Polák'
    ws['D10'] = 1
    
    # Activity 3 - has date
    ws['E6'] = '10.09.2024'
    ws['E7'] = 'Aktivizující metody'
    ws['E8'] = 'Ctení a těžba'
    ws['E9'] = 'Imrich'
    ws['E10'] = 1
    
    # Student names starting from row 11
    students = [
        'Borůvka Albert',
        'Čech Martin',
        'Dvořák Jan',
        'Eliáš Petr',
        'Filip Karel'
    ]
    
    for i, student in enumerate(students):
        ws.cell(row=11+i, column=2, value=student)
        # Add attendance marks
        ws.cell(row=11+i, column=3, value='ano')  # Activity 1
        ws.cell(row=11+i, column=4, value='ano')  # Activity 2
        ws.cell(row=11+i, column=5, value='ano')  # Activity 3
    
    # Save file
    output_path = "/root/vyvoj_sw/electron_app/test_32h_missing_dates.xlsx"
    wb.save(output_path)
    wb.close()
    
    print(f"Test file created: {output_path}")
    print("Features:")
    print("- 3 activities (32-hour format)")
    print("- Activity 2 has MISSING DATE in cell D6")
    print("- 5 students")
    
    return output_path

if __name__ == "__main__":
    create_test_file()