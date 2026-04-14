from __future__ import annotations

import os
import re
import unicodedata
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List
from xml.etree import ElementTree

import pandas as pd
from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook

from .base_tool import BaseTool


SUPPORTED_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
TEMP_FILE_PREFIX = "~$"
WORKBOOK_NS = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
CANDIDATE_SHEETS = {"podpory", "evidence"}


@dataclass(frozen=True)
class WorkbookMatch:
    file_path: Path
    relative_path: str
    sheet_name: str
    header_row: int
    project_number: str
    report_number: int | None
    participant_count: int
    modified_time: float


class DvppReportProcessor(BaseTool):
    """Scan project workbooks and generate a DVPP HTML summary report."""

    def validate_inputs(self, files: List[str], options: Dict[str, Any]) -> bool:
        self.clear_messages()

        project_dir = options.get("project_dir")
        if not project_dir:
            self.add_error("Nebyl zadán projektový adresář")
            return False

        project_path = self.normalize_input_path(str(project_dir))
        if not project_path.is_dir():
            self.add_error(f"Projektový adresář neexistuje: {project_dir}")
            return False

        if not files:
            self.add_error("Nebyl vybrán žádný DVPP soubor ke zpracování")
            return False

        for file_path in files:
            normalized = self.normalize_input_path(str(file_path))
            if not normalized.is_file():
                self.add_error(f"Soubor neexistuje: {file_path}")
                return False

        return True

    def process(self, files: List[str], options: Dict[str, Any]) -> Dict[str, Any]:
        if not self.validate_inputs(files, options):
            return self.get_result(False)

        project_dir = self.normalize_input_path(str(options["project_dir"]))
        selected_matches = self._build_matches_from_paths(project_dir, files)

        if not selected_matches:
            self.add_error("Vybrané soubory neodpovídají DVPP struktuře")
            return self.get_result(False)

        report_filename = f"{project_dir.name}_dvpp_report.html"
        report_path = project_dir / report_filename
        title = f"Souhrnný report podpory DVPP - {project_dir.name}"

        unique_persons = self.render_html_report(selected_matches, title, report_path)

        processed_data = {
            "files_processed": len(selected_matches),
            "unique_participants": unique_persons,
            "report_path": str(report_path),
            "report_filename": report_filename,
            "selected_files": [match.relative_path for match in selected_matches],
            "report_numbers": [match.report_number for match in selected_matches if match.report_number is not None],
        }

        self.add_info(f"DVPP report uložen: {report_path}||{report_filename}")
        return self.get_result(True, processed_data)

    def scan_project_directory(self, project_dir: str | Path) -> List[Dict[str, Any]]:
        root = self.normalize_input_path(str(project_dir))
        matches: List[WorkbookMatch] = []

        for file_path in self.iter_excel_files(root):
            if not self.has_candidate_sheet_name(file_path):
                continue
            try:
                match = self.inspect_workbook(file_path, root)
            except Exception as exc:
                self.add_warning(f"Soubor {file_path.name} nelze zpracovat: {exc}")
                continue
            if match is not None:
                matches.append(match)

        matches.sort(key=lambda item: item.relative_path.lower())
        return [self._serialize_match(match) for match in matches]

    def normalize_input_path(self, path_str: str) -> Path:
        windows_match = re.match(r"^([a-zA-Z]):[\\/](.*)$", path_str)
        if windows_match:
            drive = windows_match.group(1).lower()
            tail = windows_match.group(2).replace("\\", "/")
            return Path(f"/mnt/{drive}/{tail}")
        return Path(path_str)

    def normalize_text(self, value: object) -> str:
        if value is None:
            return ""
        text = str(value).strip().lower()
        text = unicodedata.normalize("NFKD", text)
        text = "".join(char for char in text if not unicodedata.combining(char))
        text = re.sub(r"\s+", " ", text)
        return text

    def iter_excel_files(self, project_dir: Path) -> List[Path]:
        files: List[Path] = []
        for path in project_dir.rglob("*"):
            if not path.is_file():
                continue
            if path.name.startswith(TEMP_FILE_PREFIX):
                continue
            if path.suffix.lower() in SUPPORTED_SUFFIXES:
                files.append(path)
        return sorted(files)

    def has_candidate_sheet_name(self, file_path: Path) -> bool:
        try:
            with zipfile.ZipFile(file_path) as archive:
                workbook_xml = archive.read("xl/workbook.xml")
        except (KeyError, OSError, zipfile.BadZipFile):
            return True

        root = ElementTree.fromstring(workbook_xml)
        for sheet in root.findall("main:sheets/main:sheet", WORKBOOK_NS):
            name = self.normalize_text(sheet.attrib.get("name", ""))
            if name in CANDIDATE_SHEETS:
                return True
        return False

    def find_header_mapping(self, sheet) -> tuple[int, dict[str, int]] | None:
        for row_index in range(1, min(sheet.max_row, 40) + 1):
            headers: dict[str, int] = {}
            for col_index in range(1, min(sheet.max_column, 25) + 1):
                cell_text = self.normalize_text(sheet.cell(row_index, col_index).value)
                if not cell_text:
                    continue
                if cell_text == "jmeno":
                    headers["first_name"] = col_index
                elif cell_text == "prijmeni":
                    headers["last_name"] = col_index
                elif "pocet hodin podpory" in cell_text:
                    headers["support_hours"] = col_index
                elif "datum narozeni" in cell_text or "datum nar" in cell_text:
                    headers["birth_date"] = col_index
            if {"first_name", "last_name", "support_hours"}.issubset(headers):
                return row_index, headers
        return None

    def extract_metadata(self, sheet) -> tuple[str, int | None]:
        project_number = ""
        report_number = None

        for row_index in range(1, min(sheet.max_row, 15) + 1):
            for col_index in range(1, min(sheet.max_column, 15) + 1):
                cell_text = self.normalize_text(sheet.cell(row_index, col_index).value)
                if not cell_text:
                    continue

                if "registracni cislo projektu" in cell_text and col_index + 2 <= sheet.max_column:
                    raw_value = sheet.cell(row_index, col_index + 2).value
                    if raw_value:
                        project_number = str(raw_value).strip()

                if "zprava o realizaci c." in cell_text and col_index + 1 <= sheet.max_column:
                    raw_value = sheet.cell(row_index, col_index + 1).value
                    if raw_value is not None:
                        try:
                            report_number = int(raw_value)
                        except (TypeError, ValueError):
                            report_number = None

        return project_number, report_number

    def count_participant_rows(self, sheet, header_row: int, headers: dict[str, int]) -> int:
        count = 0
        for row_index in range(header_row + 1, sheet.max_row + 1):
            first_name = sheet.cell(row_index, headers["first_name"]).value
            last_name = sheet.cell(row_index, headers["last_name"]).value
            if first_name and last_name:
                count += 1
        return count

    def inspect_workbook(self, file_path: Path, project_dir: Path) -> WorkbookMatch | None:
        workbook = load_workbook(file_path, data_only=True, read_only=True)
        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                header_info = self.find_header_mapping(sheet)
                if header_info is None:
                    continue

                header_row, headers = header_info
                project_number, report_number = self.extract_metadata(sheet)
                participant_count = self.count_participant_rows(sheet, header_row, headers)
                return WorkbookMatch(
                    file_path=file_path,
                    relative_path=file_path.relative_to(project_dir).as_posix(),
                    sheet_name=sheet_name,
                    header_row=header_row,
                    project_number=project_number,
                    report_number=report_number,
                    participant_count=participant_count,
                    modified_time=file_path.stat().st_mtime,
                )
        finally:
            workbook.close()
        return None

    def extract_participants_from_match(self, match: WorkbookMatch) -> List[Dict[str, Any]]:
        workbook = load_workbook(match.file_path, data_only=True, read_only=True)
        try:
            sheet = workbook[match.sheet_name]
            header_info = self.find_header_mapping(sheet)
            if header_info is None:
                return []
            header_row, headers = header_info

            participants: List[Dict[str, Any]] = []
            for row_index in range(header_row + 1, sheet.max_row + 1):
                first_name = sheet.cell(row_index, headers["first_name"]).value
                last_name = sheet.cell(row_index, headers["last_name"]).value
                if not first_name or not last_name:
                    continue

                raw_hours = sheet.cell(row_index, headers["support_hours"]).value
                try:
                    support_hours = float(raw_hours or 0)
                except (TypeError, ValueError):
                    support_hours = 0.0

                participants.append(
                    {
                        "first_name": str(first_name).strip(),
                        "last_name": str(last_name).strip(),
                        "support_hours": support_hours,
                    }
                )

            return participants
        finally:
            workbook.close()

    def aggregate_participants(self, matches: List[WorkbookMatch]) -> List[Dict[str, Any]]:
        aggregated: Dict[tuple[str, str], Dict[str, Any]] = {}

        for match in matches:
            for participant in self.extract_participants_from_match(match):
                key = (
                    participant["first_name"].strip().lower(),
                    participant["last_name"].strip().lower(),
                )
                if key not in aggregated:
                    aggregated[key] = {
                        "first_name": participant["first_name"],
                        "last_name": participant["last_name"],
                        "support_hours": 0.0,
                    }
                aggregated[key]["support_hours"] += participant["support_hours"]

        rows = list(aggregated.values())
        rows.sort(key=lambda row: (str(row["last_name"]).lower(), str(row["first_name"]).lower()))
        return rows

    def build_combined_dataframe(self, matches: List[WorkbookMatch]) -> pd.DataFrame:
        frames: List[pd.DataFrame] = []
        for match in matches:
            frame = pd.read_excel(match.file_path, sheet_name=match.sheet_name, header=match.header_row - 1)
            frames.append(frame)
        if not frames:
            return pd.DataFrame()
        return pd.concat(frames, ignore_index=True)

    def build_summary_tables(self, matches: List[WorkbookMatch]) -> tuple[pd.DataFrame | None, pd.DataFrame | None, List[str]]:
        full_combined_df = self.build_combined_dataframe(matches)
        if full_combined_df.empty:
            return None, None, []

        sablona_col = "Šablona"
        tema_long_col = (
            'Téma\nZvolíte-li ve sloupci „Forma“ kteroukoliv z variant „mentoring“, '
            '„supervize“ „koučink“ nebo „kvalifikační studium DVPP“ nechte pole Téma prázdné.'
        )
        jmeno_col = "Jméno"
        prijmeni_col = "Příjmení"
        hodiny_col = "Počet  hodin podpory"

        final_themes_df = None
        if sablona_col in full_combined_df.columns and tema_long_col in full_combined_df.columns:
            df_themes = full_combined_df[[sablona_col, tema_long_col]].copy().rename(columns={tema_long_col: "Téma"})
            df_themes.dropna(how="all", inplace=True)
            df_themes["Šablona"] = df_themes["Šablona"].astype(str).str.strip()
            df_themes["Téma"] = df_themes["Téma"].astype(str).str.strip()
            df_themes.drop_duplicates(inplace=True)
            grouped_themes = df_themes.groupby("Šablona")["Téma"].apply(list).to_dict()
            final_themes_df = pd.DataFrame(dict([(key, pd.Series(value)) for key, value in grouped_themes.items()]))

        person_hours_df = None
        required_cols = [jmeno_col, prijmeni_col, hodiny_col]
        if all(col in full_combined_df.columns for col in required_cols):
            df_persons = full_combined_df[required_cols].copy()
            df_persons.dropna(subset=[jmeno_col, prijmeni_col], how="all", inplace=True)
            df_persons[hodiny_col] = pd.to_numeric(df_persons[hodiny_col], errors="coerce").fillna(0)
            df_persons["Celé Jméno"] = (
                df_persons[jmeno_col].astype(str).str.strip() + " " + df_persons[prijmeni_col].astype(str).str.strip()
            ).str.strip()
            person_hours_df = df_persons.groupby("Celé Jméno", as_index=False)[hodiny_col].sum()
            person_hours_df.rename(columns={hodiny_col: "Celkem hodin"}, inplace=True)
            person_hours_df = person_hours_df.sort_values(by="Celkem hodin", ascending=False)

        names_list = person_hours_df["Celé Jméno"].tolist() if person_hours_df is not None else []
        return final_themes_df, person_hours_df, names_list

    def render_html_report(self, matches: List[WorkbookMatch], title: str, output_path: Path) -> int:
        final_themes_df, person_hours_df, names_list = self.build_summary_tables(matches)

        templates_dir = Path(__file__).resolve().parents[1] / "templates"
        env = Environment(loader=FileSystemLoader(str(templates_dir)))
        template = env.get_template("dvpp_report_template.html")

        themes_html = (
            final_themes_df.to_html(classes="table table-bordered table-striped", na_rep="", justify="left")
            if final_themes_df is not None
            else "<p>Tabulku témat se nepodařilo vygenerovat.</p>"
        )
        persons_html = (
            person_hours_df.to_html(index=False, classes="table table-hover table-striped", na_rep="")
            if person_hours_df is not None
            else "<p>Tabulku hodin se nepodařilo vygenerovat.</p>"
        )

        html_content = template.render(
            title=title,
            table1=themes_html,
            table2=persons_html,
            unique_persons_count=len(person_hours_df) if person_hours_df is not None else 0,
            names_list=names_list,
        )

        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(html_content, encoding="utf-8")
        return len(person_hours_df) if person_hours_df is not None else 0

    def _build_matches_from_paths(self, project_dir: Path, files: List[str]) -> List[WorkbookMatch]:
        matches: List[WorkbookMatch] = []
        for file_path in files:
            normalized = self.normalize_input_path(str(file_path))
            match = self.inspect_workbook(normalized, project_dir)
            if match is not None:
                matches.append(match)
        matches.sort(key=lambda item: item.relative_path.lower())
        return matches

    def _serialize_match(self, match: WorkbookMatch) -> Dict[str, Any]:
        return {
            "file_path": str(match.file_path),
            "relative_path": match.relative_path,
            "sheet_name": match.sheet_name,
            "project_number": match.project_number,
            "report_number": match.report_number,
            "participant_count": match.participant_count,
        }
