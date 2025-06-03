"""
Validation utilities for tools
"""

import os
import pandas as pd
from typing import List, Dict, Any, Optional, Tuple, Union
from pathlib import Path
from openpyxl import load_workbook


class FileValidator:
    """Validator for file operations"""
    
    @staticmethod
    def validate_file_exists(file_path: str) -> Tuple[bool, Optional[str]]:
        """
        Validate that file exists
        
        Returns:
            Tuple of (is_valid, error_message)
        """
        if not file_path:
            return False, "No file path provided"
            
        if not os.path.exists(file_path):
            return False, f"File not found: {file_path}"
            
        if not os.path.isfile(file_path):
            return False, f"Path is not a file: {file_path}"
            
        return True, None
        
    @staticmethod
    def validate_directory_exists(dir_path: str) -> Tuple[bool, Optional[str]]:
        """
        Validate that directory exists
        
        Returns:
            Tuple of (is_valid, error_message)
        """
        if not dir_path:
            return False, "No directory path provided"
            
        if not os.path.exists(dir_path):
            return False, f"Directory not found: {dir_path}"
            
        if not os.path.isdir(dir_path):
            return False, f"Path is not a directory: {dir_path}"
            
        return True, None
        
    @staticmethod
    def validate_file_extension(file_path: str, 
                              allowed_extensions: List[str]) -> Tuple[bool, Optional[str]]:
        """
        Validate file extension
        
        Args:
            file_path: Path to file
            allowed_extensions: List of allowed extensions (e.g., ['.xlsx', '.xls'])
            
        Returns:
            Tuple of (is_valid, error_message)
        """
        if not file_path:
            return False, "No file path provided"
            
        ext = Path(file_path).suffix.lower()
        
        if ext not in allowed_extensions:
            return False, f"Invalid file extension: {ext}. Allowed: {allowed_extensions}"
            
        return True, None
        
    @staticmethod
    def validate_file_readable(file_path: str) -> Tuple[bool, Optional[str]]:
        """
        Validate that file is readable
        
        Returns:
            Tuple of (is_valid, error_message)
        """
        try:
            with open(file_path, 'rb') as f:
                # Try to read first few bytes
                f.read(1024)
            return True, None
        except PermissionError:
            return False, f"Permission denied: {file_path}"
        except Exception as e:
            return False, f"Cannot read file: {str(e)}"


class ExcelValidator:
    """Validator for Excel files"""
    
    @staticmethod
    def validate_excel_file(file_path: str) -> Tuple[bool, Optional[str]]:
        """
        Validate that file is a valid Excel file
        
        Returns:
            Tuple of (is_valid, error_message)
        """
        # First check file extension
        valid, msg = FileValidator.validate_file_extension(
            file_path, ['.xlsx', '.xls', '.xlsm']
        )
        if not valid:
            return False, msg
            
        # Try to open with openpyxl
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            wb.close()
            return True, None
        except Exception as e:
            return False, f"Invalid Excel file: {str(e)}"
            
    @staticmethod
    def validate_sheet_exists(file_path: str, sheet_name: str) -> Tuple[bool, Optional[str]]:
        """
        Validate that sheet exists in Excel file
        
        Returns:
            Tuple of (is_valid, error_message)
        """
        try:
            wb = load_workbook(file_path, read_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return False, f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            wb.close()
            return True, None
        except Exception as e:
            return False, f"Error checking sheet: {str(e)}"
            
    @staticmethod
    def validate_cell_value(file_path: str, sheet_name: str, 
                          cell_ref: str, expected_value: Optional[str] = None) -> Tuple[bool, Optional[str]]:
        """
        Validate cell value in Excel file
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name
            cell_ref: Cell reference (e.g., 'B1')
            expected_value: Expected value (if None, just checks if cell has value)
            
        Returns:
            Tuple of (is_valid, error_message)
        """
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            if sheet_name not in wb.sheetnames:
                wb.close()
                return False, f"Sheet '{sheet_name}' not found"
                
            value = wb[sheet_name][cell_ref].value
            wb.close()
            
            if expected_value is not None:
                if str(value).strip() != str(expected_value).strip():
                    return False, f"Cell {cell_ref} value mismatch. Expected: '{expected_value}', Got: '{value}'"
            else:
                if value is None or str(value).strip() == '':
                    return False, f"Cell {cell_ref} is empty"
                    
            return True, None
            
        except Exception as e:
            return False, f"Error reading cell: {str(e)}"


class DataValidator:
    """Validator for data structures"""
    
    @staticmethod
    def validate_dataframe(df: pd.DataFrame, 
                         min_rows: Optional[int] = None,
                         min_cols: Optional[int] = None,
                         required_columns: Optional[List[str]] = None) -> Tuple[bool, Optional[str]]:
        """
        Validate DataFrame structure
        
        Returns:
            Tuple of (is_valid, error_message)
        """
        if df is None:
            return False, "DataFrame is None"
            
        if df.empty:
            return False, "DataFrame is empty"
            
        if min_rows is not None and len(df) < min_rows:
            return False, f"Not enough rows. Expected at least {min_rows}, got {len(df)}"
            
        if min_cols is not None and len(df.columns) < min_cols:
            return False, f"Not enough columns. Expected at least {min_cols}, got {len(df.columns)}"
            
        if required_columns:
            missing = set(required_columns) - set(df.columns)
            if missing:
                return False, f"Missing required columns: {missing}"
                
        return True, None
        
    @staticmethod
    def validate_numeric_range(value: Union[int, float], 
                             min_value: Optional[Union[int, float]] = None,
                             max_value: Optional[Union[int, float]] = None,
                             field_name: str = "Value") -> Tuple[bool, Optional[str]]:
        """
        Validate numeric value is within range
        
        Returns:
            Tuple of (is_valid, error_message)
        """
        try:
            num_value = float(value)
        except (TypeError, ValueError):
            return False, f"{field_name} is not a valid number: {value}"
            
        if min_value is not None and num_value < min_value:
            return False, f"{field_name} too small: {num_value} < {min_value}"
            
        if max_value is not None and num_value > max_value:
            return False, f"{field_name} too large: {num_value} > {max_value}"
            
        return True, None
        
    @staticmethod
    def validate_percentage(value: Union[int, float], 
                          field_name: str = "Percentage") -> Tuple[bool, Optional[str]]:
        """
        Validate percentage value (0-100)
        
        Returns:
            Tuple of (is_valid, error_message)
        """
        return DataValidator.validate_numeric_range(value, 0, 100, field_name)