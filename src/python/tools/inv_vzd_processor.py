"""
InvVzd Processor - Process innovative education attendance files

DEBUG FINDINGS (2025-05-30):
- Files are loading correctly when paths are valid
- Processing correctly calculates total hours (e.g., 57 hours, not 100%)
- The tool requires Windows with MS Excel due to xlwings dependency
- Enhanced logging with [INVVZD] prefix shows all processing steps
- Platform check added to provide clear error message on non-Windows systems
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import warnings
from typing import Dict, Any, List, Optional, Tuple
import platform
try:
    import xlwings as xw
    XLWINGS_AVAILABLE = True
except ImportError:
    XLWINGS_AVAILABLE = False
from datetime import datetime
import unicodedata
import re

from .base_tool import BaseTool

warnings.filterwarnings('ignore')

# Version-specific constants
VERSIONS = {
    "32": {
        "hours": 32,
        "template": {
            "B1": "32 hodin"  # Simplified to match actual template
        },
        "source": {
            "B6": "datum aktivity",
            "B7": "Forma výuky",
            "sheet": "List1"
        },
        "output_prefix": "32_hodin_inovativniho_vzdelavani",
        "short_prefix": "32_inv",
        "data_start_row": 11,
        "data_start_col": 2,  # Column B
        "hours_row": 10,
        "name_col": 2,  # Column B
        "attendance_start_col": 3,  # Column C
        "skiprows": 9,  # Skip first 9 rows in template
        "hours_total_cell": "B10"  # Cell for total hours
    },
    "16": {
        "hours": 16,
        "template": {
            "B1": "16 hodin"  # Simplified to match actual template
        },
        "source": {
            "B2": "Pořadové číslo aktivity",
            "sheet": "Seznam aktivit"
        },
        "output_prefix": "16_hodin_inovativniho_vzdelavani",
        "short_prefix": "16_inv",
        "data_start_row": 3,
        "data_start_col": 2,  # Column B
        "column_mapping": {
            "datum": "C",
            "cas": "D", 
            "hodin": "E",
            "forma": "F",
            "tema": "G",
            "ucitel": "H"
        }
    }
}


class InvVzdProcessor(BaseTool):
    """Processor for innovative education attendance (16/32 hours)"""
    
    def __init__(self, version=None, logger=None):
        super().__init__(logger)
        self.hours_total = 0
        self.version = version
        self.config = VERSIONS.get(version) if version else None
        
    def validate_inputs(self, files: List[str], options: Dict[str, Any]) -> bool:
        """Validate input files and options"""
        self.clear_messages()
        self.logger.info(f"[INVVZD] === VALIDATE INPUTS START ===")
        self.logger.info(f"[INVVZD] Validating inputs: {len(files)} files")
        self.logger.info(f"[INVVZD] Files: {files}")
        self.logger.info(f"[INVVZD] Options: {options}")
        
        # Check if files provided
        if not files:
            self.add_error("Žádné soubory nebyly poskytnuty")
            self.logger.error("[INVVZD] ERROR: No files provided")
            return False
            
        # Check if template provided
        template = options.get('template')
        self.logger.info(f"[INVVZD] Template path: {template}")
        if not template:
            self.add_error("Šablona nebyla poskytnuta")
            self.logger.error("[INVVZD] ERROR: No template provided")
            return False
            
        # Check if template exists
        self.logger.info(f"[INVVZD] Checking if template exists: {template}")
        template_exists = self.file_exists(template)
        self.logger.info(f"[INVVZD] Template exists: {template_exists}")
        if not template_exists:
            self.add_error(f"Šablona neexistuje: {template}")
            self.logger.error(f"[INVVZD] ERROR: Template does not exist: {template}")
            return False
            
        # Validate all source files exist
        for file in files:
            self.logger.info(f"[INVVZD] Checking if source file exists: {file}")
            file_exists = self.file_exists(file)
            self.logger.info(f"[INVVZD] File exists: {file_exists}")
            if not file_exists:
                self.add_error(f"Soubor neexistuje: {file}")
                self.logger.error(f"[INVVZD] ERROR: Source file does not exist: {file}")
                return False
                
        # Detect template version
        self.logger.info(f"[INVVZD] Detecting template version...")
        template_version = self._detect_template_version(template)
        self.logger.info(f"[INVVZD] Detected template version: {template_version}")
        if not template_version:
            self.add_error("Nepodařilo se detekovat verzi šablony")
            self.logger.error(f"[INVVZD] ERROR: Failed to detect template version")
            return False
            
        self.version = template_version
        self.config = VERSIONS[template_version]
        # Template version detected - no general message needed
        self.logger.info(f"[INVVZD] === VALIDATE INPUTS SUCCESS ===")
        self.logger.info(f"[INVVZD] Version set to: {self.version}")
        self.logger.info(f"[INVVZD] Config: {self.config}")
        
        return True
        
    def process(self, files: List[str], options: Dict[str, Any]) -> Dict[str, Any]:
        """Process attendance files"""
        self.logger.info(f"[INVVZD] === PROCESS START ===")
        self.logger.info(f"[INVVZD] InvVzdProcessor.process called with {len(files)} files")
        self.logger.info(f"[INVVZD] Files: {files}")
        self.logger.info(f"[INVVZD] Options: {options}")
        
        if not self.validate_inputs(files, options):
            self.logger.error("[INVVZD] ERROR: Input validation failed")
            result = self.get_result(False)
            self.logger.info(f"[INVVZD] Returning validation failure result: {result}")
            return result
            
        try:
            template = options.get('template')
            keep_filename = options.get('keep_filename', True)
            optimize = options.get('optimize', False)
            output_dir = options.get('output_dir', os.path.dirname(files[0]))
            
            results = []
            
            for source_file in files:
                self.logger.info(f"[INVVZD] Processing file: {source_file}")
                
                # Clear messages for this file (keep only the template version message)
                self.clear_file_messages()
                
                # Validate version match
                self.logger.info(f"[INVVZD] Validating version match...")
                version_match = self._validate_version_match(source_file, template)
                self.logger.info(f"[INVVZD] Version match: {version_match}")
                
                if not version_match:
                    self.logger.error(f"[INVVZD] Version mismatch, skipping file")
                    # Still add to results with error status
                    results.append({
                        "source": source_file,
                        "output": None,
                        "hours": 0,
                        "status": "error",
                        "errors": list(self.errors),  # Copy current errors
                        "warnings": list(self.warnings),  # Copy current warnings  
                        "info": list(self.info_messages)  # Copy current info
                    })
                    continue
                    
                # Process the file
                output_file = self._process_single_file(
                    source_file, 
                    template, 
                    output_dir,
                    keep_filename,
                    optimize
                )
                
                # Collect messages for this specific file
                file_info = list(self.info_messages)
                file_warnings = list(self.warnings)
                file_errors = list(self.errors)
                
                if output_file:
                    self.logger.info(f"[INVVZD] File processed successfully: {output_file}")
                    self.logger.info(f"[INVVZD] Total hours: {self.hours_total}")
                    results.append({
                        "source": source_file,
                        "output": output_file,
                        "hours": self.hours_total,
                        "status": "success" if not file_errors else "warning",
                        "errors": file_errors,
                        "warnings": file_warnings,
                        "info": file_info
                    })
                else:
                    self.logger.error(f"[INVVZD] Failed to process file: {source_file}")
                    results.append({
                        "source": source_file,
                        "output": None,
                        "hours": 0,
                        "status": "error",
                        "errors": file_errors,
                        "warnings": file_warnings,
                        "info": file_info
                    })
                    
            # Always return data structure, even if empty
            data = {"processed_files": results}
            
            if results:
                # Overall success - no general message needed
                result = self.get_result(True, data)
                self.logger.info(f"[INVVZD] === PROCESS SUCCESS ===")
                self.logger.info(f"[INVVZD] Returning success result: {result}")
            else:
                # No files succeeded, but return data structure with empty results
                result = self.get_result(False, data)
                self.logger.error(f"[INVVZD] === PROCESS FAILED - No results ===")
                self.logger.info(f"[INVVZD] Returning failure result with empty data: {result}")
                
            return result
                
        except Exception as e:
            self.logger.error(f"[INVVZD] === PROCESS EXCEPTION ===")
            self.logger.error(f"[INVVZD] Exception: {str(e)}")
            import traceback
            self.logger.error(f"[INVVZD] Traceback: {traceback.format_exc()}")
            self.add_error(f"Chyba při zpracování: {str(e)}")
            result = self.get_result(False)
            self.logger.info(f"[INVVZD] Returning exception result: {result}")
            return result
            
    def _detect_template_version(self, template_path: str) -> Optional[str]:
        """Detect version from template content"""
        try:
            self.logger.info(f"[INVVZD] Loading template workbook: {template_path}")
            wb = load_workbook(template_path, read_only=True)
            sheet = wb.worksheets[0]
            self.logger.info(f"[INVVZD] First sheet name: {sheet.title}")
            
            for version, config in VERSIONS.items():
                self.logger.info(f"[INVVZD] Checking for version {version}...")
                match = True
                for cell, expected_value in config["template"].items():
                    actual_value = sheet[cell].value
                    self.logger.info(f"[INVVZD]   Checking cell {cell}: expected='{expected_value}', actual='{actual_value}'")
                    if expected_value.lower() not in str(actual_value).lower():
                        match = False
                        break
                if match:
                    wb.close()
                    self.logger.info(f"[INVVZD] Template version detected: {version}")
                    return version
                    
            wb.close()
            self.logger.error(f"[INVVZD] No matching template version found")
            return None
        except Exception as e:
            self.logger.error(f"[INVVZD] Exception in template version detection: {str(e)}")
            import traceback
            self.logger.error(f"[INVVZD] Traceback: {traceback.format_exc()}")
            self.add_error(f"Chyba při detekci verze šablony: {str(e)}")
            return None
            
    def _detect_source_version(self, source_file: str) -> Optional[str]:
        """Detect version from source content"""
        self.logger.info(f"[INVVZD] === DETECT SOURCE VERSION START ===")
        self.logger.info(f"[INVVZD] Source file: {source_file}")
        
        try:
            wb = load_workbook(source_file, read_only=True)
            sheet_names = wb.sheetnames
            self.logger.info(f"[INVVZD] Sheet names in file: {sheet_names}")
            
            # Priority 1: Check for "zdroj-dochazka" sheet (preferred format)
            if "zdroj-dochazka" in sheet_names:
                self.logger.info(f"[INVVZD] Found 'zdroj-dochazka' sheet")
                sheet = wb["zdroj-dochazka"]
                b7_value = sheet["B7"].value
                self.logger.info(f"[INVVZD] B7 value in zdroj-dochazka: '{b7_value}'")
                
                # If B7 contains "čas zahájení" then it's 16h version
                if b7_value and "čas zahájení" in str(b7_value).lower():
                    wb.close()
                    self.logger.info(f"[INVVZD] Detected version: 16h (found 'čas zahájení' in B7)")
                    return "16"
                else:
                    # If zdroj-dochazka exists but no "čas zahájení" in B7, it's 32h
                    wb.close()
                    self.logger.info(f"[INVVZD] Detected version: 32h (zdroj-dochazka exists but no 'čas zahájení')") 
                    return "32"
            
            # Priority 2: If no "zdroj-dochazka", use first sheet and check B6/B7
            if sheet_names:
                self.logger.info(f"[INVVZD] No 'zdroj-dochazka' sheet, checking first sheet: {sheet_names[0]}")
                sheet = wb[sheet_names[0]]  # Take first sheet regardless of name
                b6_value = sheet["B6"].value
                b7_value = sheet["B7"].value
                self.logger.info(f"[INVVZD] B6 value: '{b6_value}'")
                self.logger.info(f"[INVVZD] B7 value: '{b7_value}'")
                
                # Check if B6 contains "datum aktivity"
                if b6_value and "datum aktivity" in str(b6_value).lower():
                    self.logger.info(f"[INVVZD] Found 'datum aktivity' in B6")
                    # If B7 contains "čas zahájení" then 16h, otherwise 32h
                    if b7_value and "čas zahájení" in str(b7_value).lower():
                        wb.close()
                        self.logger.info(f"[INVVZD] Detected version: 16h (found 'čas zahájení' in B7)")
                        return "16"
                    else:
                        wb.close()
                        self.logger.info(f"[INVVZD] Detected version: 32h (B6 has 'datum aktivity' but no 'čas zahájení' in B7)")
                        return "32"
            
            # Fallback: Check legacy formats
            self.logger.info(f"[INVVZD] Checking legacy formats...")
            # Check for 16 hour version (complex sheet structure)
            if "Seznam aktivit" in sheet_names:
                self.logger.info(f"[INVVZD] Found 'Seznam aktivit' sheet")
                sheet = wb["Seznam aktivit"]
                b2_value = sheet["B2"].value
                self.logger.info(f"[INVVZD] B2 value in Seznam aktivit: '{b2_value}'")
                
                if b2_value and "pořadové číslo aktivity" in str(b2_value).lower():
                    wb.close()
                    self.logger.info(f"[INVVZD] Detected version: 16h (found 'pořadové číslo aktivity' in Seznam aktivit B2)")
                    return "16"
                    
            wb.close()
            self.logger.warning(f"[INVVZD] Could not detect version - no matching patterns found")
            return None
        except Exception as e:
            self.logger.error(f"[INVVZD] Exception in _detect_source_version: {str(e)}")
            import traceback
            self.logger.error(f"[INVVZD] Traceback: {traceback.format_exc()}")
            self.add_error(f"Chyba při detekci verze zdroje: {str(e)}")
            return None
            
    def _validate_version_match(self, source_file: str, template_path: str) -> bool:
        """Validate that source data version matches template version"""
        source_version = self._detect_source_version(source_file)
        
        if source_version is None:
            self.add_error(f"Nepodařilo se detekovat verzi souboru: {os.path.basename(source_file)}")
            return False
            
        if source_version != self.version:
            source_hours = VERSIONS[source_version]["hours"]
            template_hours = self.config["hours"]
            self.add_error(
                f"Nesoulad verzí: zdrojový soubor má {source_hours} hodin, "
                f"ale šablona je pro {template_hours} hodin"
            )
            return False
            
        return True
        
    def _process_single_file(self, source_file: str, template_path: str, 
                           output_dir: str, keep_filename: bool, 
                           optimize: bool) -> Optional[str]:
        """Process a single attendance file"""
        try:
            self.logger.info(f"[INVVZD] === PROCESS SINGLE FILE START ===")
            self.logger.info(f"[INVVZD] Source: {source_file}")
            self.logger.info(f"[INVVZD] Template: {template_path}")
            self.logger.info(f"[INVVZD] Output dir: {output_dir}")
            # Read source data
            self.logger.info(f"[INVVZD] Reading source data...")
            source_data = self._read_source_data(source_file)
            if source_data is None:
                self.logger.error(f"[INVVZD] Failed to read source data")
                return None
            self.logger.info(f"[INVVZD] Source data read successfully, shape: {source_data.shape}")
                
            # Optimize if requested
            if optimize:
                source_data = self._optimize_data(source_data)
                
            # Create output filename
            output_file = self._create_output_filename(
                source_file, output_dir, keep_filename
            )
            
            # Copy template and fill with data
            self.logger.info(f"[INVVZD] Copying template and filling with data...")
            self._copy_template_with_data(
                template_path, output_file, source_data, source_file
            )
            self.logger.info(f"[INVVZD] Template copied and filled successfully")
            
            self.add_info(f"Vytvořen výstupní soubor: {os.path.basename(output_file)}")
            self.logger.info(f"[INVVZD] === PROCESS SINGLE FILE SUCCESS ===")
            return output_file
            
        except Exception as e:
            self.logger.error(f"[INVVZD] === PROCESS SINGLE FILE EXCEPTION ===")
            self.logger.error(f"[INVVZD] Exception: {str(e)}")
            import traceback
            self.logger.error(f"[INVVZD] Traceback: {traceback.format_exc()}")
            self.add_error(f"Chyba při zpracování souboru {os.path.basename(source_file)}: {str(e)}")
            return None
            
    def _read_source_data(self, source_file: str) -> Optional[pd.DataFrame]:
        """Read and process source data"""
        try:
            self.logger.info(f"[INVVZD] Reading source data for version: {self.version}")
            if self.version == "16":
                return self._read_16_hour_data(source_file)
            elif self.version == "32":
                return self._read_32_hour_data(source_file)
            else:
                self.add_error(f"Nepodporovaná verze: {self.version}")
                self.logger.error(f"[INVVZD] Unsupported version: {self.version}")
                return None
                
        except Exception as e:
            self.add_error(f"Chyba při čtení zdrojových dat: {str(e)}")
            return None
            
    def _read_16_hour_data(self, source_file: str) -> Optional[pd.DataFrame]:
        """Read data from 16 hour source file (zdroj-dochazka sheet)"""
        try:
            wb = load_workbook(source_file)
            
            # Find the correct sheet - prefer zdroj-dochazka
            sheet_name = None
            if "zdroj-dochazka" in wb.sheetnames:
                sheet_name = "zdroj-dochazka"
            elif "List1" in wb.sheetnames:
                sheet_name = "List1"
            else:
                sheet_name = wb.sheetnames[0]
                
            sheet = wb[sheet_name]
            self.add_info(f"Čtu 16h data z listu: {sheet_name}")
            
            # 16h format structure (from actual file inspection):
            # Row 6: dates (datum aktivity)
            # Row 7: times (čas zahájení) - specific to 16h
            # Row 8: forms (forma výuky)
            # Row 9: topics (téma výuky)
            # Row 10: teachers (jméno pedagoga)
            # Row 11: hours (počet hodin)
            
            data = []
            col = 3  # Start from column C (first activity)
            
            # Debug: Check the first few cells
            self.logger.info(f"[INVVZD] 16h Debug - Starting to read from column {col}")
            for test_col in range(3, 6):  # Check columns C, D, E
                hours_test = sheet.cell(row=11, column=test_col).value
                self.logger.info(f"[INVVZD] 16h Debug - Row 11, Column {test_col}: {repr(hours_test)} (type: {type(hours_test)})")
            
            while True:
                # Check if there's data in this column by checking hours (row 11 for 16h)
                hours_cell = sheet.cell(row=11, column=col).value
                self.logger.info(f"[INVVZD] 16h Debug - Checking column {col}, hours_cell: {repr(hours_cell)}")
                
                if hours_cell is None or str(hours_cell).strip() == '':
                    self.logger.info(f"[INVVZD] 16h Debug - Hours cell is empty, breaking at column {col}")
                    break
                    
                try:
                    hours = int(float(str(hours_cell)))
                    self.logger.info(f"[INVVZD] 16h Debug - Converted hours: {hours}")
                    if hours <= 0:
                        self.logger.info(f"[INVVZD] 16h Debug - Hours <= 0, breaking")
                        break
                except (ValueError, TypeError) as e:
                    self.logger.info(f"[INVVZD] 16h Debug - Failed to convert hours: {e}")
                    break
                
                # Get date (row 6)
                date_cell = sheet.cell(row=6, column=col).value
                self.logger.info(f"[INVVZD] 16h Debug - Date cell: {repr(date_cell)}")
                
                if date_cell:
                    if hasattr(date_cell, 'strftime'):
                        datum = date_cell.strftime('%d.%m.%Y')
                    else:
                        datum = str(date_cell).strip()
                else:
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(col)
                    self.add_error(f"Chybí datum aktivity v buňce {col_letter}6")
                    datum = None
                
                # Get time (row 7) - specific to 16h
                time_cell = sheet.cell(row=7, column=col).value
                cas = str(time_cell) if time_cell else ''
                
                # Get form (row 8)
                forma_cell = sheet.cell(row=8, column=col).value
                forma = str(forma_cell) if forma_cell else 'Neurčeno'
                
                # Get topic (row 9)
                tema_cell = sheet.cell(row=9, column=col).value
                tema = str(tema_cell) if tema_cell else 'Neurčeno'
                
                # Get teacher (row 10)
                ucitel_cell = sheet.cell(row=10, column=col).value
                ucitel = str(ucitel_cell) if ucitel_cell else 'Neurčeno'
                
                # Only add data if datum is valid
                if datum is not None:
                    activity_data = {
                        'datum': datum,
                        'cas': cas,
                        'forma': forma,
                        'tema': tema,
                        'ucitel': ucitel,
                        'hodin': hours
                    }
                    self.logger.info(f"[INVVZD] 16h Debug - Adding activity data: {activity_data}")
                    data.append(activity_data)
                else:
                    self.logger.info(f"[INVVZD] 16h Debug - Skipping data due to invalid datum")
                
                col += 1
            
            wb.close()
            
            self.logger.info(f"[INVVZD] 16h Debug - Total data collected: {len(data)} activities")
            
            if not data:
                self.add_error("Nenalezena žádná data aktivit")
                return None
                
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # No need to convert hours - they're already numeric from reading
            # Filter is already done during reading (hours > 0)
            
            # Format dates properly - ensure they include full date (DD.MM.YYYY)
            if 'datum' in df.columns:
                # First try to fix incomplete dates if they exist
                # For 16h template, data starts at row 6 (C6)
                df['datum'] = self._fix_incomplete_dates(df['datum'], start_row=6)
                # Then convert to datetime and format - SPECIFY dayfirst=True for DD.MM.YYYY format!
                df['datum'] = pd.to_datetime(df['datum'], format='%d.%m.%Y', dayfirst=True, errors='coerce')
                
                # Check for any failed date conversions
                nan_count = df['datum'].isna().sum()
                if nan_count > 0:
                    # Find which rows have invalid dates and report specific cells
                    failed_indices = df[df['datum'].isna()].index.tolist()
                    for idx in failed_indices:
                        # Row number in Excel is idx + 3 (header at 1, data starts at 3, 0-based index)
                        excel_row = idx + 3
                        self.add_error(f"Chybí nebo neplatné datum v řádku {excel_row}")
                    
                    self.add_info("Zkontrolujte správnost a případně soubor opravte a spusťte znovu")
                    return None
                
                # Format as DD.MM.YYYY
                df['datum'] = df['datum'].dt.strftime('%d.%m.%Y')
            
            # Format time if present
            if 'cas' in df.columns:
                # Keep time as is or format if needed
                df['cas'] = df['cas'].astype(str)
            
            # Calculate total hours
            self.hours_total = df['hodin'].sum()
            self.add_info(f"Celkem hodin: {self.hours_total}")
            self.add_info(f"Načteno {len(df)} aktivit")
            
            # Select required columns for output
            required_columns = ['datum', 'cas', 'hodin', 'forma', 'tema', 'ucitel']
            df = df[required_columns]
            
            return df
            
        except Exception as e:
            self.add_error(f"Chyba při čtení 16 hodinových dat: {str(e)}")
            return None
            
    def _read_32_hour_data(self, source_file: str) -> Optional[pd.DataFrame]:
        """Read data from 32 hour source file (List1 or zdroj-dochazka sheet)"""
        try:
            wb = load_workbook(source_file)
            
            # Try to find the correct sheet - prefer zdroj-dochazka, fallback to List1
            sheet_name = None
            
            if "zdroj-dochazka" in wb.sheetnames:
                sheet_name = "zdroj-dochazka"
            elif "List1" in wb.sheetnames:
                sheet_name = "List1"
            else:
                # Use first sheet as fallback
                sheet_name = wb.sheetnames[0]
                
            sheet = wb[sheet_name]
            self.add_info(f"Čtu 32h data z listu: {sheet_name}")
            
            if sheet_name == "zdroj-dochazka":
                # New format with zdroj-dochazka sheet
                data = []
                
                # Read data from specific rows
                # Row 6: dates, Row 7: forms, Row 8: topics, Row 9: teachers, Row 10: hours
                col = 3  # Start from column C (first activity)
                
                while True:
                    # Check if there's data in this column
                    hours_cell = sheet.cell(row=10, column=col).value
                    if hours_cell is None or str(hours_cell).strip() == '':
                        break
                        
                    try:
                        hours = int(float(str(hours_cell)))
                        if hours <= 0:
                            break
                    except (ValueError, TypeError):
                        break
                    
                    # Get date (row 6)
                    date_cell = sheet.cell(row=6, column=col).value
                    if date_cell:
                        if hasattr(date_cell, 'strftime'):
                            # It's already a datetime object
                            datum = date_cell.strftime('%d.%m.%Y')
                        else:
                            # Try to parse as string - store raw value for later fixing
                            datum = str(date_cell).strip()
                    else:
                        # ERROR: Missing date in activity column
                        from openpyxl.utils import get_column_letter
                        col_letter = get_column_letter(col)
                        self.add_error(f"Chybí datum aktivity v buňce {col_letter}6")
                        datum = None  # Mark as invalid
                    
                    # Get form (row 7)
                    forma_cell = sheet.cell(row=7, column=col).value
                    forma = str(forma_cell) if forma_cell else 'Neurčeno'
                    
                    # Get topic (row 8)
                    tema_cell = sheet.cell(row=8, column=col).value
                    tema = str(tema_cell) if tema_cell else 'Neurčeno'
                    
                    # Get teacher (row 9)
                    ucitel_cell = sheet.cell(row=9, column=col).value
                    ucitel = str(ucitel_cell) if ucitel_cell else 'Neurčeno'
                    
                    # Only add data if datum is valid
                    if datum is not None:
                        data.append({
                            'datum': datum,
                            'hodin': hours,
                            'forma': forma,
                            'tema': tema,
                            'ucitel': ucitel
                        })
                    
                    col += 1
                    
            else:
                # Legacy format with List1 sheet - still read dates from row 6!
                data = []
                
                # Read data from specific rows (same as zdroj-dochazka format)
                # Row 6: dates, Row 7: forms, Row 8: topics, Row 9: teachers, Row 10: hours
                col = 3  # Start from column C (first activity)
                
                while True:
                    # Check if there's data in this column
                    hours_cell = sheet.cell(row=10, column=col).value
                    if hours_cell is None or str(hours_cell).strip() == '':
                        break
                        
                    try:
                        hours = int(float(str(hours_cell)))
                        if hours <= 0:
                            break
                    except (ValueError, TypeError):
                        break
                    
                    # Get date (row 6) - same as in zdroj-dochazka format!
                    date_cell = sheet.cell(row=6, column=col).value
                    if date_cell:
                        if hasattr(date_cell, 'strftime'):
                            # It's already a datetime object
                            datum = date_cell.strftime('%d.%m.%Y')
                        else:
                            # Try to parse as string - store raw value for later fixing
                            datum = str(date_cell).strip()
                    else:
                        # ERROR: Missing date in activity column
                        from openpyxl.utils import get_column_letter
                        col_letter = get_column_letter(col)
                        self.add_error(f"Chybí datum aktivity v buňce {col_letter}6")
                        datum = None  # Mark as invalid
                    
                    # Get form (row 7)
                    forma_cell = sheet.cell(row=7, column=col).value
                    forma = str(forma_cell) if forma_cell else 'Neurčeno'
                    
                    # Get topic (row 8)
                    tema_cell = sheet.cell(row=8, column=col).value
                    tema = str(tema_cell) if tema_cell else f'Aktivita {col-2}'
                    
                    # Get teacher (row 9)
                    ucitel_cell = sheet.cell(row=9, column=col).value
                    ucitel = str(ucitel_cell) if ucitel_cell else 'Neurčeno'
                    
                    # Only add data if datum is valid
                    if datum is not None:
                        data.append({
                            'datum': datum,
                            'hodin': hours,
                            'forma': forma,
                            'tema': tema,
                            'ucitel': ucitel
                        })
                    
                    col += 1
            
            df = pd.DataFrame(data)
            
            # Check if we have any valid data
            if len(data) == 0:
                # No valid activities found
                wb.close()
                self.add_error("Soubor neobsahuje žádné platné aktivity")
                return None
            
            # Check for any errors in data
            error_count = sum(1 for msg in self.errors if "Chybí datum aktivity" in msg)
            if error_count > 0:
                # If there are data errors, don't create output file
                wb.close()
                self.add_info("Zkontrolujte správnost a případně soubor opravte a spusťte znovu")
                return None
            
            # Log basic info only if no errors
            self.add_info(f"Načteno {len(df)} aktivit z docházky")
            
            # Fix incomplete dates if needed (for 32h template, dates are in row 6, starting from column C=3)
            if 'datum' in df.columns:
                df['datum'] = self._fix_incomplete_dates(df['datum'], start_row=6, start_col=3)
                # Convert to datetime and format - SPECIFY dayfirst=True for DD.MM.YYYY format!
                df['datum'] = pd.to_datetime(df['datum'], format='%d.%m.%Y', dayfirst=True, errors='coerce')
                
                # Check for any failed conversions
                nan_count = df['datum'].isna().sum()
                if nan_count > 0:
                    self.add_warning(f"Upozornění: {nan_count} datumů se nepodařilo převést")
                    # Try to show which dates failed
                    failed_indices = df[df['datum'].isna()].index.tolist()
                    if failed_indices:
                        self.add_warning(f"Problematické řádky: {failed_indices[:5]}...")  # Show first 5
                
                # Format as DD.MM.YYYY
                df['datum'] = df['datum'].dt.strftime('%d.%m.%Y')
            
            # Calculate total hours
            self.hours_total = df['hodin'].sum()
            self.add_info(f"Celkem hodin: {self.hours_total}")
            
            wb.close()
            return df
            
        except Exception as e:
            self.add_error(f"Chyba při čtení 32 hodinových dat: {str(e)}")
            return None
            
    def _optimize_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Optimize data by removing duplicates"""
        original_count = len(df)
        df = df.drop_duplicates()
        removed_count = original_count - len(df)
        
        if removed_count > 0:
            self.add_info(f"Odstraněno {removed_count} duplicitních řádků")
            
        return df
        
    def _normalize_filename(self, filename: str) -> str:
        """
        Normalize filename: remove diacritics, replace spaces with underscores
        
        Args:
            filename: Original filename
            
        Returns:
            Normalized filename without diacritics and spaces
        """
        # Remove file extension if present (only .xlsx, .xls at the end)
        if filename.lower().endswith(('.xlsx', '.xls')):
            name_without_ext = filename.rsplit('.', 1)[0]
        else:
            name_without_ext = filename
        
        # Remove diacritics (Czech characters like á, č, ě, etc.)
        normalized = unicodedata.normalize('NFD', name_without_ext)
        ascii_text = ''.join(c for c in normalized if unicodedata.category(c) != 'Mn')
        
        # Replace problematic characters with underscores
        # Keep only alphanumeric and underscore, replace everything else
        clean_text = re.sub(r'[^\w]', '_', ascii_text)
        
        # Remove multiple consecutive underscores
        clean_text = re.sub(r'_+', '_', clean_text)
        
        # Remove leading/trailing underscores
        clean_text = clean_text.strip('_')
        
        return clean_text
        
    def _fix_incomplete_dates(self, date_series: pd.Series, start_row: int = 2, start_col: str = 'C') -> pd.Series:
        """
        Fix incomplete dates by inferring missing years from context
        
        Examples:
        - 14.5.2024, 16.6.2024, 17.6., 19.7.2024 → 17.6.2024
        - 24.1.2025, 15.2., 20.3.2025 → 15.2.2025
        
        Args:
            date_series: Series with date values (strings or datetime objects)
            
        Returns:
            Series with fixed dates
        """
        fixed_dates = []
        uncertain_fixes = []
        
        # Convert to string series for processing
        date_strings = date_series.astype(str)
        
        # Helper function to get cell reference
        def get_cell_ref(index, row, col):
            if isinstance(col, int):
                # For 32h version, col is starting column number
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(col + index)
                return f"{col_letter}{row}"
            else:
                # For 16h version, it's column letter with row offset
                return f"{col}{row + index}"
        
        for i, date_str in enumerate(date_strings):
            original_date = date_str
            
            # Skip if already a proper datetime or NaN
            if pd.isna(date_str) or date_str == 'nan':
                fixed_dates.append(date_str)
                continue
                
            # If it's already a datetime object converted to string, skip processing
            original_value = date_series.iloc[i]
            if pd.api.types.is_datetime64_any_dtype(pd.Series([original_value])):
                fixed_dates.append(date_str)
                continue
                
            # Clean up common issues
            date_str = str(date_str).strip()
            cleaned_date = date_str
            
            # Fix common typos and formatting issues
            # Replace commas with dots: "25,1.2025" → "25.1.2025"
            if ',' in date_str:
                date_str = date_str.replace(',', '.')
                cell_ref = get_cell_ref(i, start_row, start_col)
                self.add_info(f"Opravena čárka v datu v buňce {cell_ref}: {original_date} → {date_str}")
            
            # Remove extra spaces: "24 .1.2025" → "24.1.2025" or "25. 6. 2025" → "25.6.2025"
            if ' .' in date_str or '. ' in date_str:
                date_str = re.sub(r'\s+\.', '.', date_str)
                date_str = re.sub(r'\.\s+', '.', date_str)
                date_str = re.sub(r'\s+', ' ', date_str)  # Normalize multiple spaces to single space
                if date_str != cleaned_date:
                    cell_ref = get_cell_ref(i, start_row, start_col)
                    self.add_info(f"Opraveny mezery v datu v buňce {cell_ref}: {original_date} → {date_str}")
            
            # Handle dates with spaces between parts: "25 . 6 . 25" → "25.6.25"
            if re.match(r'^\d{1,2}\s+\.\s+\d{1,2}\s+\.\s+\d{2,4}$', date_str):
                date_str = date_str.replace(' ', '')
            
            # Check if date is incomplete (missing year)
            parts = date_str.split('.')
            
            
            if len(parts) == 2:  # DD.MM format - missing year
                day, month = parts
                inferred_year = self._infer_missing_year(i, date_strings, day, month)
                
                if inferred_year['confidence'] == 'high':
                    fixed_date = f"{day}.{month}.{inferred_year['year']}"
                    fixed_dates.append(fixed_date)
                    cell_ref = get_cell_ref(i, start_row, start_col)
                    self.add_info(f"Opraven datum v buňce {cell_ref}: {original_date} → {fixed_date}")
                elif inferred_year['confidence'] == 'medium':
                    fixed_date = f"{day}.{month}.{inferred_year['year']}"
                    fixed_dates.append(fixed_date)
                    cell_ref = get_cell_ref(i, start_row, start_col)
                    uncertain_fixes.append(f"Buňka {cell_ref}: {original_date} → {fixed_date}")
                else:
                    # Low confidence - add current year as fallback
                    current_year = datetime.now().year
                    fixed_date = f"{day}.{month}.{current_year}"
                    fixed_dates.append(fixed_date)
                    cell_ref = get_cell_ref(i, start_row, start_col)
                    uncertain_fixes.append(f"Buňka {cell_ref}: {original_date} → {fixed_date} (neistý)")
                    
            elif len(parts) == 3:  # DD.MM.YYYY format - check if year is empty
                day, month, year = parts
                if not year or year.strip() == "":  # Empty year part
                    inferred_year = self._infer_missing_year(i, date_strings, day, month)
                    
                    if inferred_year['confidence'] == 'high':
                        fixed_date = f"{day}.{month}.{inferred_year['year']}"
                        fixed_dates.append(fixed_date)
                        cell_ref = get_cell_ref(i, start_row, start_col)
                        self.add_info(f"Opraven datum v buňce {cell_ref}: {original_date} → {fixed_date}")
                    elif inferred_year['confidence'] == 'medium':
                        fixed_date = f"{day}.{month}.{inferred_year['year']}"
                        fixed_dates.append(fixed_date)
                        cell_ref = get_cell_ref(i, start_row, start_col)
                        uncertain_fixes.append(f"Buňka {cell_ref}: {original_date} → {fixed_date}")
                    else:
                        # Low confidence - add current year as fallback
                        current_year = datetime.now().year
                        fixed_date = f"{day}.{month}.{current_year}"
                        fixed_dates.append(fixed_date)
                        cell_ref = get_cell_ref(i, start_row, start_col)
                        uncertain_fixes.append(f"Buňka {cell_ref}: {original_date} → {fixed_date} (neistý)")
                else:
                    # Complete date
                    fixed_dates.append(date_str)
            else:
                # Invalid format - check for dd.mm.yy format
                if re.match(r'^\d{1,2}\.\d{1,2}\.\d{2}$', date_str):
                    # Two digit year format
                    parts = date_str.split('.')
                    day, month, year = parts
                    # Convert 2-digit year to 4-digit
                    year_int = int(year)
                    if year_int < 30:
                        full_year = 2000 + year_int
                    else:
                        full_year = 1900 + year_int
                    fixed_date = f"{day}.{month}.{full_year}"
                    fixed_dates.append(fixed_date)
                    cell_ref = get_cell_ref(i, start_row, start_col)
                    self.add_info(f"Opraven formát roku v buňce {cell_ref}: {original_date} → {fixed_date}")
                else:
                    # Really invalid format
                    fixed_dates.append(date_str)
                    cell_ref = get_cell_ref(i, start_row, start_col)
                    self.add_warning(f"Neplatný formát data v buňce {cell_ref}: {original_date} (očekáván formát DD.MM.YYYY)")
        
        # Report uncertain fixes
        if uncertain_fixes:
            self.add_info("Následující data byla opravena s nejistotou:")
            for fix in uncertain_fixes:
                self.add_info(f"  {fix}")
            self.add_info("Zkontrolujte správnost a případně soubor opravte a spusťte znovu")
            
        return pd.Series(fixed_dates)
        
    def _infer_missing_year(self, index: int, date_strings: pd.Series, day: str, month: str) -> dict:
        """
        Infer missing year from neighboring dates
        
        Args:
            index: Current position in series
            date_strings: All date strings
            day: Day part of incomplete date
            month: Month part of incomplete date
            
        Returns:
            Dict with 'year' and 'confidence' ('high', 'medium', 'low')
        """
        years_found = []
        
        # Look for complete dates in nearby positions (±3 positions)
        search_range = range(max(0, index-3), min(len(date_strings), index+4))
        
        for i in search_range:
            if i == index:
                continue
                
            date_str = str(date_strings.iloc[i]).strip()
            parts = date_str.split('.')
            
            if len(parts) == 3:  # Complete date DD.MM.YYYY
                try:
                    year = int(parts[2])
                    if 2020 <= year <= 2030:  # Reasonable year range
                        years_found.append(year)
                except ValueError:
                    continue
        
        if not years_found:
            # No nearby years found - use current year
            return {'year': datetime.now().year, 'confidence': 'low'}
            
        # Find most common year
        year_counts = {}
        for year in years_found:
            year_counts[year] = year_counts.get(year, 0) + 1
            
        most_common_year = max(year_counts.keys(), key=lambda y: year_counts[y])
        
        # Determine confidence based on consistency
        if len(set(years_found)) == 1:
            # All nearby years are the same
            confidence = 'high'
        elif year_counts[most_common_year] >= len(years_found) * 0.7:
            # Most common year appears in 70%+ of cases
            confidence = 'medium'
        else:
            # Mixed years
            confidence = 'low'
            
        # Additional logic: check if month sequence makes sense
        try:
            current_month = int(month)
            
            # Look at previous and next complete dates to see if year makes sense
            for i in [index-1, index+1]:
                if 0 <= i < len(date_strings):
                    neighbor_parts = str(date_strings.iloc[i]).strip().split('.')
                    if len(neighbor_parts) == 3:
                        try:
                            neighbor_month = int(neighbor_parts[1])
                            neighbor_year = int(neighbor_parts[2])
                            
                            # If this date's month fits chronologically, increase confidence
                            if i == index-1 and current_month > neighbor_month and most_common_year == neighbor_year:
                                confidence = 'high'
                            elif i == index+1 and current_month < neighbor_month and most_common_year == neighbor_year:
                                confidence = 'high'
                                
                        except ValueError:
                            continue
                            
        except ValueError:
            pass
            
        return {'year': most_common_year, 'confidence': confidence}
        
    def _create_output_filename(self, source_file: str, output_dir: str, 
                               keep_filename: bool) -> str:
        """Create output filename with proper prefix and normalization"""
        if keep_filename:
            # Get original filename without extension
            base_name = os.path.splitext(os.path.basename(source_file))[0]
            
            # Normalize filename (remove diacritics, replace spaces)
            normalized_name = self._normalize_filename(base_name)
            
            # Add version prefix
            version_prefix = f"{self.version}h_inv_"
            output_filename = f"{version_prefix}{normalized_name}_MSMT.xlsx"
            
            return os.path.join(output_dir, output_filename)
        else:
            # Fallback with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            version_prefix = f"{self.version}h_inv_"
            return os.path.join(output_dir, f"{version_prefix}MSMT_{timestamp}.xlsx")
            
    def _copy_template_with_data(self, template_path: str, output_path: str, 
                                data: pd.DataFrame, source_file: str):
        """Copy template file and fill with data using xlwings - following original approach"""
        try:
            self.logger.info(f"[INVVZD] === COPY TEMPLATE START ===")
            self.logger.info(f"[INVVZD] Template: {template_path}")
            self.logger.info(f"[INVVZD] Output: {output_path}")
            self.logger.info(f"[INVVZD] Data shape: {data.shape}")
            self.logger.info(f"[INVVZD] Source file: {source_file}")
            
            # Check platform compatibility
            current_platform = platform.system()
            self.logger.info(f"[INVVZD] Current platform: {current_platform}")
            
            if current_platform != 'Windows':
                self.add_error(f"Nástroj InvVzd vyžaduje Windows s nainstalovaným MS Excel. Aktuální platforma: {current_platform}")
                self.add_warning("Pro zpracování souborů použijte Windows počítač s MS Excel")
                self.logger.error(f"[INVVZD] Platform not supported: {current_platform}. xlwings requires Windows with Excel.")
                raise Exception("Platform not supported for xlwings")
            
            if not XLWINGS_AVAILABLE:
                self.add_error("xlwings není dostupný. Ujistěte se, že je nainstalován.")
                raise Exception("xlwings not available")
            
            # Following original approach exactly: visible=True, no unprotect
            self.logger.info(f"[INVVZD] Opening Excel application...")
            app = xw.App(visible=True)
            self.logger.info(f"[INVVZD] Loading template workbook...")
            wb = xw.Book(template_path)
            
            # STEP 1: Write student names to "Seznam účastníků" sheet at B4
            self.logger.info(f"[INVVZD] STEP 1: Writing student names...")
            sheet = wb.sheets['Seznam účastníků']
            
            # Extract student names from source file (column B, from row 11 until two empty rows)
            student_names = self._extract_student_names_from_data(source_file)
            self.logger.info(f"[INVVZD] Extracted {len(student_names)} student names")
            
            # Write student names exactly like original
            if len(student_names) > 0:
                sheet.range("B4").options(ndim="expand", transpose=True).value = student_names
            
            # STEP 2: Write activities to "Seznam aktivit" sheet at C3
            self.logger.info(f"[INVVZD] STEP 2: Writing activities...")
            sheet = wb.sheets['Seznam aktivit']
            
            # Prepare activities data for export (following original export_columns)
            if len(data) > 0:
                # For 16h version include time column, for 32h version exclude it
                if self.version == "16":
                    export_columns = ["datum", "cas", "hodin", "forma", "tema", "ucitel"]
                else:
                    export_columns = ["datum", "hodin", "forma", "tema", "ucitel"]
                activities_data = data[export_columns] if all(col in data.columns for col in export_columns) else data
                
                
                self.add_info(f"Zapisuji {len(activities_data)} aktivit do Seznam aktivit")
                sheet.range("C3").options(ndim="expand").value = activities_data.values
            
            # STEP 3: Write overview to "Přehled" sheet at C3
            self.logger.info(f"[INVVZD] STEP 3: Writing overview...")
            sheet = wb.sheets['Přehled']
            
            # Create overview data (student-activity combinations)
            overview_data = self._create_overview_data(student_names, data)
            if len(overview_data) > 0:
                self.add_info(f"Zapisuji {len(overview_data)} záznamů do Přehled")
                sheet.range("C3").options(ndim="expand").value = overview_data
            
            # STEP 4: Control check - verify SDP sums match activities total
            self.logger.info(f"[INVVZD] STEP 4: Verifying SDP sums...")
            self._verify_sdp_sums(wb)
            
            # Save as new file and close
            self.logger.info(f"[INVVZD] Saving output file: {output_path}")
            wb.save(output_path)
            self.logger.info(f"[INVVZD] Closing workbook...")
            wb.close()
            self.logger.info(f"[INVVZD] Quitting Excel application...")
            app.quit()
            self.logger.info(f"[INVVZD] === COPY TEMPLATE SUCCESS ===")
            
        except Exception as e:
            self.logger.error(f"[INVVZD] === COPY TEMPLATE EXCEPTION ===")
            self.logger.error(f"[INVVZD] Exception: {str(e)}")
            import traceback
            self.logger.error(f"[INVVZD] Traceback: {traceback.format_exc()}")
            self.add_error(f"Chyba při kopírování šablony: {str(e)}")
            # Try to close Excel if still open
            try:
                if 'wb' in locals():
                    wb.close()
                if 'app' in locals():
                    app.quit()
            except:
                pass
            raise
    
    def _extract_student_names_from_data(self, source_file: str) -> List[str]:
        """Extract student names from source file column B"""
        try:
            wb = load_workbook(source_file, read_only=True)
            sheet_name = "zdroj-dochazka" if "zdroj-dochazka" in wb.sheetnames else wb.sheetnames[0]
            sheet = wb[sheet_name]
            
            student_names = []
            empty_count = 0
            
            # Start from row 11 for 32h version, row 12 for 16h version  
            start_row = 11 if self.version == "16" else 10  # 0-indexed
            for row in range(start_row, sheet.max_row):
                cell_value = sheet.cell(row=row+1, column=2).value  # Column B
                
                if cell_value is None or str(cell_value).strip() == "":
                    empty_count += 1
                    if empty_count >= 2:  # Two consecutive empty rows
                        break
                else:
                    empty_count = 0
                    student_names.append(str(cell_value).strip())
            
            wb.close()
            self.add_info(f"Načteno {len(student_names)} jmen žáků")
            return student_names
            
        except Exception as e:
            self.add_error(f"Chyba při načítání jmen žáků: {str(e)}")
            return []
    
    def _create_overview_data(self, student_names: List[str], activities_data: pd.DataFrame) -> List[List]:
        """Create overview data combining students with activity numbers"""
        try:
            # Following original logic from create_orginal_file
            overview_result = []
            
            # For each activity (numbered 1, 2, 3...)
            for activity_num, (_, activity) in enumerate(activities_data.iterrows(), 1):
                # Add all students for this activity
                for student_name in student_names:
                    overview_result.append([activity_num, student_name])
            
            self.add_info(f"Vytvořen přehled: {len(activities_data)} aktivit × {len(student_names)} žáků = {len(overview_result)} záznamů")
            return overview_result
            
        except Exception as e:
            self.add_error(f"Chyba při vytváření přehledu: {str(e)}")
            return []
    
    def select_folder(self, folder_path: str) -> Dict[str, Any]:
        """Scan folder for attendance files and filter them"""
        self.logger.info(f"[INVVZD] === SELECT FOLDER START ===")
        self.logger.info(f"[INVVZD] Folder path: {folder_path}")
        
        result = {"success": False, "files": [], "message": ""}
        
        try:
            # Check if folder exists
            if not os.path.exists(folder_path):
                self.logger.error(f"[INVVZD] Folder does not exist: {folder_path}")
                result["message"] = f"Složka neexistuje: {folder_path}"
                return result
                
            # List all files in folder
            all_files = os.listdir(folder_path)
            self.logger.info(f"[INVVZD] All files in folder: {all_files}")
            
            attendance_files = []
            
            # Scan for Excel files
            for file in all_files:
                if file.endswith(('.xlsx', '.xls')) and not file.startswith('~$'):
                    full_path = os.path.join(folder_path, file)
                    self.logger.info(f"[INVVZD] Checking Excel file: {file}")
                    
                    # Skip output files (already processed)
                    if file.startswith(('32h_inv_', '16h_inv_', '32_hodin_inovativniho_vzdelavani_', '16_hodin_inovativniho_vzdelavani_')):
                        self.logger.info(f"[INVVZD] Skipping output file: {file}")
                        continue
                    
                    # Skip template files
                    if 'sablona' in file.lower() or 'template' in file.lower():
                        self.logger.info(f"[INVVZD] Skipping template file: {file}")
                        continue
                    
                    # Try to detect version from content
                    self.logger.info(f"[INVVZD] Detecting version for: {full_path}")
                    version = self._detect_source_version(full_path)
                    self.logger.info(f"[INVVZD] Detected version: {version}")
                    
                    if version:
                        # Check if version matches current template
                        if self.version and version != self.version:
                            self.logger.info(f"[INVVZD] Skipping file {file} - version mismatch (file: {version}h, template: {self.version}h)")
                            continue
                            
                        self.logger.info(f"[INVVZD] Valid attendance file found: {file} (version {version}h)")
                        attendance_files.append({
                            "path": full_path,
                            "name": file,
                            "version": f"{version} hodin",
                            "compatible": True
                        })
                    else:
                        self.logger.info(f"[INVVZD] Not an attendance file: {file}")
                        
            # Log summary
            self.logger.info(f"[INVVZD] Total compatible files found: {len(attendance_files)}")
            
            # Include info about template version in message
            template_hours = self.config["hours"] if self.version else "?"
            
            if attendance_files:
                result["success"] = True
                result["files"] = attendance_files
                if self.version:
                    result["message"] = f"Nalezeno {len(attendance_files)} souborů kompatibilních se šablonou {template_hours} hodin"
                else:
                    result["message"] = f"Nalezeno {len(attendance_files)} souborů s docházkou"
                self.logger.info(f"[INVVZD] SUCCESS: {result['message']}")
            else:
                if self.version:
                    result["message"] = f"Ve složce nebyly nalezeny žádné soubory kompatibilní se šablonou {template_hours} hodin"
                else:
                    result["message"] = "Ve složce nebyly nalezeny žádné soubory s docházkou"
                self.logger.warning(f"[INVVZD] WARNING: {result['message']}")
                
        except Exception as e:
            result["message"] = f"Chyba při procházení složky: {str(e)}"
            self.logger.error(f"[INVVZD] ERROR: {result['message']}")
            import traceback
            self.logger.error(f"[INVVZD] Traceback: {traceback.format_exc()}")
            
        self.logger.info(f"[INVVZD] === SELECT FOLDER END ===")
        self.logger.info(f"[INVVZD] Result: {result}")
        return result
    
    def _verify_sdp_sums(self, wb):
        """Verify SDP sums match total hours - following original control logic"""
        try:
            # Calculate activities total (Seznam aktivit column depends on version)
            aktivit_sheet = wb.sheets['Seznam aktivit']
            activities_total = 0
            row = 3
            
            # For 16h version, hours are in column E (includes time column)
            # For 32h version, hours are in column D  
            hours_column = "E" if self.version == "16" else "D"
            
            while True:
                cell_value = aktivit_sheet.range(f"{hours_column}{row}").value
                if cell_value is None or str(cell_value).strip() == '':
                    break
                try:
                    activities_total += int(float(str(cell_value)))
                except:
                    pass
                row += 1
            
            # Check SDP sheet sums
            sdp_sheet = wb.sheets['SDP']
            
            # Sum C4:C10 (forma range)
            sdp_forma_total = 0
            for row in range(4, 11):  # C4 to C10
                cell_value = sdp_sheet.range(f"C{row}").value
                if cell_value is not None:
                    try:
                        sdp_forma_total += int(float(str(cell_value)))
                    except:
                        pass
            
            # Sum C12:C28 (tema range) 
            sdp_tema_total = 0
            for row in range(12, 29):  # C12 to C28
                cell_value = sdp_sheet.range(f"C{row}").value
                if cell_value is not None:
                    try:
                        sdp_tema_total += int(float(str(cell_value)))
                    except:
                        pass
            
            # Compare and report with detailed logging
            self.logger.info(f"[INVVZD] === SDP VERIFICATION ===")
            self.logger.info(f"[INVVZD] Activities total: {activities_total} hours")
            self.logger.info(f"[INVVZD] SDP forma total (C4-C10): {sdp_forma_total} hours")
            self.logger.info(f"[INVVZD] SDP tema total (C12-C28): {sdp_tema_total} hours")
            
            self.add_info(f"Kontrola součtů:")
            self.add_info(f"  Aktivity: {activities_total}h")
            self.add_info(f"  SDP forma: {sdp_forma_total}h")
            self.add_info(f"  SDP téma: {sdp_tema_total}h")
            
            if activities_total == sdp_forma_total == sdp_tema_total:
                self.logger.info(f"[INVVZD] All sums match!")
                self.add_info(f"✅ Všechny součty souhlasí")
            else:
                self.logger.error(f"[INVVZD] ❌ SUMS DO NOT MATCH!")
                self.logger.error(f"[INVVZD] Activities: {activities_total}, Forma: {sdp_forma_total}, Tema: {sdp_tema_total}")
                
                self.add_error(f"❌ NESOUHLASÍ součty v SDP!")
                self.add_error(f"Aktivity: {activities_total}h")
                self.add_error(f"SDP forma: {sdp_forma_total}h")
                self.add_error(f"SDP téma: {sdp_tema_total}h")
                self.add_warning("ZKONTROLUJTE výsledný soubor - aktivity na listu 'Seznam aktivit'")
                
        except Exception as e:
            self.add_error(f"Chyba při kontrole SDP součtů: {str(e)}")
            
    def process_paths(self, source_files: List[str], template_path: str, 
                     output_dir: str, keep_filename: bool = True,
                     optimize: bool = False) -> Dict[str, Any]:
        """
        Process files using file paths (for testing)
        
        Args:
            source_files: List of source file paths
            template_path: Path to template file
            output_dir: Output directory
            keep_filename: Whether to keep original filename
            optimize: Whether to optimize data
            
        Returns:
            Processing result dictionary
        """
        try:
            self.clear_messages()
            
            # Validate template
            if not self.file_exists(template_path):
                self.add_error(f"Šablona neexistuje: {template_path}")
                return {"success": False, "output_files": []}
                
            # Detect and validate template version
            template_version = self._detect_template_version(template_path)
            if template_version != self.version:
                self.add_error(f"Verze šablony ({template_version}) neodpovídá procesoru ({self.version})")
                return {"success": False, "output_files": []}
                
            # Process files
            output_files = []
            files_processed = []
            
            for source_file in source_files:
                self.add_info(f"Zpracovávám soubor: {os.path.basename(source_file)}")
                
                if not self.file_exists(source_file):
                    self.add_error(f"Zdrojový soubor neexistuje: {source_file}")
                    continue
                    
                # Validate version match
                if not self._validate_version_match(source_file, template_path):
                    continue
                    
                # Process file
                output_file = self._process_single_file(
                    source_file, template_path, output_dir, 
                    keep_filename, optimize
                )
                
                if output_file:
                    output_files.append(output_file)
                    files_processed.append({
                        "source": source_file,
                        "filename": os.path.basename(output_file),
                        "hours": self.config['hours'] if self.config else 0
                    })
                else:
                    # File failed but we continue with others
                    self.add_error(f"❌ Soubor {os.path.basename(source_file)} nebyl zpracován kvůli chybám")
                    
            # Success if at least one file was processed
            success = len(output_files) > 0
            return {
                "success": success,
                "output_files": output_files,
                "files": files_processed,
                "errors": self.errors,
                "warnings": self.warnings,
                "info": self.info_messages
            }
            
        except Exception as e:
            self.add_error(f"Neočekávaná chyba: {str(e)}")
            return {"success": False, "output_files": []}
            
    def clear_messages(self):
        """Clear all messages"""
        self.errors.clear()
        self.warnings.clear()
        self.info_messages.clear()
        
    def file_exists(self, filepath: str) -> bool:
        """Check if file exists"""
        exists = os.path.isfile(filepath)
        self.logger.info(f"[INVVZD] Checking file existence: {filepath} -> {exists}")
        if not exists:
            # Additional debug info
            self.logger.info(f"[INVVZD] Current directory: {os.getcwd()}")
            self.logger.info(f"[INVVZD] Path is absolute: {os.path.isabs(filepath)}")
            if os.path.exists(filepath):
                self.logger.info(f"[INVVZD] Path exists but is not a file (maybe directory?)")
            else:
                self.logger.info(f"[INVVZD] Path does not exist at all")
        return exists