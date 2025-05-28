import pandas as pd
import glob
from openpyxl import load_workbook
import os
import warnings
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime
import logging

from .base_tool import BaseTool

warnings.filterwarnings('ignore')

# Configuration constants
DATE_RANGES = [
    ('2022-2023', '2022-09-01', '2023-08-31'),
    ('2023-2024', '2023-09-01', '2024-08-31'),
    ('2024-2025', '2024-09-01', '2025-08-31'),
    ('2025-2026', '2025-09-01', '2026-08-31')
]

TEMA_ORDER = [
    "čtenářská pre/gramotnost",
    "matematická pre/gramotnost", 
    "umělecká gramotnost",
    "mediální gramotnost",
    "cizí jazyky/komunikace v cizím jazyce",
    "inkluze včetně primární prevence",
    "přírodovědné a technické vzdělávání",
    "evvo a vzdělávání pro udržitelný rozvoj",
    "vzdělávání s využitím nových technologií",
    "kulturní povědomí a vyjádření",
    "historické povědomí, výuka moderních dějin",
    "rozvoj podnikavosti a kreativity",
    "well-being a psychohygiena",
    "genderová tematika v obsahu vzdělávání",
    "kariérové poradenství včetně identifikace a rozvoje nadání",
    "občanské vzdělávání a demokratické myšlení"
]


class ZorSpecDatProcessor(BaseTool):
    """Processor for special data items in ZoR (závěrečná zpráva o realizaci)"""
    
    def __init__(self, logger: Optional[logging.Logger] = None):
        super().__init__(logger)
        self.sheet_name = "Přehled"
        self.cols_names = ["ca", "jmena", "datum", "pocet_hodin", "forma", "tema"]
        self.cols_agg = ["forma/téma", "čislo", "cena", "typ"]
        self.result_cols_names = ["forma/téma", "číslo celkem", "cena celkem", "typ"]
        self.agg_function = [{"hash_jmena": "nunique"}, {"pocet_hodin": "sum"}]
        self.name_list = []
        
    def validate_inputs(self, files: List[str], options: Dict[str, Any]) -> bool:
        """Validate input files and options"""
        self.clear_messages()
        
        # Check if files or directory provided
        source_dir = options.get('source_dir')
        exclude_list = options.get('exclude_list', '')
        
        if not files and not source_dir:
            self.add_error("Nebyla poskytnuta žádná data ke zpracování")
            return False
            
        # If source_dir provided, get files from directory
        if source_dir:
            if not os.path.isdir(source_dir):
                self.add_error(f"Složka neexistuje: {source_dir}")
                return False
                
        # Load exclude list if provided
        if exclude_list and os.path.isfile(exclude_list):
            self._load_exclude_names(exclude_list)
            self.add_info(f"Načten seznam vyloučených: {len(self.name_list)} jmen")
            
        return True
        
    def process(self, files: List[str], options: Dict[str, Any]) -> Dict[str, Any]:
        """Process attendance files and generate ZoR special data items report"""
        if not self.validate_inputs(files, options):
            return self.get_result(False)
            
        try:
            source_dir = options.get('source_dir')
            output_dir = options.get('output_dir', source_dir or os.getcwd())
            
            # Get files to process
            if source_dir:
                excel_files = self._get_files_from_directory(source_dir)
            else:
                excel_files = self._validate_files(files)
                
            if not excel_files:
                self.add_error("Nebyl nalezen žádný platný soubor ke zpracování")
                return self.get_result(False)
                
            self.add_info(f"Nalezeno {len(excel_files)} souborů ke zpracování")
            
            # Process files and generate report
            html_report, unique_names_data = self._generate_report(excel_files)
            
            # Save outputs
            html_file = self._save_html_report(output_dir, html_report)
            txt_file = self._save_unique_names(output_dir, unique_names_data)
            
            processed_data = {
                "files_processed": len(excel_files),
                "unique_students": len(unique_names_data),
                "html_report": html_file,
                "names_list": txt_file,
                "output_files": [html_file, txt_file]
            }
            
            self.add_info(f"Report uložen: {html_file}||{os.path.basename(html_file)}")
            self.add_info(f"Seznam žáků uložen: {txt_file}||{os.path.basename(txt_file)}")
            
            return self.get_result(True, processed_data)
            
        except Exception as e:
            self.add_error(f"Chyba při zpracování: {str(e)}")
            return self.get_result(False)
            
    def _get_files_from_directory(self, source_dir: str) -> List[str]:
        """Get valid Excel files from directory"""
        # Exclude temporary Excel files (starting with ~)
        file_pattern = os.path.join(source_dir, "*.xlsx")
        file_names = [f for f in glob.glob(file_pattern) if not os.path.basename(f).startswith("~")]
        
        # Filter files that have the required sheet
        valid_files = []
        for file_path in file_names:
            if self._sheet_exists(file_path):
                valid_files.append(file_path)
            else:
                self.add_warning(f"Soubor nemá list '{self.sheet_name}': {os.path.basename(file_path)}")
                
        return valid_files
        
    def _validate_files(self, files: List[str]) -> List[str]:
        """Validate individual files"""
        valid_files = []
        for file_path in files:
            if not os.path.isfile(file_path):
                self.add_error(f"Soubor neexistuje: {file_path}")
                continue
                
            if not self._sheet_exists(file_path):
                self.add_error(f"Soubor nemá požadovaný list '{self.sheet_name}': {os.path.basename(file_path)}")
                continue
                
            valid_files.append(file_path)
            
        return valid_files
        
    def _sheet_exists(self, excel_file: str) -> bool:
        """Check if required sheet exists in Excel file"""
        try:
            wb = load_workbook(excel_file, read_only=True)
            result = self.sheet_name in wb.sheetnames
            wb.close()
            return result
        except Exception as e:
            self.add_warning(f"Chyba při kontrole souboru {os.path.basename(excel_file)}: {str(e)}")
            return False
            
    def _get_template_name(self, excel_file: str) -> str:
        """Get template name from Excel file (cell C4 on first sheet)"""
        try:
            wb = load_workbook(excel_file, read_only=True)
            value = wb.worksheets[0].cell(row=4, column=3).value
            wb.close()
            return str(value) if value else "Neznámá šablona"
        except Exception:
            return "Neznámá šablona"
            
    def _calculate_subreport(self, excel_file: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """Calculate subreport for single Excel file"""
        try:
            # Read data from Přehled sheet
            df = pd.read_excel(
                excel_file,
                sheet_name=self.sheet_name,
                usecols="C,D,E,F:H",
                names=self.cols_names,
                skiprows=1
            )
            
            # Clean string data
            df = df.applymap(lambda x: x.lower().strip() if isinstance(x, str) else x)
            
            # Standardize forma values
            replacements = {
                'projektové vzdělávání (ve škole / mimo školu)': 'projektové vzdělávání / projektová výuka',
                'propojování formálního a neformálního vzdělávání': 'propojování neformálního a formálního vzdělávání'
            }
            
            for old_val, new_val in replacements.items():
                df['forma'] = df['forma'].replace(old_val, new_val)
                
            # Add file identifier and clean data
            df['ca'] = df['ca'].astype(str) + str(hash(excel_file))
            df = df.dropna()
            df['pocet_hodin'] = df['pocet_hodin'].astype(int)
            df['hash_jmena'] = df['jmena'].apply(self._custom_hash)
            
            # Add template name
            template_name = self._get_template_name(excel_file)
            df['sablona'] = template_name
            
            # Generate aggregated results
            forma_agg = self._aggregate(df, "forma", "forma")
            tema_agg = self._aggregate(df, "tema", "téma")
            
            return df, pd.concat([forma_agg, tema_agg], axis="rows")
            
        except Exception as e:
            raise Exception(f"Chyba při zpracování souboru {os.path.basename(excel_file)}: {str(e)}")
            
    def _custom_hash(self, value: str) -> int:
        """Calculate custom hash for name (handles name variations)"""
        if pd.isna(value):
            return 0
        hash_value = "".join([x.lower() for x in str(value) if x.isalnum()])
        hash_value = hash("".join(sorted(hash_value)))
        return hash_value
        
    def _aggregate(self, df: pd.DataFrame, group_col: str, col_name: str) -> pd.DataFrame:
        """Aggregate data by specified column"""
        # Count unique students
        r1 = df.groupby([group_col], group_keys=False).agg(self.agg_function[0]).reset_index()
        
        # Sum hours
        r2 = (df.groupby(["ca", group_col, "pocet_hodin"], group_keys=False)
              .agg(self.agg_function[0]).reset_index()
              .groupby([group_col], group_keys=False)
              .agg(self.agg_function[1]).reset_index())
        
        # Merge results
        result = r1.merge(r2, on=group_col, how="left")
        result['typ'] = col_name
        result.columns = self.cols_agg
        
        return result
        
    def _generate_report(self, excel_files: List[str]) -> Tuple[str, List[List]]:
        """Generate complete HTML report"""
        concatenated = pd.DataFrame()
        html_parts = []
        
        # Process each file
        for file_path in excel_files:
            try:
                df, subreport = self._calculate_subreport(file_path)
                concatenated = pd.concat([concatenated, df], axis="rows")
                
                # Add file section to HTML
                file_name = os.path.basename(file_path)
                html_parts.append(f"<h2>{file_name}</h2>")
                html_parts.append(self._dataframe_to_html_table(subreport))
                
                self.add_info(f"Zpracován soubor: {file_name}")
                
            except Exception as e:
                self.add_error(f"Chyba při zpracování {os.path.basename(file_path)}: {str(e)}")
                continue
                
        if concatenated.empty:
            raise Exception("Žádná data nebyla úspěšně zpracována")
            
        # Filter out excluded names
        if self.name_list:
            original_count = len(concatenated)
            concatenated = concatenated[~concatenated['jmena'].isin(self.name_list)]
            excluded_count = original_count - len(concatenated)
            if excluded_count > 0:
                self.add_info(f"Vyloučeno {excluded_count} záznamů podle seznamu")
                
        # Calculate hash names for deduplication
        concatenated['hash_jmena'] = concatenated['jmena'].apply(self._custom_hash)
        
        # Generate template-based aggregation
        template_data = self._aggregate_data_by_template(concatenated)
        
        # Add template sections to HTML
        for template_name, data in template_data.items():
            html_parts.insert(0, self._dataframe_to_html_table(data))
            html_parts.insert(0, f"<h3>Šablona: {template_name}</h3>")
            
        html_parts.insert(0, "<h2>SDP ZoR</h2>")
        
        # Generate final aggregated results
        forma_result = self._aggregate(concatenated, "forma", "forma")
        tema_result = self._aggregate(concatenated, "tema", "téma")
        
        final_result = pd.concat([forma_result, tema_result], axis="rows")
        final_result.columns = self.result_cols_names
        
        # Get unique names
        unique_names, unique_count = self._get_unique_names(concatenated)
        
        # Build final HTML
        html_parts.insert(0, self._dataframe_to_html_table(final_result))
        html_parts.insert(0, f"<h3>Unikátní žáci v ZoR: {unique_count}</h3>")
        html_parts.insert(0, "<h2>Údaje do ZoR</h2>")
        html_parts.insert(0, "<h1>Specifické datové položky pro ZoR</h1>")
        
        # Create complete HTML document
        html_content = self._create_html_document("".join(html_parts))
        
        return html_content, unique_names
        
    def _aggregate_data_by_template(self, df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
        """Aggregate data by template with period analysis"""
        df_copy = df.copy()
        df_copy['datum'] = pd.to_datetime(df_copy['datum'], format='%d.%m.%Y')
        
        # Create period column
        df_copy['period'] = pd.cut(
            df_copy['datum'],
            bins=[pd.to_datetime(start) for _, start, _ in DATE_RANGES] + [pd.to_datetime('2026-09-01')],
            labels=[label for label, _, _ in DATE_RANGES],
            include_lowest=True
        )
        
        def aggregate_single_template(template_df):
            template_df = template_df.sort_values('datum')
            participant_tema_first_period = {}
            result = pd.DataFrame(columns=['tema'] + [label for label, _, _ in DATE_RANGES])
            
            for period, _, _ in DATE_RANGES:
                period_df = template_df[template_df['period'] == period]
                tema_counts = {}
                
                for _, row in period_df.iterrows():
                    participant = row['hash_jmena']
                    tema = row['tema']
                    key = (participant, tema)
                    
                    if key not in participant_tema_first_period:
                        participant_tema_first_period[key] = period
                        tema_counts[tema] = tema_counts.get(tema, 0) + 1
                        
                tema_counts_df = pd.DataFrame(list(tema_counts.items()), columns=['tema', period])
                if result.empty:
                    result = tema_counts_df
                else:
                    result = result.merge(tema_counts_df, on='tema', how='outer')
                    
            result = result.fillna(0)
            
            # Ensure all period columns exist
            for period, _, _ in DATE_RANGES:
                if period not in result.columns:
                    result[period] = 0
                    
            # Reorder columns
            cols = ['tema'] + [label for label, _, _ in DATE_RANGES]
            result = result[cols]
            
            # Convert to integers
            for col in result.columns[1:]:
                result[col] = result[col].astype(int)
                
            # Sort by tema order
            result['tema_cat'] = pd.Categorical(result['tema'], categories=TEMA_ORDER, ordered=True)
            result = result.sort_values('tema_cat').drop('tema_cat', axis=1)
            
            # Add missing temas
            for tema in TEMA_ORDER:
                if tema not in result['tema'].values:
                    new_row = pd.DataFrame({'tema': [tema], **{col: [0] for col in result.columns if col != 'tema'}})
                    result = pd.concat([result, new_row], ignore_index=True)
                    
            # Final sort and remove empty rows
            result = result.set_index('tema').loc[TEMA_ORDER].reset_index()
            result = result[result.iloc[:, 1:].sum(axis=1) > 0]
            
            return result
            
        grouped = df_copy.groupby('sablona')
        return {template: aggregate_single_template(group) for template, group in grouped}
        
    def _get_unique_names(self, df: pd.DataFrame) -> Tuple[List[List], int]:
        """Extract unique student names"""
        if "jmena" not in df.columns:
            return [], 0
            
        # Remove duplicates by hash
        unique_data = df.drop_duplicates(subset=["hash_jmena"]).sort_values(by=["jmena"])
        unique_count = df["hash_jmena"].nunique()
        
        return unique_data[["jmena", "hash_jmena"]].values.tolist(), unique_count
        
    def _load_exclude_names(self, file_path: str):
        """Load names to exclude from processing"""
        self.name_list = []
        try:
            with open(file_path, 'r', encoding="utf-8") as f:
                for line in f:
                    name = line.strip()
                    if name:
                        self.name_list.append(name)
        except Exception as e:
            self.add_warning(f"Chyba při načítání seznamu vyloučených: {str(e)}")
            self.name_list = []
            
    def _dataframe_to_html_table(self, df: pd.DataFrame) -> str:
        """Convert DataFrame to HTML table with styling"""
        return df.to_html(
            index=False,
            classes='blue_light',
            escape=False,
            table_id='data-table'
        )
        
    def _create_html_document(self, body_content: str) -> str:
        """Create complete HTML document"""
        return f"""
        <!DOCTYPE html>
        <html lang="cs">
        <head>
            <meta charset="utf-8">
            <meta http-equiv="X-UA-Compatible" content="IE=edge">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Specifické datové položky - report</title>
            <style type="text/css">
                h1, h2, h3 {{
                    color: navy;
                    font-family: Century Gothic, sans-serif;
                }}
                .blue_light {{
                    border-collapse: collapse;
                    margin: 20px 0;
                    font-size: 0.9em;
                    font-family: sans-serif;
                    min-width: 400px;
                    border-radius: 5px 5px 0 0;
                    overflow: hidden;
                    box-shadow: 0 0 20px rgba(0, 0, 0, 0.15);
                }}
                .blue_light thead tr {{
                    background-color: #009879;
                    color: #ffffff;
                    text-align: left;
                }}
                .blue_light th,
                .blue_light td {{
                    padding: 12px 15px;
                    border: 1px solid #dddddd;
                }}
                .blue_light tbody tr {{
                    border-bottom: 1px solid #dddddd;
                }}
                .blue_light tbody tr:nth-of-type(even) {{
                    background-color: #f3f3f3;
                }}
                .blue_light tbody tr:last-of-type {{
                    border-bottom: 2px solid #009879;
                }}
            </style>
        </head>
        <body>
            {body_content}
        </body>
        </html>
        """
        
    def _save_html_report(self, output_dir: str, html_content: str) -> str:
        """Save HTML report to file"""
        output_file = os.path.join(output_dir, "result.html")
        try:
            with open(output_file, 'w', encoding="utf-8") as f:
                f.write(html_content)
            return output_file
        except Exception as e:
            raise Exception(f"Chyba při ukládání HTML reportu: {str(e)}")
            
    def _save_unique_names(self, output_dir: str, unique_names: List[List]) -> str:
        """Save unique names list to text file"""
        output_file = os.path.join(output_dir, "seznam_zaku.txt")
        try:
            with open(output_file, 'w', encoding="utf-8") as f:
                for name_data in unique_names:
                    if len(name_data) >= 2:
                        f.write(f"{name_data[0]}\t\t\t{name_data[1]}\n")
            return output_file
        except Exception as e:
            raise Exception(f"Chyba při ukládání seznamu žáků: {str(e)}")
            
    def detect_file_info(self, file_path: str) -> Dict[str, Any]:
        """
        Detect if file has 'Úvod a postup vyplňování' sheet and extract version from B1 cell
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dictionary with detection results
        """
        result = {
            'has_intro_sheet': False,
            'version': None,
            'error': None
        }
        
        try:
            wb = load_workbook(file_path, read_only=True)
            
            # Check for the required sheet
            intro_sheet_name = 'Úvod a postup vyplňování'
            if intro_sheet_name in wb.sheetnames:
                result['has_intro_sheet'] = True
                
                # Get version from B1 cell
                ws = wb[intro_sheet_name]
                version_value = ws.cell(row=1, column=2).value
                
                if version_value:
                    # Extract version number from the value
                    # Expected format: "Verze X.Y" or similar
                    version_str = str(version_value).strip()
                    result['version'] = version_str
                else:
                    result['version'] = 'Neznámá'
                    
            wb.close()
            
        except Exception as e:
            result['error'] = str(e)
            self.logger.error(f"Error detecting file info for {os.path.basename(file_path)}: {str(e)}")
            
        return result
            
    def process_paths(self, file_paths: List[str], output_dir: str, options: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        Process files using file paths (for compatibility with API)
        
        Args:
            file_paths: List of file paths to process
            output_dir: Output directory for results
            options: Additional processing options
            
        Returns:
            Processing result dictionary
        """
        if options is None:
            options = {}
            
        # Add output_dir to options
        options['output_dir'] = output_dir
        
        # Use the existing process method
        result = self.process(file_paths, options)
        
        # Return in the expected format for API
        if result['success']:
            return {
                'success': True,
                'files_processed': result['data']['files_processed'],
                'unique_students': result['data']['unique_students'],
                'output_files': [
                    {
                        'filename': os.path.basename(path),
                        'path': path,
                        'type': 'html' if path.endswith('.html') else 'text'
                    }
                    for path in result['data']['output_files']
                ],
                'errors': result.get('errors', []),
                'warnings': result.get('warnings', []),
                'info': result.get('info', [])
            }
        else:
            return {
                'success': False,
                'errors': result.get('errors', []),
                'warnings': result.get('warnings', []),
                'info': result.get('info', [])
            }