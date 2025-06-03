"""
Excel utilities for working with Excel files
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
from typing import Dict, Any, Optional, Tuple, List
import unicodedata
import re


class ExcelHelper:
    """Helper class for Excel operations"""
    
    @staticmethod
    def normalize_filename(filename: str, remove_diacritics: bool = True) -> str:
        """
        Normalize filename - remove diacritics and special characters
        
        Args:
            filename: Original filename
            remove_diacritics: Whether to remove diacritics
            
        Returns:
            Normalized filename
        """
        # Get base name without extension
        base_name = os.path.splitext(filename)[0]
        extension = os.path.splitext(filename)[1]
        
        if remove_diacritics:
            # Remove diacritics
            base_name = ''.join(
                c for c in unicodedata.normalize('NFD', base_name)
                if unicodedata.category(c) != 'Mn'
            )
        
        # Replace spaces and special characters
        base_name = re.sub(r'[^\w\s-]', '', base_name)
        base_name = re.sub(r'[-\s]+', '_', base_name)
        
        return base_name + extension
        
    @staticmethod
    def read_excel_value(wb, sheet_name: str, cell: str) -> Any:
        """
        Read a single cell value from Excel workbook
        
        Args:
            wb: Openpyxl workbook
            sheet_name: Name of the sheet
            cell: Cell reference (e.g., 'B1')
            
        Returns:
            Cell value or None
        """
        try:
            if sheet_name in wb.sheetnames:
                return wb[sheet_name][cell].value
        except:
            pass
        return None
        
    @staticmethod
    def find_header_row(df: pd.DataFrame, expected_headers: List[str], 
                       max_rows: int = 20) -> Optional[int]:
        """
        Find row containing expected headers
        
        Args:
            df: DataFrame to search
            expected_headers: List of expected header values
            max_rows: Maximum rows to search
            
        Returns:
            Row index or None
        """
        for i in range(min(max_rows, len(df))):
            row_values = df.iloc[i].astype(str).str.lower().tolist()
            matches = sum(1 for header in expected_headers 
                         if any(header.lower() in val for val in row_values))
            
            if matches >= len(expected_headers) * 0.7:  # 70% match
                return i
                
        return None
        
    @staticmethod
    def detect_version_from_template(template_path: str) -> Optional[str]:
        """
        Detect version (16h or 32h) from template file
        
        Args:
            template_path: Path to template file
            
        Returns:
            '16' or '32' or None
        """
        try:
            wb = load_workbook(template_path, read_only=True, data_only=True)
            
            # Check common cells for version indicators
            version_cells = {
                'List1': ['B1', 'A1', 'C1'],
                'Sheet1': ['B1', 'A1', 'C1']
            }
            
            for sheet, cells in version_cells.items():
                if sheet in wb.sheetnames:
                    for cell in cells:
                        value = str(wb[sheet][cell].value or '').lower()
                        if '32' in value and 'hodin' in value:
                            return '32'
                        elif '16' in value and 'hodin' in value:
                            return '16'
                            
            wb.close()
        except:
            pass
            
        # Fallback to filename
        filename = os.path.basename(template_path).lower()
        if '32' in filename:
            return '32'
        elif '16' in filename:
            return '16'
            
        return None
        
    @staticmethod
    def copy_excel_with_xlwings(source_path: str, target_path: str, 
                               data_dict: Dict[str, pd.DataFrame]) -> bool:
        """
        Copy Excel file preserving formatting using xlwings
        
        Args:
            source_path: Source Excel file
            target_path: Target Excel file
            data_dict: Dictionary of sheet_name -> DataFrame
            
        Returns:
            Success status
        """
        try:
            import xlwings as xw
            
            # Create directory if needed
            os.makedirs(os.path.dirname(target_path), exist_ok=True)
            
            # Copy file first
            import shutil
            shutil.copy2(source_path, target_path)
            
            # Open with xlwings
            app = xw.App(visible=False)
            try:
                wb = app.books.open(target_path)
                
                # Update data in each sheet
                for sheet_name, df in data_dict.items():
                    if sheet_name in [s.name for s in wb.sheets]:
                        sheet = wb.sheets[sheet_name]
                        
                        # Clear existing data (keep headers)
                        last_row = sheet.range('A1').end('down').row
                        if last_row > 1:
                            sheet.range(f'A2:Z{last_row}').clear_contents()
                            
                        # Write new data
                        if not df.empty:
                            sheet.range('A2').value = df.values
                            
                wb.save()
                wb.close()
                return True
                
            finally:
                app.quit()
                
        except Exception as e:
            print(f"Error in xlwings copy: {e}")
            return False
            
    @staticmethod
    def calculate_column_totals(df: pd.DataFrame, start_col: int, 
                               end_col: int) -> Dict[int, float]:
        """
        Calculate column totals for numeric columns
        
        Args:
            df: DataFrame
            start_col: Starting column index
            end_col: Ending column index
            
        Returns:
            Dictionary of column_index -> total
        """
        totals = {}
        
        for col_idx in range(start_col, min(end_col + 1, len(df.columns))):
            col_data = pd.to_numeric(df.iloc[:, col_idx], errors='coerce')
            total = col_data.sum()
            if not pd.isna(total) and total > 0:
                totals[col_idx] = total
                
        return totals