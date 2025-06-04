#!/usr/bin/env python3
"""
Debug script to test pandas-based attendance reading in refactored InvVzdProcessor
"""

import os
import sys
import pandas as pd
import openpyxl
from datetime import datetime

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.python.tools.inv_vzd_processor_refactored import InvVzdProcessor
from src.python.tools.inv_vzd_processor_original import InvVzdProcessor as OriginalProcessor


def inspect_excel_structure(file_path, sheet_name='zdroj-dochazka'):
    """Inspect Excel file structure to understand the data layout"""
    print(f"\n{'='*80}")
    print(f"INSPECTING EXCEL STRUCTURE: {os.path.basename(file_path)}")
    print(f"{'='*80}")
    
    # Using openpyxl for direct inspection
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")
        print(f"Max row: {sheet.max_row}, Max column: {sheet.max_column}")
        
        # Print first 15 rows to see structure
        print("\nFirst 15 rows (columns A-J):")
        print("-" * 80)
        for row in range(1, min(16, sheet.max_row + 1)):
            row_data = []
            for col in range(1, min(11, sheet.max_column + 1)):
                cell = sheet.cell(row=row, column=col)
                value = cell.value
                if value is not None:
                    # Truncate long values
                    str_val = str(value)[:20] + "..." if len(str(value)) > 20 else str(value)
                    row_data.append(f"{openpyxl.utils.get_column_letter(col)}: {str_val}")
            if row_data:
                print(f"Row {row}: {' | '.join(row_data)}")
    
    wb.close()
    
    # Now use pandas to show how it reads the data
    print(f"\n{'='*50}")
    print("PANDAS READING:")
    print(f"{'='*50}")
    
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    print(f"\nDataFrame shape: {df.shape}")
    print(f"\nFirst 15 rows with pandas (showing indices):")
    print("-" * 80)
    
    # Show with both pandas index and actual Excel row numbers
    for idx in range(min(15, len(df))):
        excel_row = idx + 2  # Excel row (1-based, accounting for header)
        print(f"Pandas idx {idx} (Excel row {excel_row}):")
        row_data = []
        for col_idx, col in enumerate(df.columns[:10]):
            value = df.iloc[idx, col_idx]
            if pd.notna(value):
                str_val = str(value)[:20] + "..." if len(str(value)) > 20 else str(value)
                row_data.append(f"Col{col_idx}: {str_val}")
        if row_data:
            print(f"  {' | '.join(row_data)}")


def test_refactored_reading(processor, file_path):
    """Test the refactored processor's data reading"""
    print(f"\n{'='*80}")
    print(f"TESTING REFACTORED PROCESSOR")
    print(f"{'='*80}")
    
    try:
        # Read the Excel file
        df = pd.read_excel(file_path, sheet_name='zdroj-dochazka')
        print(f"\nDataFrame loaded: {df.shape}")
        
        # Test date extraction
        print("\n--- Testing Date Extraction ---")
        dates = processor._extract_dates_16h(df)
        if dates:
            print(f"Found {len(dates)} dates:")
            for i, date in enumerate(dates):
                print(f"  Date {i}: {date}")
        else:
            print("No dates found!")
            print(f"Errors: {processor.errors}")
            
        # Test the actual reading process
        print("\n--- Testing Full Processing ---")
        processed_data = processor._process_16h_file(file_path, None)
        
        if processed_data and 'activities' in processed_data:
            activities_df = processed_data['activities']
            print(f"\nActivities found: {len(activities_df)}")
            print("\nFirst 5 activities:")
            print(activities_df.head())
            
            print(f"\nTotal hours: {processor.hours_total}")
        else:
            print("No data processed!")
            print(f"Errors: {processor.errors}")
            
        # Debug specific cells
        print("\n--- Debug Specific Cells ---")
        config = processor.config
        print(f"Config dates_row: {config['dates_row']} (0-based, so Excel row {config['dates_row'] + 1})")
        print(f"Config hours_row: {config['hours_row']} (0-based, so Excel row {config['hours_row'] + 1})")
        print(f"Config data_start_col: {config['data_start_col']} (0-based)")
        
        # Check specific cells
        for col in range(2, 7):  # Columns C-G
            print(f"\nColumn {col} (Excel column {openpyxl.utils.get_column_letter(col + 1)}):")
            if config['dates_row'] < len(df):
                date_val = df.iloc[config['dates_row'], col]
                print(f"  Date (row {config['dates_row']}): {repr(date_val)}")
            if config['hours_row'] < len(df):
                hours_val = df.iloc[config['hours_row'], col]
                print(f"  Hours (row {config['hours_row']}): {repr(hours_val)}")
                
    except Exception as e:
        print(f"Error in refactored reading: {e}")
        import traceback
        traceback.print_exc()


def test_original_reading(processor, file_path):
    """Test the original processor's data reading for comparison"""
    print(f"\n{'='*80}")
    print(f"TESTING ORIGINAL PROCESSOR")
    print(f"{'='*80}")
    
    try:
        # Test the reading
        df = processor._read_16_hour_data(file_path)
        
        if df is not None:
            print(f"\nData read successfully: {len(df)} activities")
            print("\nFirst 5 activities:")
            print(df.head())
            print(f"\nTotal hours: {processor.hours_total}")
        else:
            print("No data read!")
            print(f"Errors: {processor.errors}")
            
    except Exception as e:
        print(f"Error in original reading: {e}")
        import traceback
        traceback.print_exc()


def compare_processors(file_path):
    """Compare original and refactored processors"""
    print(f"\n{'='*80}")
    print(f"COMPARING PROCESSORS")
    print(f"{'='*80}")
    
    # Create processors
    original = OriginalProcessor(version="16")
    refactored = InvVzdProcessor(version="16")
    
    # Test original
    print("\n1. Original Processor:")
    orig_data = original._read_16_hour_data(file_path)
    if orig_data is not None:
        print(f"   - Activities: {len(orig_data)}")
        print(f"   - Total hours: {original.hours_total}")
        print(f"   - Columns: {list(orig_data.columns)}")
    else:
        print(f"   - Failed! Errors: {original.errors}")
    
    # Test refactored
    print("\n2. Refactored Processor:")
    ref_data = refactored._process_16h_file(file_path, None)
    if ref_data and 'activities' in ref_data:
        activities_df = ref_data['activities']
        print(f"   - Activities: {len(activities_df)}")
        print(f"   - Total hours: {refactored.hours_total}")
        print(f"   - Columns: {list(activities_df.columns)}")
    else:
        print(f"   - Failed! Errors: {refactored.errors}")
        
    # Compare results if both succeeded
    if orig_data is not None and ref_data and 'activities' in ref_data:
        print("\n3. Comparison:")
        activities_df = ref_data['activities']
        
        # Compare counts
        if len(orig_data) == len(activities_df):
            print("   ✓ Same number of activities")
        else:
            print(f"   ✗ Different number of activities: {len(orig_data)} vs {len(activities_df)}")
            
        # Compare total hours
        if original.hours_total == refactored.hours_total:
            print("   ✓ Same total hours")
        else:
            print(f"   ✗ Different total hours: {original.hours_total} vs {refactored.hours_total}")
            
        # Compare data content
        print("\n   Original first activity:")
        print(f"   {orig_data.iloc[0].to_dict()}")
        print("\n   Refactored first activity:")
        print(f"   {activities_df.iloc[0].to_dict()}")


def main():
    """Main test function"""
    # Test files
    test_files = [
        "tests/regression/inputs/16h/valid_basic.xlsx",
        "tests/regression/inputs/16h/valid_full.xlsx",
        "tests/regression/inputs/16h/error_dates.xlsx",
    ]
    
    for test_file in test_files:
        if os.path.exists(test_file):
            print(f"\n{'#'*80}")
            print(f"# TESTING FILE: {test_file}")
            print(f"{'#'*80}")
            
            # First inspect the structure
            inspect_excel_structure(test_file)
            
            # Create refactored processor
            processor = InvVzdProcessor(version="16")
            
            # Test refactored reading
            test_refactored_reading(processor, test_file)
            
            # Compare with original
            compare_processors(test_file)
            
            print("\n" + "="*80 + "\n")
        else:
            print(f"Test file not found: {test_file}")


if __name__ == "__main__":
    main()