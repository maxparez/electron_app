#!/usr/bin/env python3
"""Merge multiple innovative education attendance workbooks into one workbook.

This is a standalone helper for manual OP JAK data cleanup, not part of the
Electron UI workflow.

Default usage from the repository root:

    python scripts/merge_innovation_attendance.py

By default the helper:

- reads all non-temporary XLSX files from ``tmp/`` except the output workbook;
- uses ``tmp/dochazka-inovativni_vzdelavani_MIMO-VYUKU-dochazka_OPJAK_II_TK.xlsx``
  as the output/template workbook;
- merges all students from column B on ``zdroj-dochazka``;
- copies all activity columns from C onward and sorts them globally by date;
- preserves cached source row totals from column A;
- fills ``Třídní kniha`` column B with ``date start-end`` values, where one
  teaching hour is 45 minutes, and column D with the teacher.

Use ``--input-dir`` and ``--output`` when reusing the helper for another folder
or another output workbook.
"""

from __future__ import annotations

import argparse
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DEFAULT_OUTPUT = Path("tmp/dochazka-inovativni_vzdelavani_MIMO-VYUKU-dochazka_OPJAK_II_TK.xlsx")
ATTENDANCE_SHEET = "zdroj-dochazka"
CLASS_BOOK_SHEET = "Třídní kniha"
FIRST_ACTIVITY_COLUMN = 3  # C; 16h attendance files use C as the first activity column.
STUDENT_START_ROW = 12
CLASS_BOOK_START_ROW = 7
HEADER_ROWS = range(6, 12)


@dataclass(frozen=True)
class Student:
    source_index: int
    source_name: str
    source_row: int
    name: str
    hour_total: Any


@dataclass(frozen=True)
class Activity:
    source_index: int
    source_name: str
    source_column: int
    sort_date: date
    original_order: int
    header_values: dict[int, Any]
    attendance_by_source_row: dict[int, Any]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Merge innovative education attendance XLSX files from one folder."
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path("tmp"),
        help="Folder with source XLSX files. Defaults to tmp.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Output/template XLSX path. Defaults to {DEFAULT_OUTPUT}.",
    )
    return parser.parse_args()


def select_attendance_sheet(workbook):
    if ATTENDANCE_SHEET in workbook.sheetnames:
        return workbook[ATTENDANCE_SHEET]
    return workbook[workbook.sheetnames[0]]


def normalize_date(value: Any, source_name: str, column: int) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        stripped = value.strip()
        for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(stripped, fmt).date()
            except ValueError:
                pass
    column_letter = get_column_letter(column)
    raise ValueError(f"{source_name}: nelze přečíst datum v {column_letter}6: {value!r}")


def normalize_start_time(value: Any, source_name: str, column: int) -> time:
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    if isinstance(value, str):
        stripped = value.strip().replace(".", ":")
        if "-" in stripped or "–" in stripped or "—" in stripped:
            stripped = stripped.replace("–", "-").replace("—", "-").split("-", maxsplit=1)[0].strip()
        for fmt in ("%H:%M", "%H:%M:%S"):
            try:
                return datetime.strptime(stripped, fmt).time()
            except ValueError:
                pass
    column_letter = get_column_letter(column)
    raise ValueError(f"{source_name}: nelze přečíst čas začátku v {column_letter}7: {value!r}")


def normalize_hour_count(value: Any, source_name: str, column: int) -> float:
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError) as exc:
        column_letter = get_column_letter(column)
        raise ValueError(f"{source_name}: nelze přečíst počet hodin v {column_letter}11: {value!r}") from exc


def format_time(value: time) -> str:
    return f"{value.hour}:{value.minute:02d}"


def format_class_book_datetime(activity: Activity) -> str:
    start_time = normalize_start_time(activity.header_values[7], activity.source_name, activity.source_column)
    hour_count = normalize_hour_count(activity.header_values[11], activity.source_name, activity.source_column)
    starts_at = datetime.combine(activity.sort_date, start_time)
    ends_at = starts_at + timedelta(minutes=round(hour_count * 45))
    return (
        f"{activity.sort_date.day}.{activity.sort_date.month}.{activity.sort_date.year} "
        f"{format_time(starts_at.time())}-{format_time(ends_at.time())}"
    )


def has_activity(sheet, column: int) -> bool:
    hours = sheet.cell(row=11, column=column).value
    activity_date = sheet.cell(row=6, column=column).value
    return hours not in (None, "") and activity_date not in (None, "")


def read_students(sheet, values_sheet, source_index: int, source_name: str) -> list[Student]:
    students: list[Student] = []
    empty_count = 0
    max_row = sheet.max_row or 0
    for row in range(STUDENT_START_ROW, max_row + 1):
        value = sheet.cell(row=row, column=2).value
        if value is None or str(value).strip() == "":
            empty_count += 1
            if empty_count >= 2:
                break
            continue
        empty_count = 0
        students.append(
            Student(
                source_index=source_index,
                source_name=source_name,
                source_row=row,
                name=str(value).strip(),
                hour_total=values_sheet.cell(row=row, column=1).value,
            )
        )
    if not students:
        raise ValueError(f"{source_name}: ve sloupci B nebyli nalezeni žádní žáci")
    return students


def read_activities(sheet, source_index: int, source_name: str) -> list[Activity]:
    activities: list[Activity] = []
    max_column = sheet.max_column or 0
    for column in range(FIRST_ACTIVITY_COLUMN, max_column + 1):
        if not has_activity(sheet, column):
            continue
        header_values = {row: sheet.cell(row=row, column=column).value for row in HEADER_ROWS}
        attendance_by_source_row = {
            row: sheet.cell(row=row, column=column).value
            for row in range(STUDENT_START_ROW, (sheet.max_row or 0) + 1)
        }
        activities.append(
            Activity(
                source_index=source_index,
                source_name=source_name,
                source_column=column,
                sort_date=normalize_date(header_values[6], source_name, column),
                original_order=len(activities),
                header_values=header_values,
                attendance_by_source_row=attendance_by_source_row,
            )
        )
    if not activities:
        raise ValueError(f"{source_name}: nebyly nalezeny žádné aktivity od sloupce C dál")
    return activities


def find_source_files(input_dir: Path, output: Path) -> list[Path]:
    output = output.resolve()
    files = []
    for path in sorted(input_dir.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        if path.resolve() == output:
            continue
        files.append(path)
    if not files:
        raise ValueError(f"Ve složce {input_dir} nejsou žádné zdrojové XLSX soubory")
    return files


def copy_cell_format(source_cell, target_cell) -> None:
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)
    if source_cell.fill:
        target_cell.fill = copy(source_cell.fill)
    if source_cell.font:
        target_cell.font = copy(source_cell.font)
    if source_cell.border:
        target_cell.border = copy(source_cell.border)
    if source_cell.protection:
        target_cell.protection = copy(source_cell.protection)


def ensure_activity_capacity(sheet, activity_count: int) -> None:
    last_needed_column = FIRST_ACTIVITY_COLUMN + activity_count - 1
    template_column = min(sheet.max_column or last_needed_column, last_needed_column)
    for column in range((sheet.max_column or 0) + 1, last_needed_column + 1):
        for row in range(1, (sheet.max_row or STUDENT_START_ROW) + 1):
            copy_cell_format(sheet.cell(row=row, column=template_column), sheet.cell(row=row, column=column))
        sheet.column_dimensions[get_column_letter(column)].width = sheet.column_dimensions[
            get_column_letter(template_column)
        ].width


def clear_target(sheet, activity_count: int) -> None:
    max_clear_column = max(sheet.max_column or 0, FIRST_ACTIVITY_COLUMN + activity_count - 1)
    max_clear_row = max(sheet.max_row or 0, STUDENT_START_ROW + 120)
    for row in HEADER_ROWS:
        for column in range(FIRST_ACTIVITY_COLUMN, max_clear_column + 1):
            sheet.cell(row=row, column=column).value = None
    for row in range(STUDENT_START_ROW, max_clear_row + 1):
        sheet.cell(row=row, column=1).value = None
        sheet.cell(row=row, column=2).value = None
        for column in range(FIRST_ACTIVITY_COLUMN, max_clear_column + 1):
            sheet.cell(row=row, column=column).value = None


def write_output(sheet, students: list[Student], activities: list[Activity]) -> None:
    ensure_activity_capacity(sheet, len(activities))
    clear_target(sheet, len(activities))

    last_activity_column = FIRST_ACTIVITY_COLUMN + len(activities) - 1
    last_activity_letter = get_column_letter(last_activity_column)
    sheet.cell(row=11, column=1).value = f"=SUM(C11:{last_activity_letter}11)"

    output_row_by_student = {
        (student.source_index, student.source_row): STUDENT_START_ROW + index
        for index, student in enumerate(students)
    }

    for index, student in enumerate(students):
        output_row = STUDENT_START_ROW + index
        sheet.cell(row=output_row, column=1).value = student.hour_total
        sheet.cell(row=output_row, column=2).value = student.name

    for activity_index, activity in enumerate(activities):
        output_column = FIRST_ACTIVITY_COLUMN + activity_index
        for row, value in activity.header_values.items():
            sheet.cell(row=row, column=output_column).value = value

        for source_row, attendance in activity.attendance_by_source_row.items():
            output_row = output_row_by_student.get((activity.source_index, source_row))
            if output_row is None:
                continue
            sheet.cell(row=output_row, column=output_column).value = attendance


def clear_class_book(class_book) -> None:
    for row in range(CLASS_BOOK_START_ROW, (class_book.max_row or CLASS_BOOK_START_ROW) + 1):
        class_book.cell(row=row, column=2).value = None
        class_book.cell(row=row, column=4).value = None


def write_class_book(class_book, activities: list[Activity]) -> None:
    clear_class_book(class_book)
    for index, activity in enumerate(activities):
        row = CLASS_BOOK_START_ROW + index
        class_book.cell(row=row, column=2).value = format_class_book_datetime(activity)
        class_book.cell(row=row, column=4).value = activity.header_values[10]


def merge_workbooks(input_dir: Path, output: Path) -> dict[str, Any]:
    source_files = find_source_files(input_dir, output)
    all_students: list[Student] = []
    all_activities: list[Activity] = []

    for source_index, source_file in enumerate(source_files):
        workbook = load_workbook(source_file, data_only=False)
        values_workbook = load_workbook(source_file, data_only=True)
        try:
            sheet = select_attendance_sheet(workbook)
            values_sheet = select_attendance_sheet(values_workbook)
            all_students.extend(read_students(sheet, values_sheet, source_index, source_file.name))
            all_activities.extend(read_activities(sheet, source_index, source_file.name))
        finally:
            workbook.close()
            values_workbook.close()

    all_activities.sort(key=lambda item: (item.sort_date, item.source_index, item.source_column))

    output_workbook = load_workbook(output)
    try:
        output_sheet = select_attendance_sheet(output_workbook)
        write_output(output_sheet, all_students, all_activities)
        if CLASS_BOOK_SHEET in output_workbook.sheetnames:
            write_class_book(output_workbook[CLASS_BOOK_SHEET], all_activities)
        output_workbook.save(output)
    finally:
        output_workbook.close()

    return {
        "sources": [path.name for path in source_files],
        "students": len(all_students),
        "activities": len(all_activities),
        "first_activity_date": all_activities[0].sort_date.isoformat(),
        "last_activity_date": all_activities[-1].sort_date.isoformat(),
        "output": str(output),
    }


def main() -> None:
    args = parse_args()
    report = merge_workbooks(args.input_dir, args.output)
    print("Sloučení dokončeno")
    print(f"Zdroje: {len(report['sources'])}")
    for source in report["sources"]:
        print(f"  - {source}")
    print(f"Žáci: {report['students']}")
    print(f"Aktivity: {report['activities']}")
    print(f"Rozsah dat: {report['first_activity_date']} až {report['last_activity_date']}")
    print(f"Výstup: {report['output']}")


if __name__ == "__main__":
    main()
