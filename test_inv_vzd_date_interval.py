#!/usr/bin/env python3
"""
Focused checks for 16h innovative education date interval warnings.
"""

import os
import sys
import tempfile
from datetime import datetime

from openpyxl import Workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "src", "python"))

from tools.inv_vzd_processor import InvVzdProcessor


def create_16h_source(path, dates):
    wb = Workbook()
    ws = wb.active
    ws.title = "zdroj-dochazka"
    ws["B6"] = "datum aktivity"
    ws["B7"] = "čas zahájení"

    for offset, activity_date in enumerate(dates):
        col = 3 + offset
        ws.cell(row=6, column=col).value = activity_date
        ws.cell(row=7, column=col).value = "08:00"
        ws.cell(row=8, column=col).value = "Prezenční"
        ws.cell(row=9, column=col).value = "Téma"
        ws.cell(row=10, column=col).value = "Pedagog"
        ws.cell(row=11, column=col).value = 1

    wb.save(path)


def test_16h_date_interval_warnings():
    with tempfile.TemporaryDirectory() as tmp_dir:
        valid_path = os.path.join(tmp_dir, "valid_16h.xlsx")
        invalid_path = os.path.join(tmp_dir, "invalid_16h.xlsx")

        create_16h_source(valid_path, [
            datetime(2025, 8, 25),
            datetime(2026, 1, 14),
        ])
        create_16h_source(invalid_path, [
            datetime(2025, 8, 25),
            datetime(2026, 1, 14),
            datetime(2026, 2, 14),
        ])

        valid_processor = InvVzdProcessor("16")
        valid_data = valid_processor._read_16_hour_data(valid_path)
        assert valid_data is not None
        assert not any("mimo povolený interval" in warning for warning in valid_processor.warnings)

        invalid_processor = InvVzdProcessor("16")
        invalid_data = invalid_processor._read_16_hour_data(invalid_path)
        assert invalid_data is not None
        assert any(
            "Datum 14.02.2026 je mimo povolený interval 01.08.2025 - 31.01.2026" in warning
            for warning in invalid_processor.warnings
        ), invalid_processor.warnings


def test_16h_date_interval_warning_is_in_file_result():
    with tempfile.TemporaryDirectory() as tmp_dir:
        source_path = os.path.join(tmp_dir, "invalid_16h.xlsx")
        template_path = os.path.join(tmp_dir, "template_16h.xlsx")

        create_16h_source(source_path, [
            datetime(2025, 8, 25),
            datetime(2026, 1, 14),
            datetime(2026, 2, 14),
        ])

        template = Workbook()
        template.active["B1"] = "16 hodin"
        template.save(template_path)

        processor = InvVzdProcessor("16")
        processor._copy_template_with_data = lambda *args: None

        result = processor.process([source_path], {
            "template": template_path,
            "output_dir": tmp_dir,
        })

        processed_file = result["data"]["processed_files"][0]
        assert processed_file["status"] == "success"
        assert any(
            "Datum 14.02.2026 je mimo povolený interval 01.08.2025 - 31.01.2026" in warning
            for warning in processed_file["warnings"]
        ), processed_file


if __name__ == "__main__":
    test_16h_date_interval_warnings()
    test_16h_date_interval_warning_is_in_file_result()
    print("OK")
