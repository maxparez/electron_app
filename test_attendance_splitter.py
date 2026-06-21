#!/usr/bin/env python3

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

REPO_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(REPO_ROOT / "src" / "python"))

from tools.attendance_splitter import AttendanceSplitter


def add_16h_attendance_sheet(workbook: Workbook, title: str, *, hidden: bool = False):
    sheet = workbook.create_sheet(title)
    sheet["B6"] = "datum aktivity"
    sheet["B7"] = "čas zahájení"
    sheet["B10"] = "Jméno pedag. pracovníka"
    sheet["B11"] = "počet hodin"
    if hidden:
        sheet.sheet_state = "hidden"
    return sheet


def add_32h_attendance_sheet(workbook: Workbook, title: str):
    sheet = workbook.create_sheet(title)
    sheet["B6"] = "datum aktivity"
    sheet["B7"] = "Forma výuky"
    sheet["B9"] = "Jméno pedag. pracovníka"
    sheet["B10"] = "počet hodin"
    return sheet


def build_workbook(path: Path, sheet_builder) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheet_builder(workbook)
    workbook.save(path)


class AttendanceSplitterInspectionTests(unittest.TestCase):
    def test_inspect_workbook_detects_visible_attendance_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "combined.xlsx"

            def create_sheets(workbook):
                add_16h_attendance_sheet(workbook, "1. ročník")
                add_32h_attendance_sheet(workbook, "2. a 3. ročník")
                add_16h_attendance_sheet(workbook, "Skrytá docházka", hidden=True)
                data_sheet = workbook.create_sheet("Data")
                data_sheet.sheet_state = "hidden"

            build_workbook(workbook_path, create_sheets)

            result = AttendanceSplitter().inspect_workbook(workbook_path)

            self.assertTrue(result["eligible"])
            self.assertEqual(["1. ročník", "2. a 3. ročník"], result["attendance_sheets"])
            self.assertEqual(
                [
                    {"sheet_name": "Skrytá docházka", "reason": "Skrytý list"},
                    {"sheet_name": "Data", "reason": "Skrytý list"},
                ],
                result["skipped_sheets"],
            )

    def test_inspect_workbook_requires_at_least_two_attendance_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "single.xlsx"

            def create_sheets(workbook):
                add_16h_attendance_sheet(workbook, "1. ročník")
                workbook.create_sheet("Poznámky")

            build_workbook(workbook_path, create_sheets)

            result = AttendanceSplitter().inspect_workbook(workbook_path)

            self.assertFalse(result["eligible"])
            self.assertEqual(["1. ročník"], result["attendance_sheets"])
            self.assertEqual(
                [{"sheet_name": "Poznámky", "reason": "List neodpovídá formátu docházky"}],
                result["skipped_sheets"],
            )

    def test_normalize_sheet_name_removes_diacritics_and_unsafe_characters(self) -> None:
        splitter = AttendanceSplitter()

        normalized = splitter.normalize_sheet_name(" 2. a 3. ročník / Žáci ")

        self.assertEqual("2_a_3_rocnik_zaci", normalized)

    def test_build_output_path_does_not_overwrite_existing_files(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            (output_dir / "dochazka_inovace_1_rocnik.xlsx").touch()
            (output_dir / "dochazka_inovace_1_rocnik_2.xlsx").touch()

            output_path = AttendanceSplitter().build_output_path(output_dir, "1. ročník")

            self.assertEqual(
                output_dir / "dochazka_inovace_1_rocnik_3.xlsx",
                output_path,
            )

    def test_scan_folder_returns_only_eligible_xlsx_workbooks(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            folder = Path(temp_dir)

            build_workbook(
                folder / "combined.xlsx",
                lambda workbook: (
                    add_16h_attendance_sheet(workbook, "1. ročník"),
                    add_16h_attendance_sheet(workbook, "2. ročník"),
                ),
            )
            build_workbook(
                folder / "single.xlsx",
                lambda workbook: add_16h_attendance_sheet(workbook, "1. ročník"),
            )
            (folder / "~$combined.xlsx").touch()
            (folder / "notes.txt").write_text("not an excel file", encoding="utf-8")

            matches = AttendanceSplitter().scan_folder(folder)

            self.assertEqual(1, len(matches))
            self.assertEqual("combined.xlsx", matches[0]["name"])
            self.assertEqual(["1. ročník", "2. ročník"], matches[0]["attendance_sheets"])


class AttendanceSplitterProcessingTests(unittest.TestCase):
    def test_process_creates_outputs_in_dedicated_subfolder(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "combined.xlsx"

            def create_sheets(workbook):
                add_16h_attendance_sheet(workbook, "1. ročník")
                add_16h_attendance_sheet(workbook, "2. a 3. ročník")
                data_sheet = workbook.create_sheet("Data")
                data_sheet.sheet_state = "hidden"

            build_workbook(source_path, create_sheets)
            copy_calls = []

            def fake_copier(source, sheet_name, output_path):
                copy_calls.append((Path(source), sheet_name, Path(output_path)))
                Path(output_path).touch()

            result = AttendanceSplitter(sheet_copier=fake_copier).process([source_path])

            output_dir = Path(temp_dir) / "rozdelene_dochazky"
            self.assertTrue(result["success"])
            self.assertEqual("success", result["status"])
            self.assertEqual(2, result["data"]["created_count"])
            self.assertEqual(
                [
                    (source_path, "1. ročník", output_dir / "dochazka_inovace_1_rocnik.xlsx"),
                    (
                        source_path,
                        "2. a 3. ročník",
                        output_dir / "dochazka_inovace_2_a_3_rocnik.xlsx",
                    ),
                ],
                copy_calls,
            )
            file_result = result["data"]["files"][0]
            self.assertEqual("success", file_result["status"])
            self.assertEqual(
                [{"sheet_name": "Data", "reason": "Skrytý list"}],
                file_result["skipped_sheets"],
            )

    def test_process_continues_after_one_sheet_copy_fails(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "combined.xlsx"
            build_workbook(
                source_path,
                lambda workbook: (
                    add_16h_attendance_sheet(workbook, "1. ročník"),
                    add_16h_attendance_sheet(workbook, "2. ročník"),
                ),
            )

            def flaky_copier(source, sheet_name, output_path):
                if sheet_name == "1. ročník":
                    raise RuntimeError("Excel copy failed")
                Path(output_path).touch()

            result = AttendanceSplitter(sheet_copier=flaky_copier).process([source_path])

            self.assertTrue(result["success"])
            self.assertEqual("partial", result["status"])
            self.assertEqual(1, result["data"]["created_count"])
            self.assertEqual(1, result["data"]["failed_count"])
            file_result = result["data"]["files"][0]
            self.assertEqual("partial", file_result["status"])
            self.assertIn("1. ročník: Excel copy failed", file_result["errors"])
            self.assertEqual("2. ročník", file_result["created_files"][0]["sheet_name"])

    def test_process_reports_ineligible_workbook_without_copying(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "single.xlsx"
            build_workbook(
                source_path,
                lambda workbook: add_16h_attendance_sheet(workbook, "1. ročník"),
            )

            result = AttendanceSplitter(sheet_copier=lambda *_: None).process([source_path])

            self.assertFalse(result["success"])
            self.assertEqual("error", result["status"])
            self.assertEqual(0, result["data"]["created_count"])
            self.assertEqual("error", result["data"]["files"][0]["status"])
            self.assertIn(
                "Soubor neobsahuje alespoň dva vhodné listy docházky",
                result["data"]["files"][0]["errors"],
            )

    def test_process_reports_when_existing_output_forces_renamed_file(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "combined.xlsx"
            build_workbook(
                source_path,
                lambda workbook: (
                    add_16h_attendance_sheet(workbook, "1. ročník"),
                    add_16h_attendance_sheet(workbook, "2. ročník"),
                ),
            )
            output_dir = Path(temp_dir) / "rozdelene_dochazky"
            output_dir.mkdir()
            (output_dir / "dochazka_inovace_1_rocnik.xlsx").touch()

            def fake_copier(source, sheet_name, output_path):
                Path(output_path).touch()

            result = AttendanceSplitter(sheet_copier=fake_copier).process([source_path])

            created_file = result["data"]["files"][0]["created_files"][0]
            self.assertTrue(created_file["renamed_to_avoid_overwrite"])
            self.assertEqual(
                "dochazka_inovace_1_rocnik.xlsx",
                created_file["requested_filename"],
            )
            self.assertEqual(
                "dochazka_inovace_1_rocnik_2.xlsx",
                created_file["filename"],
            )


if __name__ == "__main__":
    unittest.main()
