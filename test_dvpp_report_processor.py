#!/usr/bin/env python3

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parent / "src" / "python"))

from tools.dvpp_report_processor import DvppReportProcessor


def build_dvpp_workbook(path: Path, *, report_number: int, rows: list[tuple[str, str, float]]) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    workbook.create_sheet("postup vyplňování")
    sheet = workbook.create_sheet("podpory")
    workbook.create_sheet("data")

    sheet["B6"] = "        Registrační číslo projektu"
    sheet["D6"] = "CZ.TEST/00/00/00/00000"
    sheet["H6"] = "Zpráva o realizaci č. "
    sheet["I6"] = report_number
    sheet["B10"] = "Příjmení"
    sheet["C10"] = "Jméno"
    sheet["D10"] = "Datum narození"
    sheet["E10"] = "Šablona"
    sheet["F10"] = 'Téma\nZvolíte-li ve sloupci „Forma“ kteroukoliv z variant „mentoring“, „supervize“ „koučink“ nebo „kvalifikační studium DVPP“ nechte pole Téma prázdné.'
    sheet["G10"] = "Počet  hodin podpory"

    row_index = 11
    for last_name, first_name, hours in rows:
        sheet.cell(row_index, 2).value = last_name
        sheet.cell(row_index, 3).value = first_name
        sheet.cell(row_index, 5).value = "DVPP/1"
        sheet.cell(row_index, 6).value = "Formativní hodnocení"
        sheet.cell(row_index, 7).value = hours
        row_index += 1

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def build_dvpp_candidate_workbook(path: Path) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    default_sheet.title = "postup vyplňování"
    workbook.create_sheet("podpory")
    workbook.create_sheet("data")

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


class DvppReportProcessorTests(unittest.TestCase):
    def test_scan_project_directory_returns_matching_workbooks(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            project_dir = Path(temp_dir) / "Project_12968"
            build_dvpp_workbook(
                project_dir / "1ZoR" / "first.xlsx",
                report_number=1,
                rows=[("Novák", "Jan", 8), ("Svobodová", "Eva", 16)],
            )
            build_dvpp_workbook(
                project_dir / "2ZoR" / "nested" / "second.xlsx",
                report_number=2,
                rows=[("Novák", "Jan", 12)],
            )

            processor = DvppReportProcessor()
            matches = processor.scan_project_directory(str(project_dir))

            self.assertEqual(2, len(matches))
            self.assertEqual("1ZoR/first.xlsx", matches[0]["relative_path"])
            self.assertEqual("2ZoR/nested/second.xlsx", matches[1]["relative_path"])
            self.assertEqual(1, matches[0]["report_number"])
            self.assertEqual(2, matches[1]["report_number"])
            self.assertEqual(2, matches[0]["participant_count"])

    def test_scan_project_directory_lists_candidates_with_required_sheets_only(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            project_dir = Path(temp_dir) / "Project_12968"
            build_dvpp_candidate_workbook(project_dir / "1ZoR" / "candidate.xlsx")

            processor = DvppReportProcessor()
            matches = processor.scan_project_directory(str(project_dir))

            self.assertEqual(1, len(matches))
            self.assertEqual("1ZoR/candidate.xlsx", matches[0]["relative_path"])
            self.assertEqual("podpory", matches[0]["sheet_name"])
            self.assertEqual(0, matches[0]["participant_count"])

    def test_process_generates_html_report_in_project_directory(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            project_dir = Path(temp_dir) / "Project_12968"
            build_dvpp_workbook(
                project_dir / "1ZoR" / "first.xlsx",
                report_number=1,
                rows=[("Novák", "Jan", 8), ("Svobodová", "Eva", 16)],
            )
            build_dvpp_workbook(
                project_dir / "2ZoR" / "second.xlsx",
                report_number=2,
                rows=[("Novák", "Jan", 12), ("Svobodová", "Eva", 8)],
            )

            processor = DvppReportProcessor()
            matches = processor.scan_project_directory(str(project_dir))
            selected_files = [match["file_path"] for match in matches]

            result = processor.process(selected_files, {"project_dir": str(project_dir)})

            self.assertTrue(result["success"])
            self.assertEqual(2, result["data"]["files_processed"])
            self.assertEqual(2, result["data"]["unique_participants"])
            self.assertEqual("Project_12968_dvpp_report.html", result["data"]["report_filename"])

            report_path = Path(result["data"]["report_path"])
            self.assertEqual(project_dir / "Project_12968_dvpp_report.html", report_path)
            self.assertTrue(report_path.exists())

            html = report_path.read_text(encoding="utf-8")
            self.assertIn("Souhrnný report podpory DVPP - Project_12968", html)
            self.assertIn("Témata dle Šablon", html)
            self.assertIn("Formativní hodnocení", html)
            self.assertIn("Jan Novák", html)
            self.assertIn("Eva Svobodová", html)
            self.assertIn("Souhrn podpory DVPP podle pedagogů", html)


if __name__ == "__main__":
    unittest.main()
